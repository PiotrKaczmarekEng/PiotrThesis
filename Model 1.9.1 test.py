# -*- coding: utf-8 -*-
"""
Created on Thu May  2 11:41:52 2024

@author: Piotr Kaczmarek
"""

#%% --- Preamble ---

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import geopy.distance
import math
from openpyxl import load_workbook
from feedinlib import era5
import pvlib
from feedinlib import Photovoltaic
from feedinlib import get_power_plant_data
from windpowerlib.modelchain import ModelChain
from windpowerlib.wind_turbine import WindTurbine
from windpowerlib import get_turbine_types
from datetime import timedelta
from gurobipy import *
import plotly.express as px
import plotly.io as pio
import plotly.graph_objects as go
from sklearn.linear_model import LinearRegression
import statistics
import xarray as xr


import warnings

def fxn():
    warnings.warn("deprecated", DeprecationWarning)

with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    fxn()

# # Or if you are using > Python 3.11:
# with warnings.catch_warnings(action="ignore"):
#     fxn()
pd.set_option('future.no_silent_downcasting', True)


#%% --- Excel Parameters ---

#loading excel to later retrieve input data using openpyxl
data_file = os.getcwd() 
data_file = os.path.dirname(os.path.realpath('__file__')) + '\Inputdata econ.xlsx'
wb = load_workbook(data_file,data_only=True) # creating workbook

# general = wb['General'] #selecting data from sheet called general
# wind = wb['Wind'] #selecting data from sheet called wind
solar = wb['Solar'] #selecting data from sheet called solar
electrolyzer = wb['Electrolyzer'] #selecting data from sheet called electrolyzer
desalination = wb['Desalination'] #selecting data from sheet called desalination
ammonia = wb['Ammonia'] #selecting data from sheet called ammonia
liquidhydrogen = wb['Liquid hydrogen'] #selecting data from sheet called liquid hydrogen
landtransport = wb['Land transport'] #selecting data from sheet called landtransport
storage = wb['Storage'] #selecting data from sheet called storage
fpso = wb['FPSO'] #selecting data from sheet called fpso

# Time period data
Startyear = 2020 #first year to be reviewed
timeperiod = 30 #total time period of simulation in years
timestep = 3 #time step of simulation in years
Nsteps = int(timeperiod/timestep+1) #number of time steps (important for data selection from excel and loop at the end)

# Coordinates production location

## North-Sea case 
# Production central coordinates
latitude = 54.35
longitude = 6.28

#  Eemshaven (Groningen)
coords_demand = (53.43, 6.84) 
coords_port = (53.43, 6.84) 
# #  Port of Rotterdam
# coords_demand = (51.95, 4.14) 

# Era 5 weather data file location
era5_netcdf_filename = 'Era 5 test data\ERA5_weather_data_NorthSea_010120-300920.nc' #North-Sea Case data 9 months in 2020


# Solar and Wind 
multsolar = 524*1000/206.7989   # Installed capacity per unit [kWp] * 1000 (to make it [W]) / capacity from PVlib [Wp]
multwind = 12/3   # Capacity per unit [MW] / capacity from Windlib [MW]

# Set to the time period ratio of the CDS dataset considered (1/12 means 1 month)
DataYearRatio = 9/12

# Discount rate
DiscountRate = 0.08

# Demand parameter (needed for now for transport calculation)
capacity = 700 # [MW] theoretical capacity of north sea case
demand = capacity*8760/1000/0.0505 # [tonH2/yr] theoretical demand at 100% CF

# Energy Medium (0: NH3 ship, 1: LH2 ship, 2: GH2 pipe, 3: NH3 pipe)
E = 2

# Excel parameter replacement from TC
distanceseafactor = 1


# Technical parameters
# capelectrolyzerhour = electrolyzer.cell(row=9, column=2).value #hourly output capacity of electrolyzers in tons of hydrogen per hour
# alpha_j
beta = 0.4       # [tonH2]
gamma = 50500000 # [Wh/tonH2]
# epsilon
# eta_ij
# phi_j
# nu_j

# Excel parameter replacement from model parameters


#%% Wind Parameters - Excel replacement

#### Input params

learning_rate_wind = 0.088       # [%/100]
reduction_factor = np.log2(1 - learning_rate_wind)
Capacity_Wind = 12          # [MW]
CAPEX_wind_initial = 1944          # [Eur/kW] in 2020
CAPEX_wind_initial_MW = CAPEX_wind_initial * 1000 # [Eur/MW]
OPEX_wind_initial = 64 # OPEX [Eur/kW] in 2020
OPEX_wind_initial_MW = OPEX_wind_initial*1000 # OPEX [Eur/MW]
lftm = 25 # lifetime of turbine [Years]
a_wind = (DiscountRate*(1+DiscountRate)**lftm) / ((1+DiscountRate)**lftm - 1) # amortization factor

#### Computation

years = np.arange(2020, 2051)  # Years from 2020 to 2050

annual_increase = (250000 - 50000) / (2031 - 2021) # [GW/y] global cumulative capacity increase per year
cumulative_capacity = 50000 + (years - 2021) * annual_increase

CAPEX_wind_year = CAPEX_wind_initial_MW * (cumulative_capacity / 50000) ** reduction_factor
CAPEX_wind_year_list = list(zip(years, CAPEX_wind_year*Capacity_Wind))

OPEX_wind_year = OPEX_wind_initial_MW * (cumulative_capacity / 50000) ** reduction_factor
OPEX_wind_year_list = list(zip(years, OPEX_wind_year*Capacity_Wind))



Wind_Costs = np.zeros((5,31))

yearstep = 0
for i in range(31):
    Wind_Costs[0][i] = 2020 + yearstep
    yearstep += 1
    
for i in range(31):
    Wind_Costs[1][i] = CAPEX_wind_year_list[i][1]

for i in range(31):
    Wind_Costs[2][i] = lftm
    
for i in range(31):
    Wind_Costs[3][i] = OPEX_wind_year_list[i][1]

Wind_Costs[4] = a_wind*Wind_Costs[1] + Wind_Costs[3]

def rel_wind_array(SY, NS, TS, WiCo):
    index = np.where(Wind_Costs[0] == SY)[0][0]
    relevant_array = np.zeros(NS)
    for i in range(NS):    
        relevant_array[i] = WiCo[4][index+i*TS]
    
    return relevant_array

#%% Ammonia Conversion and Reconversion Parameters - Excel Replacement

# Tycho's numbers
r = 0.08 # Interest rate [%/100]
lftm = 25 #lifetime [years]
a = (DiscountRate*(1+DiscountRate)**lftm) / ((1+DiscountRate)**lftm - 1) # amortization factor

ThroughputConv = 100000/365 #tonNH3/day
ThroughputReconv = 1200 #tonNH3/day
TotalDirectCostConv = 80 #million euro
TotalDirectCostReconv = 63 #million euro
FactorConv = (TotalDirectCostConv/TotalDirectCostReconv)*(ThroughputReconv/ThroughputConv)

# Reconv
CAPEX_1 = 94500000 # euro
OPEX_1 = 1600000 # euro/yr
Capacity_reconv = 78840 # [tonH2/yr]
Capacity_used_device = 10000

Rec_Costs = np.zeros((5,31))

yearstep = 0

for i in range(31):
    Rec_Costs[0][i] = 2020 + yearstep # Year
    Rec_Costs[1][i] = CAPEX_1*Capacity_used_device/Capacity_reconv # CAPEX
    Rec_Costs[2][i] = lftm # Lifetime
    Rec_Costs[3][i] = OPEX_1*Capacity_used_device/Capacity_reconv # OPEX
    Rec_Costs[4][i] = Rec_Costs[1][i]*a+Rec_Costs[3][i] # Total = CAPEX*a + OPEX
    yearstep += 1
    
# Conv

Con_Costs = np.zeros((5,31))

yearstep = 0

for i in range(31):
    Con_Costs[0][i] = 2020 + yearstep # Year
    Con_Costs[1][i] = FactorConv*Rec_Costs[1][i] # CAPEX
    Con_Costs[2][i] = lftm # Lifetime
    Con_Costs[3][i] = FactorConv*Rec_Costs[3][i] # OPEX
    Con_Costs[4][i] = Con_Costs[1][i]*a+Con_Costs[3][i] # Total = CAPEX*a + OPEX
    yearstep += 1

def rel_con_array(SY, NS, TS, ConNH3Co):
    index = np.where(Con_Costs[0] == SY)[0][0]
    relevant_array_con_NH3 = np.zeros(NS)
    for i in range(NS):    
        relevant_array_con_NH3[i] = ConNH3Co[4][index+i*TS]
    
    return relevant_array_con_NH3

def rel_rec_array(SY, NS, TS, RecNH3Co):
    index = np.where(Rec_Costs[0] == SY)[0][0]
    relevant_array_rec_NH3 = np.zeros(NS)
    for i in range(NS):    
        relevant_array_rec_NH3[i] = RecNH3Co[4][index+i*TS]
    
    return relevant_array_rec_NH3


#%% LH2 Conversion and Reconversion Parameters - Excel Replacement

# Conversion
# Tycho's numbers
lftm = 30   # Conversion device lifetime
a = (DiscountRate*(1+DiscountRate)**lftm) / ((1+DiscountRate)**lftm - 1) # amortization factor
Prod_conv_unit = 10000 # [tonsH2/yr] yearly production of 1 conversion unit

# Eur/ton/yr
LH2_data = np.array([
    [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044, 2045, 2046, 2047, 2048, 2049, 2050],
    [5623.93, 5567.69, 5511.45, 5455.21, 5398.97, 5342.74, 5286.50, 5230.26, 5174.02, 5117.78, 5061.54, 5010.92, 4960.31, 4909.69, 4859.08, 4808.46, 4757.85, 4707.23, 4656.62, 4606.00, 4555.38, 4504.77, 4454.15, 4403.54, 4352.92, 4302.31, 4251.69, 4201.08, 4150.46, 4099.85, 4049.23],
    [224.96, 222.71, 220.46, 218.21, 215.96, 213.71, 211.46, 209.21, 206.96, 204.71, 202.46, 200.44, 198.41, 196.39, 194.36, 192.34, 190.31, 188.29, 186.26, 184.24, 182.22, 180.19, 178.17, 176.14, 174.12, 172.09, 170.07, 168.04, 166.02, 163.99, 161.97]
])

LH2_Con_Costs = np.zeros((5,31))
LH2_Con_Costs[0] = LH2_data[0]                   #Year
LH2_Con_Costs[1] = LH2_data[1]*Prod_conv_unit    #CAPEX [Euro/yr]
LH2_Con_Costs[2] = lftm                          #Lifetime [Years]
LH2_Con_Costs[3] = LH2_data[2]*Prod_conv_unit    #OPEX [Euro/yr]
LH2_Con_Costs[4] = LH2_Con_Costs[1]*a+LH2_Con_Costs[3]             #Total yearly [Euro/yr]


# Reconversion
FactorConvLH2 = 0.2

LH2_Rec_Costs = np.zeros((5,31))

yearstep = 0
for i in range(31):
    LH2_Rec_Costs[0][i] = LH2_Con_Costs[0][i]                   #Year
    LH2_Rec_Costs[1][i] = FactorConvLH2*LH2_Con_Costs[1][i]     #CAPEX [Euro/yr]
    LH2_Rec_Costs[2][i] = lftm                                  #Lifetime [years]
    LH2_Rec_Costs[3][i] = FactorConvLH2*LH2_Con_Costs[3][i]     #OPEX [Euro/yr]
    LH2_Rec_Costs[4][i] = FactorConvLH2*LH2_Con_Costs[4][i]     #Total yearly [Euro/yr]
    yearstep += 1
    
# LH2_Con_Costs == LH2_Rec_Costs

def rel_con_arrayLH2(SY, NS, TS, ConLH2Co):
    index = np.where(LH2_Con_Costs[0] == SY)[0][0]
    relevant_array_con_LH2 = np.zeros(NS)
    for i in range(NS):    
        relevant_array_con_LH2[i] = ConLH2Co[4][index+i*TS]
    
    return relevant_array_con_LH2

def rel_rec_arrayLH2(SY, NS, TS, RecLH2Co):
    index = np.where(LH2_Rec_Costs[0] == SY)[0][0]
    relevant_array_rec_LH2 = np.zeros(NS)
    for i in range(NS):    
        relevant_array_rec_LH2[i] = RecLH2Co[4][index+i*TS]
    
    return relevant_array_rec_LH2







#%% --- Solar Function ---


# area = [longitude, latitude]
    

def func_PV(location):
    
    
    lat_location = location[1]
    lon_location = location[0]
    # get modules
    module_df = get_power_plant_data(dataset='SandiaMod') #retrieving dataset for PV modules
    
    # get inverter data
    inverter_df = get_power_plant_data(dataset='cecinverter') #retrieving dataset for inverters
    
    temp_params = pvlib.temperature.TEMPERATURE_MODEL_PARAMETERS['sapm']['open_rack_glass_polymer'] #defining temperature model parameters (leaving at default causes errors)
    
    #PV system definition
    system_data = {
        'module_name': 'Advent_Solar_Ventura_210___2008_',  # module name as in database
        'inverter_name': 'ABB__MICRO_0_25_I_OUTD_US_208__208V_',  # inverter name as in database
        'azimuth': 180, #angle of sun position with north in horizontal plane
        'tilt': 30, #angle of solar panels with horizontal
        'albedo': 0.075, #albedo (fraction of sun light reflected) of ocean water (https://geoengineering.global/ocean-albedo-modification/)
        'temperature_model_parameters': temp_params,
    }
    
    pv_system = Photovoltaic(**system_data)
    
    #getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
    # area = [lon, lat]
    pvlib_df = era5.weather_df_from_era5(     
        era5_netcdf_filename=era5_netcdf_filename,
        lib='pvlib', area=[lon_location, lat_location])
    
    #determining the zenith angle (angle of sun position with vertical in vertical plane) in the specified locations for the time instances downloaded from ERA5
    zenithcalc = pvlib.solarposition.get_solarposition(time=pvlib_df.index,latitude=latitude, longitude=longitude, altitude=None, pressure=None, method='nrel_numpy', temperature=pvlib_df['temp_air'])
    
    #determining DNI from GHI, DHI and zenith angle for the time instances downloaded from ERA5
    dni = pvlib.irradiance.dni(pvlib_df['ghi'],pvlib_df['dhi'], zenith=zenithcalc['zenith'], clearsky_dni=None, clearsky_tolerance=1.1, zenith_threshold_for_zero_dni=88.0, zenith_threshold_for_clearsky_limit=80.0)
    
    #adding DNI to dataframe with PV weather data
    pvlib_df['dni'] = dni
    
    #replacing 'NAN' in DNI column with 0 to prevent 'holes' in graphs (NANs are caused by the zenith angle being larger than the 'zenith threshold for zero dni' angle set with GHI and DHI not yet being zero. The DNI should be 0 in that case)
    pvlib_df['dni'] = pvlib_df['dni'].fillna(0)
    
    #determining PV power generation
    # location = (lat,lon)
    feedin = pv_system.feedin(
        weather=pvlib_df,
        location=(lat_location, lon_location))
    
    feedinarray = multsolar*feedin.values #hourly energy production over the year of 1 solar platform of the specified kind in the specified location
    feedinarray[feedinarray<0]=0
    
    #plotting PV power generation
    plt.figure()
    feedin.plot(title='PV feed-in '+str(lat_location)+', '+str(lon_location))
    plt.xlabel('Time')
    plt.ylabel('Power in W');
    
    feedin_index = feedin.index
    
    return feedinarray    

# feedinarray = func_PV(area)

#%% Feedin Index function


def func_Feedin_index(location):
    
    lat_location = location[1]
    lon_location = location[0]
    
    # get modules
    module_df = get_power_plant_data(dataset='SandiaMod') #retrieving dataset for PV modules
    
    # get inverter data
    inverter_df = get_power_plant_data(dataset='cecinverter') #retrieving dataset for inverters
    
    temp_params = pvlib.temperature.TEMPERATURE_MODEL_PARAMETERS['sapm']['open_rack_glass_polymer'] #defining temperature model parameters (leaving at default causes errors)
    
    #PV system definition
    system_data = {
        'module_name': 'Advent_Solar_Ventura_210___2008_',  # module name as in database
        'inverter_name': 'ABB__MICRO_0_25_I_OUTD_US_208__208V_',  # inverter name as in database
        'azimuth': 180, #angle of sun position with north in horizontal plane
        'tilt': 30, #angle of solar panels with horizontal
        'albedo': 0.075, #albedo (fraction of sun light reflected) of ocean water (https://geoengineering.global/ocean-albedo-modification/)
        'temperature_model_parameters': temp_params,
    }
    
    pv_system = Photovoltaic(**system_data)
    
    #getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
    # area = [lon, lat]
    pvlib_df = era5.weather_df_from_era5(     
        era5_netcdf_filename=era5_netcdf_filename,
        lib='pvlib', area=[lon_location,lat_location])
    
    #determining the zenith angle (angle of sun position with vertical in vertical plane) in the specified locations for the time instances downloaded from ERA5
    zenithcalc = pvlib.solarposition.get_solarposition(time=pvlib_df.index,latitude=latitude, longitude=longitude, altitude=None, pressure=None, method='nrel_numpy', temperature=pvlib_df['temp_air'])
    
    #determining DNI from GHI, DHI and zenith angle for the time instances downloaded from ERA5
    dni = pvlib.irradiance.dni(pvlib_df['ghi'],pvlib_df['dhi'], zenith=zenithcalc['zenith'], clearsky_dni=None, clearsky_tolerance=1.1, zenith_threshold_for_zero_dni=88.0, zenith_threshold_for_clearsky_limit=80.0)
    
    #adding DNI to dataframe with PV weather data
    pvlib_df['dni'] = dni
    
    #replacing 'NAN' in DNI column with 0 to prevent 'holes' in graphs (NANs are caused by the zenith angle being larger than the 'zenith threshold for zero dni' angle set with GHI and DHI not yet being zero. The DNI should be 0 in that case)
    pvlib_df['dni'] = pvlib_df['dni'].fillna(0)
    
    #determining PV power generation
    # location = (lat,lon)
    feedin = pv_system.feedin(
        weather=pvlib_df,
        location=(lat_location, lon_location))
    
    feedin_index = feedin.index
    
    return feedin_index    

#%% --- Wind function ---

def func_Wind(location):
    
    lat_location = location[1]
    lon_location = location[0]
    
    df = get_turbine_types(print_out=False)
    
    #defining the wind turbine system
    turbine_data= {
        "turbine_type": "E-101/3050",  # turbine type as in register
        "hub_height": 130,  # in m
    }
    my_turbine = WindTurbine(**turbine_data)
    
    #getting the needed weather data for wind power calculations from the file as downloaded with a seperate script from ERA5
    # [lon,lat]
    windpowerlib_df = era5.weather_df_from_era5(
        era5_netcdf_filename=era5_netcdf_filename,
        lib='windpowerlib', area=[lon_location, lat_location])
    
    #increasing the time indices by half an hour because then they match the pvlib time indices so the produced energy can be added later
    #assumed wind speeds half an hour later are similar and this will not affect the results significantly
    windpowerlib_df.index = windpowerlib_df.index + timedelta(minutes=30)
    
    # power output calculation for e126
    
    # own specifications for ModelChain setup
    modelchain_data = {
        'wind_speed_model': 'logarithmic',      # 'logarithmic' (default),
                                                # 'hellman' or
                                                # 'interpolation_extrapolation'
        'density_model': 'ideal_gas',           # 'barometric' (default), 'ideal_gas'
                                                #  or 'interpolation_extrapolation'
        'temperature_model': 'linear_gradient', # 'linear_gradient' (def.) or
                                                # 'interpolation_extrapolation'
        'power_output_model': 'power_curve',    # 'power_curve' (default) or
                                                # 'power_coefficient_curve'
        'density_correction': True,             # False (default) or True
        'obstacle_height': 0,                   # default: 0
        'hellman_exp': None}                    # None (default) or None
    
    # initialize ModelChain with own specifications and use run_model method to
    # calculate power output
    mc_my_turbine = ModelChain(my_turbine, **modelchain_data).run_model(windpowerlib_df)
    # write power output time series to WindTurbine object
    my_turbine.power_output = mc_my_turbine.power_output
    
    # title = 'Wind turbine power production at ' + str(location)
    # # plot turbine power output
    # plt.figure()
    # my_turbine.power_output.plot(title=title)
    # plt.xlabel('Time')
    # plt.ylabel('Power in W')
    # plt.show()
    
    my_turbinearray = multwind*my_turbine.power_output.values #hourly energy production over the year of 1 wind turbine of the specified kind in the specified location
    
    return my_turbinearray


#%% --- Transport Cost Function & Parameters ---

# TC , Transport cost


coords_production = (latitude, longitude)
# coords_port = (53.43, 6.84) 
# coords_demand = (53.43, 6.84)  



# distancesea = distanceseafactor*geopy.distance.geodesic(coords_production, coords_port).km #distance to be travelled over sea from production to demand location in km
distanceland = geopy.distance.geodesic(coords_port, coords_demand).km #distance to be travelled over land from production to demand location in km

Xtransport = demand #amount to be transported by ship in tons of hydrogen, assumed to be equal to demand as hydrogen losses are almost zero 
# Xkmsea = Xtransport*distancesea #yearly amount of tonkm to be made over sea
Xkmland = demand*distanceland #yealy amount of tonkm over land

Ckmammonia = np.array([float(cell.value) for cell in ammonia[158][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as ammonia in 10^3 euros over several years
Cbasetransportammonia = np.array([float(cell.value) for cell in ammonia[159][2:2+Nsteps]])
Ckmliquid = np.array([float(cell.value) for cell in liquidhydrogen[168][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as liquid hydrogen in 10^3 euros over several years
Cbasetransportliquid =  np.array([float(cell.value) for cell in liquidhydrogen[169][2:2+Nsteps]])
Ckmland = np.array([float(cell.value) for cell in landtransport[33][2:2+Nsteps]]) #costs per year per km of overland pipeline for 1 ton of hydrogen

# Changes by transport mode: pipeline, vs shipping
# Shipping cost
Ckmsea = [Ckmammonia, Ckmliquid] #costs per ton per km of overseas transport, depending on whether ammonia or liquid hydrogen is chosen
Cbasetransport = [Cbasetransportammonia, Cbasetransportliquid] #baserate of the transport per ton hydrogen

# Transport Cost Pipeline computation

# Parameters NH3 (j=4)
Lifetime_Pipe_NH3 = 30 # Project Lifetime NH3 pipeline [years]
a_pipe_NH3 =(DiscountRate*(1+DiscountRate)**Lifetime_Pipe_NH3)/((1+DiscountRate)**Lifetime_Pipe_NH3) # Amortization factor
C_pipe_NH3 = 771000 # NH3 Pipeline cost per km [Eur]
C_pump_NH3 = 1800000 # NH3 pump cost per station [Eur/pumpstation]
Pump_distance = 128.8 # Distance above which an additional pump is required [km]
Offshore_factor = 2 # Multiplier for estimating pipe cost (1: onshore, 2: offshore)

# Function calculating the transport cost parameter for NH3 pipeline
def Transport_NH3_Pipe(distance):
    # Input: distance [km], Output: 3xNsteps Cost array for each year l, row 1: total yearly cost, row 2: CAPEX, row 3: OPEX

    n_pump = math.ceil(distance/128.8) # Number of pump stations required at selected distance [pumpstations]
    CAPEX_pipe_NH3 = a_pipe_NH3*Offshore_factor*(C_pipe_NH3*distance+C_pump_NH3*n_pump) # Yearly CAPEX
    OPEX_pipe_NH3 = 0.02*Offshore_factor*(C_pipe_NH3*distance+C_pump_NH3*n_pump) # Yearly OPEX (2% of CAPEX)
    
    # Prepare output cost array
    C_pipeline_NH3 = np.zeros(Nsteps)
    
    # Create data array for each studied year
    for l in range(Nsteps):
        C_pipeline_NH3[l] = CAPEX_pipe_NH3 + OPEX_pipe_NH3
    return C_pipeline_NH3


# Parameters GH2 (j=3)
Lifetime_Pipe_GH2 = 30 # Project Lifetime NH3 pipeline [years]
peak_x3 = 35 # Peak number of electrolyzers required in project [# of electrolyzers]
a_pipe_GH2 =(DiscountRate*(1+DiscountRate)**Lifetime_Pipe_GH2)/((1+DiscountRate)**Lifetime_Pipe_GH2) # Amortization factor [-]
v_pipe_GH2 = 15 # flow rate [m/s]
rho_pipe_GH2 = 8 # density GH2 [kg/m3]
Q_pipe_GH2 = 1000*beta*peak_x3/3600 # mass flow rate [kg/s]
F_pipe_GH2 = Q_pipe_GH2/rho_pipe_GH2
D_pipe_GH2 = math.sqrt((4*F_pipe_GH2)/(math.pi*v_pipe_GH2)) # Inner diameter GH2 pipe [m]
C_pipe_GH2 = 4000000*D_pipe_GH2*D_pipe_GH2 + 598600*D_pipe_GH2 + 329000 # GH2 Pipeline cost per km [Eur]


def Transport_GH2_Pipe(distance):
    
    CAPEX_pipe_GH2 = a_pipe_GH2*Offshore_factor*C_pipe_GH2*distance
    OPEX_pipe_GH2 = 0.02*Offshore_factor*C_pipe_GH2*distance
    
    # Prepare output cost array
    C_pipeline_GH2 = np.zeros(Nsteps)
    
    # Create data array for each studied year
    for l in range(Nsteps):
        C_pipeline_GH2[l] = CAPEX_pipe_GH2 + OPEX_pipe_GH2
    
    return C_pipeline_GH2


# Energy medium set J (1 NH3 ship, 2 LH2 ship, 3 GH2 pipe, 4 NH3 pipe)
J = [1, 2, 3, 4]


if E==0:
    print('Energy Medium set to: NH3 ship')
elif E==1:
    print('Energy Medium set to: LH2 ship')
elif E==2:
    print('Energy Medium set to: GH2 pipe')
elif E==3:
    print('Energy Medium set to: NH3 pipe')


# Prepare TC_jl cost parameter array
def func_TC(location, E):
    location = [location[1],location[0]]
    distancesea = distanceseafactor*geopy.distance.geodesic(location, coords_port).km
    Xkmsea = Xtransport*distancesea
    Cpipeline = np.array([Transport_NH3_Pipe(distancesea), Transport_GH2_Pipe(distancesea)])
    
    TC = []
    for j in J[0:2]:    # Shipping
        TCyear = Xtransport*Cbasetransport[j-1] + Xkmsea*Ckmsea[j-1] + Xkmland*Ckmland #Euro
        TC.append(TCyear)

    for j in J[2:4]:  # Pipeline
        TCyear = Cpipeline[j-3] # [Euro]
        TC.append(TCyear)
            
    return TC

# TC = func_TC(area,E)



#%% Location selection

# # Define the dimensions of the matrix (play with these 4 values to determine start location)
size = 6  # This will create a size*size matrix
# Calculate the range for rows and columns
start = -3.5 # Starting position relative to longitude
end = 4 # Ending position relative to latitude
resolution_map = 1 # Distance between locations



# Create a matrix of the specified size, with each element being a tuple of length 2
loc_matrix = np.empty((size, size), dtype=object)
for i in range(size):
    for j in range(size):
        loc_matrix[i,j] = (start + j*resolution_map + longitude, end - i*resolution_map + latitude)
        

df_map_test = pd.DataFrame(columns=['longitude','latitude','z'],index=[list(range(size*size))])

fig_map = px.density_mapbox(df_map_test, lat = 'latitude', lon = 'longitude', z = 'z',
                        radius = 15,
                        center = dict(lat = latitude, lon = longitude),
                        zoom = 3,
                        mapbox_style = 'open-street-map',
                        color_continuous_scale = 'Rainbow')



for i in range(size):
    for j in range(size):
        fig_map.add_trace(go.Scattermapbox(
            lat=[loc_matrix[i,j][1]],
            lon=[loc_matrix[i,j][0]],
            mode='markers',
            marker=dict(size=10, color="Orange"),
            name="Location "+str(loc_matrix[i,j]),
    
        ))  


pio.renderers.default='browser'
fig_map.show()

fig_map.data = []

#%% Bathymetry

def get_water_depth(lat, lon, bathymetry_file='NetCDF/north_sea_bathymetry.nc'):
    """
    Returns the water depth (bathymetry) for a given latitude and longitude from a bathymetry dataset.

    Parameters:
    lat (float): Latitude of the location
    lon (float): Longitude of the location
    bathymetry_file (str): Path to the NetCDF bathymetry file

    Returns:
    float: Water depth at the specified lat/lon location in meters. Positive values indicate depth.
    """

    # Open the NetCDF file containing bathymetry data
    try:
        ds = xr.open_dataset(bathymetry_file)
    except FileNotFoundError:
        print("Bathymetry file not found.")
        return None

    # Check if the file contains 'latitude', 'longitude', and 'elevation' variables
    if 'lat' not in ds.variables or 'lon' not in ds.variables or 'elevation' not in ds.variables:
        print("NetCDF file does not contain required variables.")
        return None

    # Find the nearest lat/lon indices in the dataset
    lat_idx = np.abs(ds['lat'] - lat).argmin().item()
    lon_idx = np.abs(ds['lon'] - lon).argmin().item()

    # Extract the water depth at the nearest lat/lon indices
    water_depth = ds['elevation'][lat_idx, lon_idx].item()

    # Close the dataset
    ds.close()

    return water_depth



#%% Prepare location loop



list_vert_locs = []
for vert in range(size):
    
    
    dict_of_df = {}
    for loci in loc_matrix[vert]:
        dict_of_df[loci] = pd.DataFrame(index=['Demand (tons of hydrogen)', 'Usage location', 'Year','Production location','Transfer port','Total costs per year (euros)','Costs per kg hydrogen (euros)','Wind turbines', 'Solar platforms','Electrolyzers','Desalination equipment', 'Storage volume (m3)','Conversion devices','Reconversion devices','Transport medium', 'FPSO volume (m3)', 'Distance sea (km)','Distance land (km)'])
    
    feedinlist = []
    counter = 0
    
    for loc in loc_matrix[vert]:
        # Select location 'loc' (varies by longitude)
        #loc = (lon,lat)
        
        # Calculate PV
        feedinarray = func_PV(list(loc))
        
        # Calculate Wind
        my_turbinearray = func_Wind(loc)
        
        lon_prod = loc[0]
        lat_prod = loc[1]
        
        # Calculate TC
        TC = func_TC(loc,E)
        print('Currently in location: (lat=', str(lat_prod),', lon=',str(lon_prod),')')
        #(lat, lon)
        distancesea = distanceseafactor*geopy.distance.geodesic((lat_prod,lon_prod), coords_port).km

    
        #%% --- Model parameters and sets ---
        
        # Set model name
        model = Model('GFPSO Cost Optimization')
        
        # ---- Sets ----
        
        I = [1, 2] # Conv devices (1: Conv, 2: Reconv)
        J = [1, 2, 3, 4] # Energy Medium (1: Ammonia, 2: LiquidH2)
        K = [1, 2, 3, 4] # Device types (1: Wind, 2: Solar, 3: Elec, 4: Desal)
        L = range(Nsteps) # Years in time period
        N = [1, 2] # Volume based equipment (1: Storage, 2: FPSO)
        T = list(range(0,len(feedinarray))) # Operational hours in year
        
        
        # ---- Parameters ----
        
        # Cost parameters
        
        Cconvammonia = DataYearRatio*rel_con_array(Startyear, Nsteps, timestep, Con_Costs)
        # Cconvliquid = DataYearRatio*rel_con_arrayLH2(Startyear, Nsteps, timestep, LH2_Con_Costs)
        # Creconvammonia = DataYearRatio*rel_rec_array(Startyear, Nsteps, timestep, Rec_Costs)
        # Creconvliquid = DataYearRatio*rel_rec_arrayLH2(Startyear, Nsteps, timestep, LH2_Rec_Costs)
        
        Cconvammonia = DataYearRatio*np.array([float(cell.value) for cell in ammonia[48][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
        Cconvliquid = DataYearRatio*np.array([float(cell.value) for cell in liquidhydrogen[61][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years
        Creconvammonia = DataYearRatio*np.array([float(cell.value) for cell in ammonia[111][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
        Creconvliquid = DataYearRatio*np.array([float(cell.value) for cell in liquidhydrogen[125][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years
        Cconvgas = np.zeros(Nsteps) # j=3
        Creconvgas = np.zeros(Nsteps) #j=3
        
        
        # A_lij , l = year in time period, i = conversion device type, j = medium
        A = []
        for l in range(Nsteps):
            Aarray = np.array([[Cconvammonia[l], Cconvliquid[l], Cconvgas[l], Cconvammonia[l]], [Creconvammonia[l], Creconvliquid[l], Creconvgas[l], Creconvammonia[l]]])
            A.append(Aarray)
        
        Cs1 = DataYearRatio*np.array([float(cell.value) for cell in solar[51][2:2+Nsteps]])
        # Cw1_floating = DataYearRatio*np.array([float(cell.value) for cell in wind[48][2:2+Nsteps]])
        Cw1 = DataYearRatio*rel_wind_array(Startyear, Nsteps, timestep, Wind_Costs)
        Ce = DataYearRatio*np.array([float(cell.value) for cell in electrolyzer[50][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an electrolyzer in 10^3 euros over several years
        Cd = DataYearRatio*np.array([float(cell.value) for cell in desalination[49][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a desalination installation in 10^3 euros over several years
        
        # B_lk , l = year in time period, k = device type (solar, wind, elec, desal)
        B = []
        for l in L:
            Barray = np.array([Cw1[l], 
                               Cs1[l], 
                               Ce[l], 
                               Cd[l]])
            B.append(Barray)
        
                
        # C_lnj
        Cstliquid =  DataYearRatio*np.array([float(cell.value) for cell in storage[25][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of storage per m3 in euros over several years
        Cstammonia =  DataYearRatio*np.array([float(cell.value) for cell in storage[54][2:2+Nsteps]])
        Cstgas = np.zeros(Nsteps)  # Assumption: Pipeline counts as storage
        Cstammoniapipe = np.zeros(Nsteps) # Assumption: Pipeline counts as storage
        
        Cfpso = DataYearRatio* np.array([float(cell.value) for cell in fpso[53][2:2+Nsteps]]) #np.array([1,1,0.8,0.6,0.4]) #cost per year (depreciation+OPEX) of FPSO per m3 in 10^3 euros over several years
        Cst = [Cstammonia, Cstliquid, Cstgas, Cstammoniapipe]
        
        
        
        C = []
        for l in L:
            Carray = np.array([Cst[E][l], 
                               Cfpso[l]])
            C.append(Carray)
        
        # D , Yearly demand
        D = (DataYearRatio)*demand # [tonH2/yr]
        
            
        
        
        # Non-cost parameters
        
        electrolyzer_energy = 1000*electrolyzer.cell(row=7, column=2).value #energy requirement of electrolyzer in Wh per ton of hydrogen
        gamma = electrolyzer_energy #50500000 [Wh/tonH2]
        
        Ed = 35
        Ec1 = 750/0.18
        Ec2 = 10180
        Ec3 = 0
        Ec4 = 750/0.18
        Es1 = 210
        Es2 = 600
        Es3 = 0
        Es4 = 0
        
        fracpowerelectrolyzerliquid = electrolyzer.cell(row=11, column=2).value #fraction of energy used by electrolyzers when using liquid hydrogen
        fracpowerelectrolyzerammonia = electrolyzer.cell(row=12, column=2).value #fraction of energy used by elecyrolyzers when using ammonia
        
        alpha_1 = (gamma/1000) / (Ed + Ec1 + Es1 + (gamma/1000))
        alpha_2 = (gamma/1000) / (Ed + Ec2 + Es2 + (gamma/1000))
        alpha_3 = (gamma/1000) / (Ed + Ec3 + Es3 + (gamma/1000))
        alpha_4 = (gamma/1000) / (Ed + Ec4 + Es4 + (gamma/1000))
        
        #alpha_j
        
        alpha = [alpha_1, alpha_2, alpha_3, alpha_4]
        
        # capelectrolyzerhour = electrolyzer.cell(row=9, column=2).value #hourly output capacity of electrolyzers in tons of hydrogen per hour
        # beta = capelectrolyzerhour
        
        
        
        electrolyzer_water = electrolyzer.cell(row=6, column=2).value #water requirement of electrolyzer in m3 of water per ton of hydrogen
        capdesalinationhour = desalination.cell(row=14, column=2).value #hourly output capacity desalination in m3 of water per hour
        epsilon = electrolyzer_water/capdesalinationhour
        
        eta_conversionammonia = ammonia.cell(row=13, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
        eta_conversionliquid =  liquidhydrogen.cell(row=16, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
        eta_reconversionammonia = eta_conversionammonia #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
        eta_reconversionliquid = eta_conversionliquid #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
        eta_conversiongas = 1
        eta_reconversiongas = 1
        
        # eta_ij , i = conv device type, j = energy medium and mode
        eta = 1/np.array([[eta_conversionammonia, eta_conversionliquid, eta_conversiongas, eta_conversionammonia],
                         [eta_reconversionammonia, eta_reconversionliquid, eta_reconversiongas, eta_reconversionammonia]])
        
        #nu_j
        volumefpsoliquid = 20*324.5*60.69*33.8/600 # (j=2) electrolyzer.cell(row=15, column=2).value #FPSO volume per electrolyzer liquid hydrogen
        volumefpsoammonia = volumefpsoliquid-(volumefpsoliquid*0.4469-(volumefpsoliquid*0.4469/(0.681*0.18/0.071)))  #(j=1) electrolyzer.cell(row=16, column=2).value #FPSO volume per electrolyzer ammonia
        volumefpsogaseouspipe = volumefpsoliquid # (j=3)
        volumefpsoammoniapipe = volumefpsoammonia # (j=4)
        
        nu = [volumefpsoammonia, volumefpsoliquid, volumefpsogaseouspipe, volumefpsoammoniapipe]
        
        #phi_j
        ratiostoragefpsoliquid = fpso.cell(row=18, column=2).value #ratio storage tanks in m3/fpso volume in m3 (j=2)
        ratiostoragefpsoammonia = fpso.cell(row=19, column=2).value #ratio storage tanks in m3/fpso volume in m3 (j=1)
        ratiostoragefpsogaseouspipe = 0 # Assumption: Pipeline counts as storage j=3
        ratiostoragefpsoammoniapipe = 0 # Assumption: Pipeline counts as storage j=4
        
        phi = [ratiostoragefpsoammonia, ratiostoragefpsoliquid, ratiostoragefpsogaseouspipe, ratiostoragefpsoammoniapipe]
        
        
        # Conversion
        capconvammonia = DataYearRatio*ammonia.cell(row=10, column=2).value #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for ammonia
        capconvliquid = DataYearRatio*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for liquid hydrogen
        
        w11 = math.ceil(1.6*D/(eta[1][0]*capconvammonia))   # NH3 j=1
        w12 = math.ceil(1.6*D/(eta[1][1]*capconvliquid))               # LH2 j=2
        w13 = 0                                             # GH2 j=3 (no need for conversion)
        w14 = math.ceil(1.6*D/(eta[1][0]*capconvammonia))   # NH3 j=4 (same as w11, both are ammonia)
        
        
        # Reconversion
        capreconvammonia = DataYearRatio*ammonia.cell(row=75, column=2).value #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for ammonia
        capreconvliquid = DataYearRatio*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for liquid hydrogen
        
        w21 = math.ceil(D/capreconvammonia)    # NH3 j=1
        w22 = math.ceil(D/capreconvliquid)     # LH2 j=2
        w23 = 0                     # GH2 j=3
        w24 = math.ceil(D/capreconvammonia)    # NH3 j=4
        
        
        # w_ij Number of conversion and reconversion devices 
        W = [[w11, w12, w13, w14], [w21, w22, w23, w24]]
        
        
        #  For now the converters and reconverter amount is constant WC[i][j]
        WC = quicksum(W[i-1][E]*A[l][i-1][E] for i in I) 
        
        #%% --- Variables ---
        
        # # w[i,j]
        # w = {}
        # for i in I:
        #     for j in J:
        #         w[i,j] = model.addVar (lb = 0, vtype = GRB.INTEGER, name = 'w[' + str(i) + ',' + str(j) + ']' )
        
        # x[k]
        x = {}
        for k in K:
            x[k] = model.addVar (lb = 0, vtype = GRB.INTEGER, name = 'x[' + str(k) + ']' )
            
        # y[n]
        y = {}
        for n in N:
            y[n] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'y[' + str(n) + ']' )
        
        PU = {}
        for t in T:
            PU[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PU[' + str(t) + ']' )
            
        # Power Generated (used to lower RHS values)
        PG = {}
        for t in T:
            PG[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PG[' + str(t) + ']' )
        
        # Hydrogen produced
        h = {}
        for t in T:
            h[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'h[' + str(t) + ']' )
        
        
        
        #%% --- Integrate new variables ---
        model.update()
        
        
        #%% --- Constraints ---
        
        for t in T:
            model.addConstr(PG[t] == my_turbinearray[t]*x[1] + feedinarray[t]*x[2])
            model.addConstr(PU[t] <= alpha[E]*PG[t])
            model.addConstr(PU[t] <= beta*gamma*x[3])
            model.addConstr(h[t] == PU[t]/gamma)
        
        model.addConstr(quicksum(h[t] for t in T) >= D/(eta[0][E]*eta[1][E]))  
        
        model.addConstr(x[4] >= x[3]*beta*epsilon)
        
        model.addConstr(y[2] == x[3]*nu[E])
        
        model.addConstr(y[1] == y[2]*phi[E])
        
        # model.addConstr(x[1] == 0)# Force no wind, only solar
        # print('Forcing wind turbines to 0 units (Check Constraints)')
        
        
        #%% --- Run Optimization ---
        model.update()
        
        model.setParam( 'OutputFlag', False) # gurobi output or not (If you want ouput, keep the line. If you dont want output, comment line out)
        model.Params.NumericFocus = 1
        model.Params.timeLimit = 400
        # model.optimize()
        
        Runtime_List = []
        Result = []
        feedin_index = func_Feedin_index(list(loc))
        hdata = pd.DataFrame(index=feedin_index) #for plotting the hydrogen production
        exceldf = pd.DataFrame(index=['Demand (tons of hydrogen)', 'Usage location', 'Year','Production location','Transfer port','Total costs per year (euros)','Costs per kg hydrogen (euros)','Wind turbines', 'Solar platforms','Electrolyzers','Desalination equipment', 'Storage volume (m3)','Conversion devices','Reconversion devices','Transport medium', 'FPSO volume (m3)', 'Distance sea (km)','Distance land (km)'])
        demandlocation = 'Groningen' #location where hydrogen is asked
        productionlocation = 'North-Sea' #location where hydrogen is produced
        transferport = 'Groningen' #port where hydrogen is transferred from sea to land transport
        
        
        transportmedium = str('')
        if E == 0:
            transportmedium = 'NH3 ship'
        elif E == 1:
            transportmedium = 'LH2 ship'
        elif E == 2:
            transportmedium = 'GH2 pipe'
        elif E == 3:
            transportmedium = 'NH3 pipe'
        
        # --- Objective function ---
        
        for l in L: 
            model.reset()
            model.setObjective (WC + quicksum(x[k]*B[l][k-1] for k in K) + quicksum(y[n]*C[l][n-1] for n in N) + TC[E][l])
            model.modelSense = GRB.MINIMIZE
            model.update ()
            model.optimize()
            Result.append(model.ObjVal)
            Runtime_List.append(model.Runtime)
            exceldf[timestep*l+Startyear] = [demand,demandlocation,timestep*l+Startyear, loc, transferport, model.ObjVal*(1/DataYearRatio), model.ObjVal/D/1000,x[1].X, x[2].X, x[3].X, x[4].x, y[1].X, W[0][E], W[1][E], transportmedium, y[2].X,distancesea,distanceland]
            dict_of_df[loc][timestep*l+Startyear] = [demand,demandlocation,timestep*l+Startyear, loc, transferport, model.ObjVal*(1/DataYearRatio), model.ObjVal/D/1000,x[1].X, x[2].X, x[3].X, x[4].x, y[1].X, W[0][E], W[1][E], transportmedium, y[2].X,distancesea,distanceland]
            print('------- Completed run: ', l+1, ' out of ', max(L)+1 , '     (Year: ',Startyear+timestep*l ,')')
        print('--------- Completed horizontal loc runs: ', counter+1, 'out of ', size)    
        print('--------- Completed vertical loc runs: ', vert+1, 'out of ', size)
        
        
        
        
        #%% --- Post-Processing ---
        
        # Data with latitude/longitude and values
        df = pd.DataFrame(columns=['longitude','latitude','Cost_per_kg'],index=[list(range(size))])
        # Add resulting cost per kg of h2 in 2050
        Cost_per_kg = dict_of_df[loc].loc['Costs per kg hydrogen (euros)',2050]
        df.loc[counter,'Cost_per_kg'] = Cost_per_kg
        df.loc[counter,'longitude'] = loc_matrix[vert][counter][0]
        df.loc[counter,'latitude'] = loc_matrix[vert][counter][1]
        
        list_vert_locs.append(dict_of_df)
        # dict_of_df[loc][timestep*l+Startyear] = [demand*(1/DataYearRatio),demandlocation,timestep*l+Startyear, productionlocation, transferport, model.ObjVal*(1/DataYearRatio), model.ObjVal/demand/1000,x[1].X, x[2].X, x[3].X, x[4].x, y[1].X, W[0][E], W[1][E], transportmedium, y[2].X,distancesea,distanceland]
        counter = counter + 1
        # df.loc[0,'ObjVal'] = model.ObjVal

df.to_csv("test_csv.csv")


print("--- Runtime Statistics ---")
print("Total Runtime: ",sum(Runtime_List))
print("Average Runtime: ", statistics.mean(Runtime_List))
print("Max Runtime: ", max(Runtime_List))
print("Min Runtime: ", min(Runtime_List))
    
if E==0:
    title_str = ' NH3 ship'
elif E==1:
    title_str = ' LH2 ship'
elif E==2:
    title_str = ' GH2 pipe'
elif E==3:
    title_str = ' NH3 pipe'


#%% --- Plotting ---
counter = 0
# list_dfs = []
# for vert in range(size):
#     for i in range(len(list_vert_locs)):
#         for j in range(size):
#             list_dfs.append(list_vert_locs[i][loc_matrix[vert][j]])

list_dfs2 = []
for vert in range(size):
    for j in range(size):
        # print(counter)
        # print('Iteration: ', counter, ' Location: ', loc_matrix[vert][j]) 
        list_dfs2.append(list_vert_locs[counter][loc_matrix[vert][j]])
        counter = counter + 1

COMPLETE_OUTPUT = list_dfs2

# All values in 2050
All_Headers = exceldf.index.to_list()

counter = 0
df_full_all = pd.DataFrame(columns=['longitude','latitude','Cost_per_kg', 'Wind turbines', 'Solar platforms', 'Electrolyzers', 'Desalination equipment', 'Storage volume', 'Conversion devices', 'Reconversion devices', 'Transport medium', 'FPSO volume', 'Distance sea'],index=[list(range(size*size))])
for vert in range(size):
    counter2=0
    for j in range(size):
        Cost_per_kg = COMPLETE_OUTPUT[counter].loc['Costs per kg hydrogen (euros)',2050]
        df_full_all.loc[counter,'Cost_per_kg'] = Cost_per_kg
        df_full_all.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
        df_full_all.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
        df_full_all.loc[counter,'Wind turbines'] = COMPLETE_OUTPUT[counter].loc['Wind turbines',2050]
        df_full_all.loc[counter,'Solar platforms']= COMPLETE_OUTPUT[counter].loc['Solar platforms',2050]
        df_full_all.loc[counter,'Electrolyzers'] = COMPLETE_OUTPUT[counter].loc['Electrolyzers',2050]
        df_full_all.loc[counter,'Desalination equipment'] = COMPLETE_OUTPUT[counter].loc['Desalination equipment',2050]
        df_full_all.loc[counter,'Storage volume'] = COMPLETE_OUTPUT[counter].loc['Storage volume (m3)',2050]
        df_full_all.loc[counter,'Conversion devices'] = COMPLETE_OUTPUT[counter].loc['Conversion devices',2050]
        df_full_all.loc[counter,'Reconversion devices'] = COMPLETE_OUTPUT[counter].loc['Reconversion devices',2050]
        df_full_all.loc[counter,'Transport medium'] = COMPLETE_OUTPUT[counter].loc['Transport medium',2050]
        df_full_all.loc[counter,'FPSO volume'] = COMPLETE_OUTPUT[counter].loc['FPSO volume (m3)',2050]
        df_full_all.loc[counter,'Distance sea'] = COMPLETE_OUTPUT[counter].loc['Distance sea (km)',2050]
        counter2 = counter2 + 1     
        counter = counter + 1  



# Cost_per_kg in 2050
counter = 0
df_full = pd.DataFrame(columns=['longitude','latitude','Cost_per_kg'],index=[list(range(size*size))])
for vert in range(size):
    counter2=0
    for j in range(size):
        Cost_per_kg = list_dfs2[counter].loc['Costs per kg hydrogen (euros)',2050]
        df_full.loc[counter,'Cost_per_kg'] = Cost_per_kg
        df_full.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
        df_full.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
        counter2 = counter2 + 1     
        counter = counter + 1      

# # LCOH in across years
# counter = 0
# df_full = pd.DataFrame(columns=['longitude','latitude','LCOH'],index=[list(range(size*size))])
# for vert in range(size):
#     counter2=0
#     for j in range(size):
#         LCOH = list_dfs2[counter].loc['Costs per kg hydrogen (euros)',2050]
#         df_full.loc[counter,'LCOH'] = LCOH
#         df_full.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
#         df_full.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
#         counter2 = counter2 + 1     
#         counter = counter + 1   




df_full_backup = df_full
df_full = df_full_backup
avg = df_full['Cost_per_kg'].mean()
# df_full['Cost_per_kg'] = df_full['Cost_per_kg'].apply(lambda x: x-avg)      
# df_full['Cost_per_kg'] = df_full['Cost_per_kg'].apply(lambda x: x*100)  

# # Filter the dataset to only include values between lat=54.55-53.55, lon=6.08-8.28
# df_relevant_lat = df_full.iloc[0:160]


# lon_value_to_remove = 8.48 #8.48, 8.68, 8.88, 9.08, 9.28, 9.48, 9.68, 9.88
# remove_array = [8.48, 8.68, 8.88, 9.08, 9.28, 9.48, 9.68, 9.88]
# for i in remove_array:
#     # df_filtered = df_relevant_lat[df_relevant_lat['longitude'] != i]
#     # df_relevant_lat = df_filtered
#     df_relevant_lat = df_relevant_lat[df_relevant_lat['longitude'] != i]
# print(df_relevant_lat)

# df_relevant_lat = df_full.loc[(df_full['latitude'] >=53.55) & (df_full['longitude'] <=8.28)]
df_relevant_lat = df_full



#%% Only smaller region

# Color palettes: 'RdBu', 
fig = px.density_mapbox(df_relevant_lat, lat = 'latitude', lon = 'longitude', z = 'Cost_per_kg',
                        radius = 15,
                        center = dict(lat = latitude, lon = longitude),
                        zoom = 3,
                        mapbox_style = 'open-street-map',
                        title = title_str,
                        color_continuous_scale = 'Rainbow')



# Adjust color of heatmap by adding more points for density
fig.add_trace(
    go.Scattermapbox(
        lat=df_relevant_lat["latitude"],
        lon=df_relevant_lat["longitude"],
        mode="markers",
        showlegend=False,
        hoverinfo="skip",
        marker={
            "color": df_relevant_lat["Cost_per_kg"],
            "size": df_relevant_lat["Cost_per_kg"].fillna(0).infer_objects(copy=False),
            "coloraxis": "coloraxis",
            # desired max size is 15. see https://plotly.com/python/bubble-maps/#united-states-bubble-map
            "sizeref": (df_relevant_lat["Cost_per_kg"].max()) / 15 ** 2,
            "sizemode": "area",
        },
    )
)

# Usage location
fig.add_trace(go.Scattermapbox(
        lat=[coords_demand[0]],
        lon=[coords_demand[1]],
        mode='markers',
        marker=dict(size=10, color="Orange"),
        name="Usage Location",
    
    ))


pio.renderers.default='browser'
fig.show()



#%%

fig = px.density_mapbox(df_full, lat = 'latitude', lon = 'longitude', z = 'Cost_per_kg',
                        radius = 15,
                        center = dict(lat = latitude, lon = longitude),
                        zoom = 3,
                        title = title_str,
                        mapbox_style = 'open-street-map',
                        color_continuous_scale = 'Rainbow')



# Adjust color of heatmap by adding more points for density
fig.add_trace(
    go.Scattermapbox(
        lat=df_full["latitude"],
        lon=df_full["longitude"],
        mode="markers",
        showlegend=False,
        hoverinfo="skip",
        marker={
            "color": df_full["Cost_per_kg"],
            "size": df_full["Cost_per_kg"].fillna(0).infer_objects(copy=False),
            "coloraxis": "coloraxis",
            # desired max size is 15. see https://plotly.com/python/bubble-maps/#united-states-bubble-map
            "sizeref": (df_full["Cost_per_kg"].max()) / 15 ** 2,
            "sizemode": "area",
        },
    )
)

# Usage location
fig.add_trace(go.Scattermapbox(
        lat=[coords_demand[0]],
        lon=[coords_demand[1]],
        mode='markers',
        marker=dict(size=10, color="Orange"),
        name="Usage Location",
    
    ))


pio.renderers.default='browser'
fig.show()

#%% Electrolyzer map 2050
df_map = df_full_all
# df_map = df_map.loc[(df_map['latitude'] >=53.55) & (df_map['longitude'] <=8.28)]
# df_map = df_map.loc[(df_map['Electrolyzers'] <70)]



# Color palettes: 'RdBu', 
fig = px.density_mapbox(df_map, lat = 'latitude', lon = 'longitude', z = 'Electrolyzers',
                        radius = 15,
                        center = dict(lat = latitude, lon = longitude),
                        zoom = 3,
                        mapbox_style = 'open-street-map',
                        title = title_str,
                        color_continuous_scale = 'Rainbow')



# Adjust color of heatmap by adding more points for density
fig.add_trace(
    go.Scattermapbox(
        lat=df_map["latitude"],
        lon=df_map["longitude"],
        mode="markers",
        showlegend=False,
        hoverinfo="skip",
        marker={
            "color": df_map['Electrolyzers'],
            "size": df_map['Electrolyzers'].fillna(0).infer_objects(copy=False),
            "coloraxis": "coloraxis",
            # desired max size is 15. see https://plotly.com/python/bubble-maps/#united-states-bubble-map
            "sizeref": (df_map['Electrolyzers'].max()) / 15 ** 2,
            "sizemode": "area",
        },
    )
)

# Usage location
fig.add_trace(go.Scattermapbox(
        lat=[coords_demand[0]],
        lon=[coords_demand[1]],
        mode='markers',
        marker=dict(size=10, color="Orange"),
        name="Usage Location",
    
    ))


pio.renderers.default='browser'
fig.show()


#%%
df_full.to_csv("df_full_csv.csv")
# df_full_all.to_csv("csv_files/Output_pipeline_2050.csv")

list_df_all = []
for year in list(COMPLETE_OUTPUT[0].columns):
    counter = 0
    df_full_all = pd.DataFrame(columns=['longitude','latitude','Cost_per_kg', 'Wind turbines', 'Solar platforms', 'Electrolyzers', 'Desalination equipment', 'Storage volume', 'Conversion devices', 'Reconversion devices', 'Transport medium', 'FPSO volume', 'Distance sea'],index=[list(range(size*size))])
    for vert in range(size):
        counter2=0
        for j in range(size):
            Cost_per_kg = COMPLETE_OUTPUT[counter].loc['Costs per kg hydrogen (euros)',year]
            df_full_all.loc[counter,'Cost_per_kg'] = Cost_per_kg
            df_full_all.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
            df_full_all.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
            df_full_all.loc[counter,'Wind turbines'] = COMPLETE_OUTPUT[counter].loc['Wind turbines',year]
            df_full_all.loc[counter,'Solar platforms']= COMPLETE_OUTPUT[counter].loc['Solar platforms',year]
            df_full_all.loc[counter,'Electrolyzers'] = COMPLETE_OUTPUT[counter].loc['Electrolyzers',year]
            df_full_all.loc[counter,'Desalination equipment'] = COMPLETE_OUTPUT[counter].loc['Desalination equipment',year]
            df_full_all.loc[counter,'Storage volume'] = COMPLETE_OUTPUT[counter].loc['Storage volume (m3)',year]
            df_full_all.loc[counter,'Conversion devices'] = COMPLETE_OUTPUT[counter].loc['Conversion devices',year]
            df_full_all.loc[counter,'Reconversion devices'] = COMPLETE_OUTPUT[counter].loc['Reconversion devices',year]
            df_full_all.loc[counter,'Transport medium'] = COMPLETE_OUTPUT[counter].loc['Transport medium',year]
            df_full_all.loc[counter,'FPSO volume'] = COMPLETE_OUTPUT[counter].loc['FPSO volume (m3)',year]
            df_full_all.loc[counter,'Distance sea'] = COMPLETE_OUTPUT[counter].loc['Distance sea (km)',year]
            df_full_all.loc[counter,'Demand [tonH2/yr]'] = COMPLETE_OUTPUT[counter].loc['Demand (tons of hydrogen)',year]
            counter2 = counter2 + 1     
            counter = counter + 1  
            
    # df_full_all.to_csv("csv_files/Output_pipeline_"+str(year)+".csv")
    list_df_all.append(df_full_all)

# create a excel writer object
with pd.ExcelWriter("csv_files/Output_j"+str(E+1)+".xlsx") as writer:
   
    # use to_excel function and specify the sheet_name and index 
    # to store the dataframe in specified sheet
    counter = 0
    for year in list(COMPLETE_OUTPUT[0].columns):
        list_df_all[counter].to_excel(writer, sheet_name=str(year), index=False)
        counter += 1

# list_vert_locs

# for loc in loc_matrix[vert]:
#     Cost_per_kg = dict_of_df[loc].loc['Costs per kg hydrogen (euros)',2050]
#     df.loc[counter,'Cost_per_kg'] = Cost_per_kg
#     df.loc[counter,'longitude'] = loc_matrix[vert][counter][0]
#     df.loc[counter,'latitude'] = loc_matrix[vert][counter][1]
#     counter = counter + 1

# fig = px.density_mapbox(df, lat = 'latitude', lon = 'longitude', z = 'Cost_per_kg',
#                         radius = 7,
#                         center = dict(lat = 26.81, lon = 126.94),
#                         zoom = 3,
#                         mapbox_style = 'open-street-map',
#                         color_continuous_scale = 'rainbow')

# # Usage location
# fig.add_trace(go.Scattermapbox(
#         lat=[35.64],
#         lon=[139.8],
#         mode='markers',
#         marker=dict(size=10, color="Orange"),
#         name="Usage Location",
    
#     ))


# pio.renderers.default='browser'
# fig.show()


