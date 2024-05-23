# -*- coding: utf-8 -*-
"""
Created on Thu May  2 11:41:52 2024

@author: spide
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import geopy.distance
import math
from openpyxl import load_workbook
from feedinlib import era5
import pvlib
from feedinlib import Photovoltaic
from feedinlib import get_power_plant_data
from windpowerlib.modelchain import ModelChain
from windpowerlib.wind_turbine import WindTurbine
from windpowerlib import get_turbine_types
from datetime import timedelta
from gurobipy import *

# Excel Parameters

#loading excel to later retrieve input data using openpyxl
data_file = os.getcwd() 
data_file = os.path.dirname(os.path.realpath('__file__')) + '\Inputdata econ.xlsx'
# data_file = 'C:/Users/tmell/Documents/TU Delft/Master/Afstuderen/Model/Inputdata econ.xlsx' #importing data from input data excel
wb = load_workbook(data_file,data_only=True) # creating workbook

general = wb['General'] #selecting data from sheet called general
wind = wb['Wind'] #selecting data from sheet called wind
solar = wb['Solar'] #selecting data from sheet called solar
electrolyzer = wb['Electrolyzer'] #selecting data from sheet called electrolyzer
desalination = wb['Desalination'] #selecting data from sheet called desalination
ammonia = wb['Ammonia'] #selecting data from sheet called ammonia
liquidhydrogen = wb['Liquid hydrogen'] #selecting data from sheet called liquid hydrogen
landtransport = wb['Land transport'] #selecting data from sheet called landtransport
storage = wb['Storage'] #selecting data from sheet called storage
fpso = wb['FPSO'] #selecting data from sheet called fpso

# Time period data
Startyear = general.cell(row=9, column=2).value #first year to be reviewed
timeperiod = general.cell(row=10, column=2).value #time period of simulation in years
timestep = general.cell(row=11, column=2).value #time step of simulation in years
Nsteps = int(timeperiod/timestep+1) #number of time steps (important for data selection from excel and loop at the end)

# Coordinates production location
latitude =  26.81
longitude = 126.94
# Solar and Wind 
multsolar = 524*1000/206.7989   # Installed capacity [kWp]
multwind = 12/3   # Capacity [MW]



#%%

# --- Solar Energy Generation ---

# set start and end date (end date will be included
# in the time period for which data is downloaded)
start_date, end_date = '2020-01-01', '2020-10-31'
# set variable set to download
variable = 'feedinlib'

era5_netcdf_filename = 'ERA5_weather_data_location4.nc' #referring to file with weather data downloaded earlier using the ERA5 API

area = [longitude, latitude] #location of production

#
#PV
#

# get modules
module_df = get_power_plant_data(dataset='SandiaMod') #retrieving dataset for PV modules

# get inverter data
inverter_df = get_power_plant_data(dataset='cecinverter') #retrieving dataset for inverters

temp_params = pvlib.temperature.TEMPERATURE_MODEL_PARAMETERS['sapm']['open_rack_glass_polymer'] #defining temperature model parameters (leaving at default causes errors)

#PV system definition
system_data = {
    'module_name': 'Advent_Solar_Ventura_210___2008_',  # module name as in database
    'inverter_name': 'ABB__MICRO_0_25_I_OUTD_US_208__208V_',  # inverter name as in database
    'azimuth': 180, #angle of sun position with north in horizontal plane
    'tilt': 30, #angle of solar panels with horizontal
    'albedo': 0.075, #albedo (fraction of sun light reflected) of ocean water (https://geoengineering.global/ocean-albedo-modification/)
    'temperature_model_parameters': temp_params,
}


pv_system = Photovoltaic(**system_data)

#getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
pvlib_df = era5.weather_df_from_era5(     
    era5_netcdf_filename='ERA5_weather_data_location4.nc',
    lib='pvlib', area=area)

#determining the zenith angle (angle of sun position with vertical in vertical plane) in the specified locations for the time instances downloaded from ERA5
zenithcalc = pvlib.solarposition.get_solarposition(time=pvlib_df.index,latitude=latitude, longitude=longitude, altitude=None, pressure=None, method='nrel_numpy', temperature=pvlib_df['temp_air'])

#determining DNI from GHI, DHI and zenith angle for the time instances downloaded from ERA5
dni = pvlib.irradiance.dni(pvlib_df['ghi'],pvlib_df['dhi'], zenith=zenithcalc['zenith'], clearsky_dni=None, clearsky_tolerance=1.1, zenith_threshold_for_zero_dni=88.0, zenith_threshold_for_clearsky_limit=80.0)

#adding DNI to dataframe with PV weather data
pvlib_df['dni'] = dni

#replacing 'NAN' in DNI column with 0 to prevent 'holes' in graphs (NANs are caused by the zenith angle being larger than the 'zenith threshold for zero dni' angle set with GHI and DHI not yet being zero. The DNI should be 0 in that case)
pvlib_df['dni'] = pvlib_df['dni'].fillna(0)

#plotting dhi and ghi
plt.figure
pvlib_df.loc[:, ['dhi', 'ghi']].plot(title='Irradiance')
plt.xlabel('Time')
plt.ylabel('Irradiance in $W/m^2$');

#determining PV power generation
feedin = pv_system.feedin(
    weather=pvlib_df,
    location=(latitude, longitude))

#plotting PV power generation
plt.figure()
feedin.plot(title='PV feed-in')
plt.xlabel('Time')
plt.ylabel('Power in W');

#calculating total PV power generated in selected period
PVpower = pd.Series.sum(feedin) #power produced over time period in Wh

print('One panel with the chosen input parameters in location', latitude,',',longitude, 'will produce',PVpower,'Wh of electricity between', start_date, 'and', end_date )



feedinarray = multsolar*feedin.values #hourly energy production over the year of 1 solar platform of the specified kind in the specified location
feedinarray[feedinarray<0]=0

#%%

# --- Wind Energy Generation ---

# get power curves
# get names of wind turbines for which power curves and/or are provided
# set print_out=True to see the list of all available wind turbines
df = get_turbine_types(print_out=False)

#defining the wind turbine system
turbine_data= {
    "turbine_type": "E-101/3050",  # turbine type as in register
    "hub_height": 130,  # in m
}
my_turbine = WindTurbine(**turbine_data)

#getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
windpowerlib_df = era5.weather_df_from_era5(
    era5_netcdf_filename='ERA5_weather_data_location4.nc',
    lib='windpowerlib', area=area)

#increasing the time indices by half an hour because then they match the pvlib time indices so the produced energy can be added later
#assumed wind speeds half an hour later are similar and this will not affect the results significantly
windpowerlib_df.index = windpowerlib_df.index + timedelta(minutes=30)

# power output calculation for e126

# own specifications for ModelChain setup
modelchain_data = {
    'wind_speed_model': 'logarithmic',      # 'logarithmic' (default),
                                            # 'hellman' or
                                            # 'interpolation_extrapolation'
    'density_model': 'ideal_gas',           # 'barometric' (default), 'ideal_gas'
                                            #  or 'interpolation_extrapolation'
    'temperature_model': 'linear_gradient', # 'linear_gradient' (def.) or
                                            # 'interpolation_extrapolation'
    'power_output_model': 'power_curve',    # 'power_curve' (default) or
                                            # 'power_coefficient_curve'
    'density_correction': True,             # False (default) or True
    'obstacle_height': 0,                   # default: 0
    'hellman_exp': None}                    # None (default) or None

# initialize ModelChain with own specifications and use run_model method to
# calculate power output
mc_my_turbine = ModelChain(my_turbine, **modelchain_data).run_model(windpowerlib_df)
# write power output time series to WindTurbine object
my_turbine.power_output = mc_my_turbine.power_output


# plot turbine power output
plt.figure()
my_turbine.power_output.plot(title='Wind turbine power production')
plt.xlabel('Time')
plt.ylabel('Power in W')
plt.show()

my_turbinearray = multwind*my_turbine.power_output.values #hourly energy production over the year of 1 wind turbine of the specified kind in the specified location

# %% Model parameters and sets
# Set model name
model = Model('GFPSO Cost Optimization')

# ---- Sets ----

I = [1, 2] # Conv devices (1: Conv, 2: Reconv)
J = [1, 2] # Energy Medium (1: Ammonia, 2: LiquidH2)
K = [1, 2, 3, 4] # Device types (1: Wind, 2: Solar, 3: Elec, 4: Desal)
L = Nsteps # Years in time period
N = [1, 2] # Volume based equipment (1: Storage, 2: FPSO)
T = list(range(0,7320)) # Operational hours in year

# If medium choice is done as parameter
E = 0


# ---- Parameters ----

# Cost parameters

Cconvammonia = 10/12*np.array([float(cell.value) for cell in ammonia[48][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
Cconvliquid = 10/12*np.array([float(cell.value) for cell in liquidhydrogen[61][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years
Creconvammonia = 10/12*np.array([float(cell.value) for cell in ammonia[111][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
Creconvliquid = 10/12*np.array([float(cell.value) for cell in liquidhydrogen[125][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years

# A_lij , l = year in time period, i = conversion device type, j = medium
A = []
for l in range(Nsteps):
    Aarray = np.array([[Cconvammonia[l], Cconvliquid[l]], [Creconvammonia[l], Creconvliquid[l]]])
    A.append(Aarray)

Cs1 = 10/12*np.array([float(cell.value) for cell in solar[51][2:2+Nsteps]])
Cw1 = 10/12*np.array([float(cell.value) for cell in wind[48][2:2+Nsteps]])
Ce = 10/12*np.array([float(cell.value) for cell in electrolyzer[50][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an electrolyzer in 10^3 euros over several years
Cd = 10/12*np.array([float(cell.value) for cell in desalination[49][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a desalination installation in 10^3 euros over several years

# B_lk , l = year in time period, k = device type (solar, wind, elec, desal)
B = []
for l in range(Nsteps):
    Barray = np.array([Cw1[l], 
                       Cs1[l], 
                       Ce[l], 
                       Cd[l]])
    B.append(Barray)

# C_ln
Cstliquid =  10/12*np.array([float(cell.value) for cell in storage[25][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of storage per m3 in euros over several years
Cstammonia =  10/12*np.array([float(cell.value) for cell in storage[54][2:2+Nsteps]])
Cfpso = 10/12* np.array([float(cell.value) for cell in fpso[53][2:2+Nsteps]]) #np.array([1,1,0.8,0.6,0.4]) #cost per year (depreciation+OPEX) of FPSO per m3 in 10^3 euros over several years
Cst = Cstammonia

C = []
for l in range(Nsteps):
    Carray = np.array([Cst[l], 
                       Cfpso[l]])
    C.append(Carray)

# D , Yearly demand
demand = 10/12*general.cell(row=21, column=2).value #demand in tons of hydrogen per year
# D = demand
D = (10/12)*50000 # [tonH2/yr]\

# TC , Transport cost

coords_production = (general.cell(row=15, column=4).value, general.cell(row=15, column=5).value) #coordinates (latitude and longitude) of production location
coords_port = (general.cell(row=16, column=4).value, general.cell(row=16, column=5).value) #coordinates (latitude and longitude) of port
coords_demand = (general.cell(row=17, column=4).value, general.cell(row=17, column=5).value)  #coordinates (latitude and longitude) of demand location
distanceseafactor = general.cell(row=25, column=2).value
distancesea = distanceseafactor*geopy.distance.geodesic(coords_production, coords_port).km #distance to be travelled over sea from production to demand location in km
distanceland = geopy.distance.geodesic(coords_port, coords_demand).km #distance to be travelled over land from production to demand location in km

Xtransport = demand #amount to be transported by ship in tons of hydrogen, assumed to be equal to demand as hydrogen losses are almost zero 
Xkmsea = Xtransport*distancesea #yearly amount of tonkm to be made over sea
Xkmland = demand*distanceland #yealy amount of tonkm over land

Ckmammonia = np.array([float(cell.value) for cell in ammonia[158][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as ammonia in 10^3 euros over several years
Cbasetransportammonia = np.array([float(cell.value) for cell in ammonia[159][2:2+Nsteps]])
Ckmliquid = np.array([float(cell.value) for cell in liquidhydrogen[168][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as liquid hydrogen in 10^3 euros over several years
Cbasetransportliquid =  np.array([float(cell.value) for cell in liquidhydrogen[169][2:2+Nsteps]])
Ckmland = np.array([float(cell.value) for cell in landtransport[33][2:2+Nsteps]]) #costs per year per km of overland pipeline for 1 ton of hydrogen

Ckmsea = [Ckmammonia, Ckmliquid] #costs per km of overseas transport, depending on whether ammonia or liquid hydrogen is chosen
Cbasetransport = [Cbasetransportammonia, Cbasetransportliquid] #baserate of the transport per ton hydrogen

# TC_jl
TC = []
for j in J:
    TCyear = Xtransport*Cbasetransport[j-1] + Xkmsea*Ckmsea[j-1] + Xkmland*Ckmland
    TC.append(TCyear)

BigM = 1000000000

# Non-cost parameters
fracpowerelectrolyzerliquid = electrolyzer.cell(row=11, column=2).value #fraction of energy used by electrolyzers when using liquid hydrogen
fracpowerelectrolyzerammonia = electrolyzer.cell(row=12, column=2).value #fraction of energy used by elecyrolyzers when using ammonia
alpha = [fracpowerelectrolyzerammonia, fracpowerelectrolyzerliquid]

capelectrolyzerhour = electrolyzer.cell(row=9, column=2).value #hourly output capacity of electrolyzers in tons of hydrogen per hour
beta = capelectrolyzerhour

electrolyzer_energy = 1000*electrolyzer.cell(row=7, column=2).value #energy requirement of electrolyzer in Wh per ton of hydrogen
gamma = electrolyzer_energy

electrolyzer_water = electrolyzer.cell(row=6, column=2).value #water requirement of electrolyzer in m3 of water per ton of hydrogen
capdesalinationhour = desalination.cell(row=14, column=2).value #hourly output capacity desalination in m3 of water per hour
epsilon = electrolyzer_water/capdesalinationhour

eta_conversionammonia = ammonia.cell(row=13, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
eta_conversionliquid =  liquidhydrogen.cell(row=16, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
eta_reconversionammonia = eta_conversionammonia #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
eta_reconversionliquid = eta_conversionliquid #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
# eta_ij , i = conv device type, j = energy medium
eta = 1/np.array([[eta_conversionammonia, eta_conversionliquid],
                 [eta_reconversionammonia, eta_reconversionliquid]])

volumefpsoliquid = electrolyzer.cell(row=15, column=2).value #FPSO volume per electrolyzer liquid hydrogen
volumefpsoammonia = electrolyzer.cell(row=16, column=2).value #FPSO volume per electrolyzer ammonia
nu = [volumefpsoammonia, volumefpsoliquid]

ratiostoragefpsoliquid = fpso.cell(row=18, column=2).value #ratio storage tanks in m3/fpso volume in m3
ratiostoragefpsoammonia = fpso.cell(row=19, column=2).value #ratio storage tanks in m3/fpso volume in m3
phi = [ratiostoragefpsoammonia, ratiostoragefpsoliquid]

capconvammonia = 10/12*ammonia.cell(row=10, column=2).value #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for ammonia
capconvliquid = 10/12*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for liquid hydrogen
capreconvammonia = 10/12*ammonia.cell(row=75, column=2).value #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for ammonia
capreconvliquid = 10/12*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for liquid hydrogen

# Conversion
w11 = math.ceil(1.6*D/(eta[1][0]*capconvammonia))  # Ammonia
w12 = 1.6*D/(eta[1][1]*capconvliquid)   # Liquid
# Reconversion
w21 = D/capconvammonia # Ammonia
w22 = D/capconvammonia # Liquid
W = [[w11, w12], [w21, w22]]


#  For now the converters and reconverter amount is constant
WC = quicksum(W[i-1][E]*A[l][i-1][E] for i in I) 

#%%
# --- Variables ---

# # w[i,j]
# w = {}
# for i in I:
#     for j in J:
#         w[i,j] = model.addVar (lb = 0, vtype = GRB.INTEGER, name = 'w[' + str(i) + ',' + str(j) + ']' )
# x[k]
x = {}
for k in K:
    x[k] = model.addVar (lb = 0, vtype = GRB.INTEGER, name = 'x[' + str(k) + ']' )
    
# y[n]
y = {}
for n in N:
    y[n] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'y[' + str(n) + ']' )

# E
# E = model.addVar (vtype = GRB.BINARY, name = 'E' )
# For now just setting the medium to Ammonia (E=0)
E = 0

# For now just setting T = [0]
T = list(range(0,7320))

PU = {}
for t in T:
    PU[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PU[' + str(t) + ']' )
    
# test for more constraints
PA = {}
PG = {}
h = {}
h = model.addVars(len(feedinarray),vtype=GRB.CONTINUOUS,name='h')
for t in T:
    PG[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PG[' + str(t) + ']' )
    PA[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PA[' + str(t) + ']' )
    # h[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'h[' + str(t) + ']' )

# %%  ---- Integrate new variables ----
model.update()

# ---- Objective Function ----
# For now just setting year to first year in time period (l=0)
l = 0
model.setObjective (WC + quicksum(x[k]*B[l][k-1] for k in K) + quicksum(y[n]*C[l][n-1] for n in N) + TC[l][E])
model.modelSense = GRB.MINIMIZE
model.update ()

# %% Constraints

# constraint for power used to power generated
# power_used_g = {}

# for t in T:
#     power_used_g[t] = model.addConstr(PU[t] <= alpha[E]*(my_turbinearray[t]*x[1] + feedinarray[t]*x[2]))
    
# power_used_e = {}
# for t in T:
#     power_used_g[t] = model.addConstr(PU[t] <= beta*gamma*x[3])

# model.addConstr(quicksum(PU[t] for t in T) >= (gamma*D)/(eta[0][E]*eta[1][E]))


model.addConstr(x[4] >= x[3]*beta*epsilon)

model.addConstr(y[2] == x[3]*nu[E])

model.addConstr(y[1] == y[2]*phi[E])

# test by adding more constraints

for t in T:
    model.addConstr(PG[t] == my_turbinearray[t]*x[1] + feedinarray[t]*x[2])
    model.addConstr(PA[t] == alpha[E]*PG[t])
    model.addConstr(PU[t] <= alpha[E]*PG[t])
    model.addConstr(PU[t] <= beta*gamma*x[3])
    model.addConstr(h[t] == PU[t]/gamma)
    
model.addConstr(h.sum() >= D/(eta[0][E]*eta[1][E]))
    
    
    
    

# for i in I:
#     model.addConstr(w[i,1] <= BigM*E)
    
# for i in I:
#     model.addConstr(w[i,2] <= BigM*(1-E))

#%%
model.update()
model.Params.NumericFocus = 1
model.Params.timeLimit = 400
model.optimize()
    