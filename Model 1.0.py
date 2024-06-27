# -*- coding: utf-8 -*-
"""
Created on Thu May  2 11:41:52 2024

@author: Piotr Kaczmarek
"""

#%% --- Preamble ---

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import geopy.distance
import math
from openpyxl import load_workbook
from feedinlib import era5
import pvlib
from feedinlib import Photovoltaic
from feedinlib import get_power_plant_data
from windpowerlib.modelchain import ModelChain
from windpowerlib.wind_turbine import WindTurbine
from windpowerlib import get_turbine_types
from datetime import timedelta
from gurobipy import *
import plotly.express as px
import plotly.io as pio
import plotly.graph_objects as go
from sklearn.linear_model import LinearRegression


import warnings

def fxn():
    warnings.warn("deprecated", DeprecationWarning)

with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    fxn()

# # Or if you are using > Python 3.11:
# with warnings.catch_warnings(action="ignore"):
#     fxn()
pd.set_option('future.no_silent_downcasting', True)


#%% --- Excel Parameters ---

#loading excel to later retrieve input data using openpyxl
data_file = os.getcwd() 
data_file = os.path.dirname(os.path.realpath('__file__')) + '\Inputdata econ.xlsx'
wb = load_workbook(data_file,data_only=True) # creating workbook

general = wb['General'] #selecting data from sheet called general
wind = wb['Wind'] #selecting data from sheet called wind
solar = wb['Solar'] #selecting data from sheet called solar
electrolyzer = wb['Electrolyzer'] #selecting data from sheet called electrolyzer
desalination = wb['Desalination'] #selecting data from sheet called desalination
ammonia = wb['Ammonia'] #selecting data from sheet called ammonia
liquidhydrogen = wb['Liquid hydrogen'] #selecting data from sheet called liquid hydrogen
landtransport = wb['Land transport'] #selecting data from sheet called landtransport
storage = wb['Storage'] #selecting data from sheet called storage
fpso = wb['FPSO'] #selecting data from sheet called fpso

# Time period data
Startyear = general.cell(row=9, column=2).value #first year to be reviewed
timeperiod = general.cell(row=10, column=2).value #time period of simulation in years
timestep = general.cell(row=11, column=2).value #time step of simulation in years
Nsteps = int(timeperiod/timestep+1) #number of time steps (important for data selection from excel and loop at the end)

# Coordinates production location

# # Tokyo
# latitude =  26.81
# longitude = 126.94

# North-Sea case
latitude = 54.35
longitude = 6.28
#  Eemshaven (Groningen)
coords_demand = (53.43, 6.84) 
coords_port = (53.43, 6.84) 
# #  Port of Rotterdam
# coords_demand = (51.95, 4.14) 

# Solar and Wind 
multsolar = 524*1000/206.7989   # Installed capacity per unit [kWp] * 1000 (to make it MW) / capacity from PVlib [kWp]
multwind = 12/3   # Capacity per unit [MW] / capacity from Windlib [MW]

# Set to the time period ratio of the CDS dataset considered (1/12 means 1 month)
DataYearRatio = 1/12

# Demand parameter (needed for now for transport calculation)
demand = DataYearRatio*general.cell(row=21, column=2).value #demand in tons of hydrogen per year

#%% Excel Parameter replacement WIND

#### Input params

# Startyear = 2047
# Nsteps = 2
# timestep = 3
learning_rate_2035 = 40/115
learning_rate_2050 = 0.5
Capacity_Wind = 12 
# CAPEX [Eur/kW] in 2020
Development = 212
Turbine = 1060
Plant_Balance = 350
Decommission = 44
Total_CAPEX = Development + Turbine + Plant_Balance + Decommission
Total_CAPEX_MW = Total_CAPEX * 1000
# OPEX [Eur/MW] in 2020
Avg_OPEX = 135000
r = 0.08 # Interest rate
lftm = 25


#### Computation

# #  TYCHOS NUMBERS (FLOATING)
# # Year
# Year = np.array([2025, 2035, 2050])
# # CAPEX
# CAPEX = np.array([4022700*Capacity_Wind, learning_rate_2035*4022700*Capacity_Wind, learning_rate_2035*learning_rate_2050*4022700*Capacity_Wind])
# # OPEX
# OPEX = np.array([979800, 340800, 170400])

Year = np.array([2020, 2035, 2050])
CAPEX = np.array([Total_CAPEX_MW*Capacity_Wind, Total_CAPEX_MW*Capacity_Wind*learning_rate_2035, Total_CAPEX_MW*Capacity_Wind*learning_rate_2035*learning_rate_2050])
OPEX = np.array([Avg_OPEX*Capacity_Wind, Avg_OPEX*Capacity_Wind*learning_rate_2035, Avg_OPEX*Capacity_Wind*learning_rate_2035*learning_rate_2050])

a = (r*(1+r)**lftm) / ((1+r)**lftm - 1) # Should this really be constant?

Wind_Costs = np.zeros((5,31))

yearstep = 0
for i in range(31):
    Wind_Costs[0][i] = 2020 + yearstep
    yearstep += 1
    
for i in range(31):
    Wind_Costs[2][i] = lftm
    
# 2020-2035 CAPEX
X = Year[0:2].reshape(-1, 1)
y = CAPEX[0:2]              
model = LinearRegression()
model.fit(X, y)
X_predict = Wind_Costs[0][0:16].reshape(-1, 1) # put the dates of which you want to predict kwh here
Wind_Costs[1][0:16] = model.predict(X_predict)
# 2035-2050 CAPEX
X = Year[1:3].reshape(-1, 1)
y = CAPEX[1:3]
model = LinearRegression()
model.fit(X, y)
X_predict = Wind_Costs[0][16:31].reshape(-1, 1) # put the dates of which you want to predict kwh here
Wind_Costs[1][16:31] = model.predict(X_predict)
# 2020-2035 OPEX
X = Year[0:2].reshape(-1, 1)
y = OPEX[0:2]
model = LinearRegression()
model.fit(X, y)
X_predict = Wind_Costs[0][0:16].reshape(-1, 1) # put the dates of which you want to predict kwh here
Wind_Costs[3][0:16] = model.predict(X_predict)
# 2035-2050 OPEX
X = Year[1:3].reshape(-1, 1)
y = OPEX[1:3]
model = LinearRegression()
model.fit(X, y)
X_predict = Wind_Costs[0][16:31].reshape(-1, 1) # put the dates of which you want to predict kwh here
Wind_Costs[3][16:31] = model.predict(X_predict)

Wind_Costs[4] = a*Wind_Costs[1] + Wind_Costs[3]

def rel_wind_array(SY, NS, TS, WiCo):
    index = np.where(Wind_Costs[0] == SY)[0][0]
    relevant_array = np.zeros(NS)
    for i in range(NS):    
        relevant_array[i] = WiCo[4][index+i*TS]
    
    return relevant_array



#%% --- Solar Function (and ERA5 setup) ---

# set start and end date (end date will be included
# in the time period for which data is downloaded)
start_date, end_date = '2020-01-01', '2020-01-30'  #Test data of 1 month
# start_date, end_date = '2020-01-01', '2020-10-31'
# set variable set to download
variable = 'feedinlib'

# era5_netcdf_filename = 'Era 5 test data\ERA5_weather_data_test_RegTokyo_Corrected.nc' #referring to file with weather data downloaded earlier using the ERA5 API
# era5_netcdf_filename = 'ERA5_weather_data_location4.nc' #referring to file with weather data downloaded earlier using the ERA5 API
era5_netcdf_filename = 'Era 5 test data\ERA5_weather_data_1month_RegNorthSea.nc' #North-Sea Case data


area = [longitude, latitude]
    

def func_PV(location):
    
    # get modules
    module_df = get_power_plant_data(dataset='SandiaMod') #retrieving dataset for PV modules
    
    # get inverter data
    inverter_df = get_power_plant_data(dataset='cecinverter') #retrieving dataset for inverters
    
    temp_params = pvlib.temperature.TEMPERATURE_MODEL_PARAMETERS['sapm']['open_rack_glass_polymer'] #defining temperature model parameters (leaving at default causes errors)
    
    #PV system definition
    system_data = {
        'module_name': 'Advent_Solar_Ventura_210___2008_',  # module name as in database
        'inverter_name': 'ABB__MICRO_0_25_I_OUTD_US_208__208V_',  # inverter name as in database
        'azimuth': 180, #angle of sun position with north in horizontal plane
        'tilt': 30, #angle of solar panels with horizontal
        'albedo': 0.075, #albedo (fraction of sun light reflected) of ocean water (https://geoengineering.global/ocean-albedo-modification/)
        'temperature_model_parameters': temp_params,
    }
    
    pv_system = Photovoltaic(**system_data)
    
    #getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
    pvlib_df = era5.weather_df_from_era5(     
        era5_netcdf_filename=era5_netcdf_filename,
        lib='pvlib', area=location)
    
    #determining the zenith angle (angle of sun position with vertical in vertical plane) in the specified locations for the time instances downloaded from ERA5
    zenithcalc = pvlib.solarposition.get_solarposition(time=pvlib_df.index,latitude=latitude, longitude=longitude, altitude=None, pressure=None, method='nrel_numpy', temperature=pvlib_df['temp_air'])
    
    #determining DNI from GHI, DHI and zenith angle for the time instances downloaded from ERA5
    dni = pvlib.irradiance.dni(pvlib_df['ghi'],pvlib_df['dhi'], zenith=zenithcalc['zenith'], clearsky_dni=None, clearsky_tolerance=1.1, zenith_threshold_for_zero_dni=88.0, zenith_threshold_for_clearsky_limit=80.0)
    
    #adding DNI to dataframe with PV weather data
    pvlib_df['dni'] = dni
    
    #replacing 'NAN' in DNI column with 0 to prevent 'holes' in graphs (NANs are caused by the zenith angle being larger than the 'zenith threshold for zero dni' angle set with GHI and DHI not yet being zero. The DNI should be 0 in that case)
    pvlib_df['dni'] = pvlib_df['dni'].fillna(0)
    
    #determining PV power generation
    feedin = pv_system.feedin(
        weather=pvlib_df,
        location=(location[1], location[0]))
    
    feedinarray = multsolar*feedin.values #hourly energy production over the year of 1 solar platform of the specified kind in the specified location
    feedinarray[feedinarray<0]=0
    
    # #plotting PV power generation
    # plt.figure()
    # feedin.plot(title='PV feed-in')
    # plt.xlabel('Time')
    # plt.ylabel('Power in W');
    
    feedin_index = feedin.index
    
    return feedinarray    

# feedinarray = func_PV(area)

#%% Feedin Index function

def func_Feedin_index(location):
    
    # get modules
    module_df = get_power_plant_data(dataset='SandiaMod') #retrieving dataset for PV modules
    
    # get inverter data
    inverter_df = get_power_plant_data(dataset='cecinverter') #retrieving dataset for inverters
    
    temp_params = pvlib.temperature.TEMPERATURE_MODEL_PARAMETERS['sapm']['open_rack_glass_polymer'] #defining temperature model parameters (leaving at default causes errors)
    
    #PV system definition
    system_data = {
        'module_name': 'Advent_Solar_Ventura_210___2008_',  # module name as in database
        'inverter_name': 'ABB__MICRO_0_25_I_OUTD_US_208__208V_',  # inverter name as in database
        'azimuth': 180, #angle of sun position with north in horizontal plane
        'tilt': 30, #angle of solar panels with horizontal
        'albedo': 0.075, #albedo (fraction of sun light reflected) of ocean water (https://geoengineering.global/ocean-albedo-modification/)
        'temperature_model_parameters': temp_params,
    }
    
    pv_system = Photovoltaic(**system_data)
    
    #getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
    pvlib_df = era5.weather_df_from_era5(     
        era5_netcdf_filename=era5_netcdf_filename,
        lib='pvlib', area=location)
    
    #determining the zenith angle (angle of sun position with vertical in vertical plane) in the specified locations for the time instances downloaded from ERA5
    zenithcalc = pvlib.solarposition.get_solarposition(time=pvlib_df.index,latitude=latitude, longitude=longitude, altitude=None, pressure=None, method='nrel_numpy', temperature=pvlib_df['temp_air'])
    
    #determining DNI from GHI, DHI and zenith angle for the time instances downloaded from ERA5
    dni = pvlib.irradiance.dni(pvlib_df['ghi'],pvlib_df['dhi'], zenith=zenithcalc['zenith'], clearsky_dni=None, clearsky_tolerance=1.1, zenith_threshold_for_zero_dni=88.0, zenith_threshold_for_clearsky_limit=80.0)
    
    #adding DNI to dataframe with PV weather data
    pvlib_df['dni'] = dni
    
    #replacing 'NAN' in DNI column with 0 to prevent 'holes' in graphs (NANs are caused by the zenith angle being larger than the 'zenith threshold for zero dni' angle set with GHI and DHI not yet being zero. The DNI should be 0 in that case)
    pvlib_df['dni'] = pvlib_df['dni'].fillna(0)
    
    #determining PV power generation
    feedin = pv_system.feedin(
        weather=pvlib_df,
        location=(location[1], location[0]))
    
    feedinarray = multsolar*feedin.values #hourly energy production over the year of 1 solar platform of the specified kind in the specified location
    feedinarray[feedinarray<0]=0
    
    # #plotting PV power generation
    # plt.figure()
    # feedin.plot(title='PV feed-in')
    # plt.xlabel('Time')
    # plt.ylabel('Power in W');
    
    feedin_index = feedin.index
    
    return feedin_index    
#%% --- Wind function ---

def func_Wind(location):
    
    df = get_turbine_types(print_out=False)
    
    #defining the wind turbine system
    turbine_data= {
        "turbine_type": "E-101/3050",  # turbine type as in register
        "hub_height": 130,  # in m
    }
    my_turbine = WindTurbine(**turbine_data)
    
    #getting the needed weather data for wind power calculations from the file as downloaded with a seperate script from ERA5
    windpowerlib_df = era5.weather_df_from_era5(
        era5_netcdf_filename=era5_netcdf_filename,
        lib='windpowerlib', area=list(location))
    
    #increasing the time indices by half an hour because then they match the pvlib time indices so the produced energy can be added later
    #assumed wind speeds half an hour later are similar and this will not affect the results significantly
    windpowerlib_df.index = windpowerlib_df.index + timedelta(minutes=30)
    
    # power output calculation for e126
    
    # own specifications for ModelChain setup
    modelchain_data = {
        'wind_speed_model': 'logarithmic',      # 'logarithmic' (default),
                                                # 'hellman' or
                                                # 'interpolation_extrapolation'
        'density_model': 'ideal_gas',           # 'barometric' (default), 'ideal_gas'
                                                #  or 'interpolation_extrapolation'
        'temperature_model': 'linear_gradient', # 'linear_gradient' (def.) or
                                                # 'interpolation_extrapolation'
        'power_output_model': 'power_curve',    # 'power_curve' (default) or
                                                # 'power_coefficient_curve'
        'density_correction': True,             # False (default) or True
        'obstacle_height': 0,                   # default: 0
        'hellman_exp': None}                    # None (default) or None
    
    # initialize ModelChain with own specifications and use run_model method to
    # calculate power output
    mc_my_turbine = ModelChain(my_turbine, **modelchain_data).run_model(windpowerlib_df)
    # write power output time series to WindTurbine object
    my_turbine.power_output = mc_my_turbine.power_output
    
    # title = 'Wind turbine power production at ' + str(location)
    # # plot turbine power output
    # plt.figure()
    # my_turbine.power_output.plot(title=title)
    # plt.xlabel('Time')
    # plt.ylabel('Power in W')
    # plt.show()
    
    my_turbinearray = multwind*my_turbine.power_output.values #hourly energy production over the year of 1 wind turbine of the specified kind in the specified location
    
    return my_turbinearray


        #%% --- Model parameters and sets ---
        
        # Set model name
        model = Model('GFPSO Cost Optimization')
        
        # ---- Sets ----
        
        I = [1, 2] # Conv devices (1: Conv, 2: Reconv)
        J = [1, 2] # Energy Medium (1: Ammonia, 2: LiquidH2)
        K = [1, 2, 3, 4] # Device types (1: Wind, 2: Solar, 3: Elec, 4: Desal)
        L = range(Nsteps) # Years in time period
        N = [1, 2] # Volume based equipment (1: Storage, 2: FPSO)
        T = list(range(0,len(feedinarray))) # Operational hours in year
        
        # If medium choice is done as parameter
        E = 0
        
        
        # ---- Parameters ----
        
        # Cost parameters
        
        Cconvammonia = DataYearRatio*np.array([float(cell.value) for cell in ammonia[48][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
        Cconvliquid = DataYearRatio*np.array([float(cell.value) for cell in liquidhydrogen[61][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years
        Creconvammonia = DataYearRatio*np.array([float(cell.value) for cell in ammonia[111][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
        Creconvliquid = DataYearRatio*np.array([float(cell.value) for cell in liquidhydrogen[125][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years
        
        # A_lij , l = year in time period, i = conversion device type, j = medium
        A = []
        for l in range(Nsteps):
            Aarray = np.array([[Cconvammonia[l], Cconvliquid[l]], [Creconvammonia[l], Creconvliquid[l]]])
            A.append(Aarray)
        
        Cs1 = DataYearRatio*np.array([float(cell.value) for cell in solar[51][2:2+Nsteps]])
        # Cw1 = DataYearRatio*np.array([float(cell.value) for cell in wind[48][2:2+Nsteps]])
        Cw1 = DataYearRatio*rel_wind_array(Startyear, Nsteps, timestep, Wind_Costs)
        Ce = DataYearRatio*np.array([float(cell.value) for cell in electrolyzer[50][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an electrolyzer in 10^3 euros over several years
        Cd = DataYearRatio*np.array([float(cell.value) for cell in desalination[49][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a desalination installation in 10^3 euros over several years
        
        # B_lk , l = year in time period, k = device type (solar, wind, elec, desal)
        B = []
        for l in L:
            Barray = np.array([Cw1[l], 
                               Cs1[l], 
                               Ce[l], 
                               Cd[l]])
            B.append(Barray)
        
        # C_ln
        Cstliquid =  DataYearRatio*np.array([float(cell.value) for cell in storage[25][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of storage per m3 in euros over several years
        Cstammonia =  DataYearRatio*np.array([float(cell.value) for cell in storage[54][2:2+Nsteps]])
        Cfpso = DataYearRatio* np.array([float(cell.value) for cell in fpso[53][2:2+Nsteps]]) #np.array([1,1,0.8,0.6,0.4]) #cost per year (depreciation+OPEX) of FPSO per m3 in 10^3 euros over several years
        Cst = Cstammonia
        
        C = []
        for l in L:
            Carray = np.array([Cst[l], 
                               Cfpso[l]])
            C.append(Carray)
        
        # D , Yearly demand
        demand = DataYearRatio*general.cell(row=21, column=2).value #demand in tons of hydrogen per year
        # D = demand
        D = (DataYearRatio)*50000 # [tonH2/yr]\
        
        # --------
        # # TC , Transport cost
        
        # coords_production = (general.cell(row=15, column=4).value, general.cell(row=15, column=5).value) #coordinates (latitude and longitude) of production location
        # coords_port = (general.cell(row=16, column=4).value, general.cell(row=16, column=5).value) #coordinates (latitude and longitude) of port
        # coords_demand = (general.cell(row=17, column=4).value, general.cell(row=17, column=5).value)  #coordinates (latitude and longitude) of demand location
        # distanceseafactor = general.cell(row=25, column=2).value
        # distancesea = distanceseafactor*geopy.distance.geodesic(coords_production, coords_port).km #distance to be travelled over sea from production to demand location in km
        # distanceland = geopy.distance.geodesic(coords_port, coords_demand).km #distance to be travelled over land from production to demand location in km
        
        # Xtransport = demand #amount to be transported by ship in tons of hydrogen, assumed to be equal to demand as hydrogen losses are almost zero 
        # Xkmsea = Xtransport*distancesea #yearly amount of tonkm to be made over sea
        # Xkmland = demand*distanceland #yealy amount of tonkm over land
        
        # Ckmammonia = np.array([float(cell.value) for cell in ammonia[158][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as ammonia in 10^3 euros over several years
        # Cbasetransportammonia = np.array([float(cell.value) for cell in ammonia[159][2:2+Nsteps]])
        # Ckmliquid = np.array([float(cell.value) for cell in liquidhydrogen[168][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as liquid hydrogen in 10^3 euros over several years
        # Cbasetransportliquid =  np.array([float(cell.value) for cell in liquidhydrogen[169][2:2+Nsteps]])
        # Ckmland = np.array([float(cell.value) for cell in landtransport[33][2:2+Nsteps]]) #costs per year per km of overland pipeline for 1 ton of hydrogen
        
        # Ckmsea = [Ckmammonia, Ckmliquid] #costs per km of overseas transport, depending on whether ammonia or liquid hydrogen is chosen
        # Cbasetransport = [Cbasetransportammonia, Cbasetransportliquid] #baserate of the transport per ton hydrogen
        
        # # # LocTC is a list of the transport costs at the different locations in region
        # # LocTC = []
        # # for Loc in loclist:
        # #     Loc = tuple(Loc)
        # #     distancesea = distanceseafactor*geopy.distance.geodesic(Loc, coords_port).km
        # #     Xkmsea = Xtransport*distancesea
            
        # # TC_jl
        # TC = []
        # for j in J:
        #     TCyear = Xtransport*Cbasetransport[j-1] + Xkmsea*Ckmsea[j-1] + Xkmland*Ckmland
        #     TC.append(TCyear)
        # ---------
        
        # LocTC.append(TC)
            
        # # TC_jl
        # TC = []
        # for j in J:
        #     TCyear = Xtransport*Cbasetransport[j-1] + Xkmsea*Ckmsea[j-1] + Xkmland*Ckmland
        #     TC.append(TCyear)
            
        
        
        # Non-cost parameters
        fracpowerelectrolyzerliquid = electrolyzer.cell(row=11, column=2).value #fraction of energy used by electrolyzers when using liquid hydrogen
        fracpowerelectrolyzerammonia = electrolyzer.cell(row=12, column=2).value #fraction of energy used by elecyrolyzers when using ammonia
        alpha = [fracpowerelectrolyzerammonia, fracpowerelectrolyzerliquid]
        
        capelectrolyzerhour = electrolyzer.cell(row=9, column=2).value #hourly output capacity of electrolyzers in tons of hydrogen per hour
        beta = capelectrolyzerhour
        
        electrolyzer_energy = 1000*electrolyzer.cell(row=7, column=2).value #energy requirement of electrolyzer in Wh per ton of hydrogen
        gamma = electrolyzer_energy
        
        electrolyzer_water = electrolyzer.cell(row=6, column=2).value #water requirement of electrolyzer in m3 of water per ton of hydrogen
        capdesalinationhour = desalination.cell(row=14, column=2).value #hourly output capacity desalination in m3 of water per hour
        epsilon = electrolyzer_water/capdesalinationhour
        
        eta_conversionammonia = ammonia.cell(row=13, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
        eta_conversionliquid =  liquidhydrogen.cell(row=16, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
        eta_reconversionammonia = eta_conversionammonia #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
        eta_reconversionliquid = eta_conversionliquid #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
        # eta_ij , i = conv device type, j = energy medium
        eta = 1/np.array([[eta_conversionammonia, eta_conversionliquid],
                         [eta_reconversionammonia, eta_reconversionliquid]])
        
        volumefpsoliquid = electrolyzer.cell(row=15, column=2).value #FPSO volume per electrolyzer liquid hydrogen
        volumefpsoammonia = electrolyzer.cell(row=16, column=2).value #FPSO volume per electrolyzer ammonia
        nu = [volumefpsoammonia, volumefpsoliquid]
        
        ratiostoragefpsoliquid = fpso.cell(row=18, column=2).value #ratio storage tanks in m3/fpso volume in m3
        ratiostoragefpsoammonia = fpso.cell(row=19, column=2).value #ratio storage tanks in m3/fpso volume in m3
        phi = [ratiostoragefpsoammonia, ratiostoragefpsoliquid]
        
        capconvammonia = DataYearRatio*ammonia.cell(row=10, column=2).value #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for ammonia
        capconvliquid = DataYearRatio*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for liquid hydrogen
        capreconvammonia = DataYearRatio*ammonia.cell(row=75, column=2).value #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for ammonia
        capreconvliquid = DataYearRatio*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for liquid hydrogen
        
        # Conversion
        w11 = math.ceil(1.6*D/(eta[1][0]*capconvammonia))  # Ammonia
        w12 = 1.6*D/(eta[1][1]*capconvliquid)   # Liquid
        # Reconversion
        w21 = D/capconvammonia # Ammonia
        w22 = D/capconvammonia # Liquid
        W = [[w11, w12], [w21, w22]]
        
        
        #  For now the converters and reconverter amount is constant
        WC = quicksum(W[i-1][E]*A[l][i-1][E] for i in I) 
        
        #%% --- Variables ---
        
        # # w[i,j]
        # w = {}
        # for i in I:
        #     for j in J:
        #         w[i,j] = model.addVar (lb = 0, vtype = GRB.INTEGER, name = 'w[' + str(i) + ',' + str(j) + ']' )
        
        # x[k]
        x = {}
        for k in K:
            x[k] = model.addVar (lb = 0, vtype = GRB.INTEGER, name = 'x[' + str(k) + ']' )
            
        # y[n]
        y = {}
        for n in N:
            y[n] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'y[' + str(n) + ']' )
        
        PU = {}
        for t in T:
            PU[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PU[' + str(t) + ']' )
            
        # Power Generated (used to lower RHS values)
        PG = {}
        for t in T:
            PG[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PG[' + str(t) + ']' )
        
        # Hydrogen produced
        h = {}
        for t in T:
            h[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'h[' + str(t) + ']' )
        
        
        
        # %%  ---- Integrate new variables ----
        model.update()
        
        # # ---- Objective Function ----
        # # For now just setting year to first year in time period (l=0)
        # # l = 1
        # model.setObjective (WC + quicksum(x[k]*B[l][k-1] for k in K) + quicksum(y[n]*C[l][n-1] for n in N) + TC[l][E])
        # model.modelSense = GRB.MINIMIZE
        # model.update ()
        
        #%% --- Constraints ---
        
        for t in T:
            model.addConstr(PG[t] == my_turbinearray[t]*x[1] + feedinarray[t]*x[2])
            model.addConstr(PU[t] <= alpha[E]*PG[t])
            model.addConstr(PU[t] <= beta*gamma*x[3])
            model.addConstr(h[t] == PU[t]/gamma)
        
        model.addConstr(quicksum(h[t] for t in T) >= D/(eta[0][E]*eta[1][E]))  
        
        model.addConstr(x[4] >= x[3]*beta*epsilon)
        
        model.addConstr(y[2] == x[3]*nu[E])
        
        model.addConstr(y[1] == y[2]*phi[E])
        
        
        #%% --- Run Optimization ---
        model.update()
        
        model.setParam( 'OutputFlag', False) # silencing gurobi output or not
        model.Params.NumericFocus = 1
        model.Params.timeLimit = 400
        # model.optimize()
        
        Result = []
        feedin_index = func_Feedin_index(list(loc))
        hdata = pd.DataFrame(index=feedin_index) #for plotting the hydrogen production
        exceldf = pd.DataFrame(index=['Demand (tons of hydrogen)', 'Usage location', 'Year','Production location','Transfer port','Total costs per year (euros)','Costs per kg hydrogen (euros)','Wind turbines', 'Solar platforms','Electrolyzers','Desalination equipment', 'Storage volume (m3)','Conversion devices','Reconversion devices','Transport medium', 'FPSO volume (m3)', 'Distance sea (km)','Distance land (km)'])
        demandlocation = general.cell(row=17, column=2).value #location where hydrogen is asked
        productionlocation = general.cell(row=15, column=2).value #location where hydrogen is produced
        transferport = general.cell(row=16, column=2).value #port where hydrogen is transferred from sea to land transport
        
        
        transportmedium = str('')
        if E == 0:
            transportmedium = 'Ammonia'
        elif E == 1:
            transportmedium = 'Liquid_H2'
        
        # --- Objective function ---
        
        for l in L: 
            model.reset()
            model.setObjective (WC + quicksum(x[k]*B[l][k-1] for k in K) + quicksum(y[n]*C[l][n-1] for n in N) + TC[E][l])
            model.modelSense = GRB.MINIMIZE
            model.update ()
            model.optimize()
            Result.append(model.ObjVal)
            exceldf[timestep*l+Startyear] = [demand*(1/DataYearRatio),demandlocation,timestep*l+Startyear, productionlocation, transferport, model.ObjVal*(1/DataYearRatio), model.ObjVal/demand/1000,x[1].X, x[2].X, x[3].X, x[4].x, y[1].X, W[0][E], W[1][E], transportmedium, y[2].X,distancesea,distanceland]
            dict_of_df[loc][timestep*l+Startyear] = [demand*(1/DataYearRatio),demandlocation,timestep*l+Startyear, productionlocation, transferport, model.ObjVal*(1/DataYearRatio), model.ObjVal/demand/1000,x[1].X, x[2].X, x[3].X, x[4].x, y[1].X, W[0][E], W[1][E], transportmedium, y[2].X,distancesea,distanceland]
            print('------- Completed run: ', l+1, ' out of ', max(L)+1 , '     (Year: ',Startyear+timestep*l ,')')
        print('--------- Completed horizontal loc runs: ', counter+1, 'out of ', size)    
        print('--------- Completed vertical loc runs: ', vert+1, 'out of ', size)
        # #plot produced hydrogen
        # hvalues = np.empty(len(h), dtype=object)
        # for t in range(len(h)):
        #     hvalues[t] = h[t].x
        # hdata['production'] = hvalues.tolist()
        # hdata.plot(title='')
        # plt.xlabel('Time')
        # plt.ylabel('Tons per hour')
        
        
        
        #%% --- Post-Processing ---
        
        # Data with latitude/longitude and values
        # df = pd.read_csv('https://raw.githubusercontent.com/R-CoderDotCom/data/main/sample_datasets/population_galicia.csv')
        
        # data_file_csv = os.path.dirname(os.path.realpath('__file__')) + '\csv_files\heatmap1.csv'
        # df = pd.read_csv(data_file_csv)
        
        df = pd.DataFrame(columns=['longitude','latitude','LCOH'],index=[list(range(size))])
        
        # Add resulting cost per kg of h2 in 2050
        LCOH = dict_of_df[loc].loc['Costs per kg hydrogen (euros)',2050]
        df.loc[counter,'LCOH'] = LCOH
        df.loc[counter,'longitude'] = loc_matrix[vert][counter][0]
        df.loc[counter,'latitude'] = loc_matrix[vert][counter][1]
        
        list_vert_locs.append(dict_of_df)
        # dict_of_df[loc][timestep*l+Startyear] = [demand*(1/DataYearRatio),demandlocation,timestep*l+Startyear, productionlocation, transferport, model.ObjVal*(1/DataYearRatio), model.ObjVal/demand/1000,x[1].X, x[2].X, x[3].X, x[4].x, y[1].X, W[0][E], W[1][E], transportmedium, y[2].X,distancesea,distanceland]
        counter = counter + 1
        # df.loc[0,'ObjVal'] = model.ObjVal

df.to_csv("test_csv.csv")


# fig = px.density_mapbox(df, lat = 'latitude', lon = 'longitude', z = 'LCOH',
#                         radius = 7,
#                         center = dict(lat = 26.81, lon = 126.94),
#                         zoom = 3,
#                         mapbox_style = 'open-street-map',
#                         color_continuous_scale = 'rainbow')

# # Usage location
# fig.add_trace(go.Scattermapbox(
#         lat=[35.64],
#         lon=[139.8],
#         mode='markers',
#         marker=dict(size=10, color="Orange"),
#         name="Usage Location",
    
#     ))


# pio.renderers.default='browser'
# fig.show()


#%% --- Plotting ---
counter = 0
# list_dfs = []
# for vert in range(size):
#     for i in range(len(list_vert_locs)):
#         for j in range(size):
#             list_dfs.append(list_vert_locs[i][loc_matrix[vert][j]])

list_dfs2 = []
for vert in range(size):
    for j in range(size):
        # print(counter)
        # print('Iteration: ', counter, ' Location: ', loc_matrix[vert][j]) 
        list_dfs2.append(list_vert_locs[counter][loc_matrix[vert][j]])
        counter = counter + 1

counter = 0
df_full = pd.DataFrame(columns=['longitude','latitude','LCOH'],index=[list(range(441))])
for vert in range(size):
    counter2=0
    for j in range(size):
        # print(counter)
        # print('Iteration: ', counter, ' Location: ', loc_matrix[vert][j], 'LCOH: ', list_dfs2[counter].loc['Costs per kg hydrogen (euros)',2050]) 
        # list_dfs2[counter].loc['Costs per kg hydrogen (euros)',2050]
        LCOH = list_dfs2[counter].loc['Costs per kg hydrogen (euros)',2050]
        df_full.loc[counter,'LCOH'] = LCOH
        df_full.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
        df_full.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
        counter2 = counter2 + 1     
        counter = counter + 1      
        
df_full_backup = df_full
df_full = df_full_backup
avg = df_full['LCOH'].mean()
# df_full['LCOH'] = df_full['LCOH'].apply(lambda x: x-avg)      
# df_full['LCOH'] = df_full['LCOH'].apply(lambda x: x*100)  


# Color palettes: 'RdBu', 
fig = px.density_mapbox(df_full, lat = 'latitude', lon = 'longitude', z = 'LCOH',
                        radius = 15,
                        center = dict(lat = latitude, lon = longitude),
                        zoom = 3,
                        mapbox_style = 'open-street-map',
                        color_continuous_scale = 'Rainbow')



# Adjust color of heatmap by adding more points for density
fig.add_trace(
    go.Scattermapbox(
        lat=df_full["latitude"],
        lon=df_full["longitude"],
        mode="markers",
        showlegend=False,
        hoverinfo="skip",
        marker={
            "color": df_full["LCOH"],
            "size": df_full["LCOH"].fillna(0).infer_objects(copy=False),
            "coloraxis": "coloraxis",
            # desired max size is 15. see https://plotly.com/python/bubble-maps/#united-states-bubble-map
            "sizeref": (df_full["LCOH"].max()) / 15 ** 2,
            "sizemode": "area",
        },
    )
)

# Usage location
fig.add_trace(go.Scattermapbox(
        lat=[coords_demand[0]],
        lon=[coords_demand[1]],
        mode='markers',
        marker=dict(size=10, color="Orange"),
        name="Usage Location",
    
    ))


pio.renderers.default='browser'
fig.show()

df_full.to_csv("df_full_csv.csv")



# list_vert_locs

# for loc in loc_matrix[vert]:
#     LCOH = dict_of_df[loc].loc['Costs per kg hydrogen (euros)',2050]
#     df.loc[counter,'LCOH'] = LCOH
#     df.loc[counter,'longitude'] = loc_matrix[vert][counter][0]
#     df.loc[counter,'latitude'] = loc_matrix[vert][counter][1]
#     counter = counter + 1

# fig = px.density_mapbox(df, lat = 'latitude', lon = 'longitude', z = 'LCOH',
#                         radius = 7,
#                         center = dict(lat = 26.81, lon = 126.94),
#                         zoom = 3,
#                         mapbox_style = 'open-street-map',
#                         color_continuous_scale = 'rainbow')

# # Usage location
# fig.add_trace(go.Scattermapbox(
#         lat=[35.64],
#         lon=[139.8],
#         mode='markers',
#         marker=dict(size=10, color="Orange"),
#         name="Usage Location",
    
#     ))


# pio.renderers.default='browser'
# fig.show()
