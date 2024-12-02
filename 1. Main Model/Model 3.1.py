# -*- coding: utf-8 -*-
"""
Created on Thu May  2 11:41:52 2024

@author: Piotr Kaczmarek
"""

#%% 1. Preamble

# Import Packages
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import geopy.distance
import math
from openpyxl import load_workbook
from feedinlib import era5
import pvlib
from feedinlib import Photovoltaic
from feedinlib import get_power_plant_data
from windpowerlib.modelchain import ModelChain
from windpowerlib.wind_turbine import WindTurbine
from windpowerlib import get_turbine_types
from datetime import timedelta
from gurobipy import *
import plotly.express as px
import plotly.io as pio
import plotly.graph_objects as go
from sklearn.linear_model import LinearRegression
import statistics
import warnings

# Suppress outdated package warning
def fxn():
    warnings.warn("deprecated", DeprecationWarning)
with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    fxn()
pd.set_option('future.no_silent_downcasting', True)


#%% 2. Excel Parameters

# Load excel parameter data file
data_file = os.getcwd() 
data_file = os.path.dirname(os.path.realpath('__file__')) + "\ " + str(1) + '. Main Model\Parameter Data.xlsx' # Locate file
data_file = data_file[0:94] + data_file[109:142]
wb = load_workbook(data_file,data_only=True) # Load Workbook

# general = wb['General']                   # Extract data from sheet called general
# wind = wb['Wind']                         # Extract data from sheet called wind
solar = wb['Solar']                         # Extract data from sheet called solar
electrolyzer = wb['Electrolyzer']           # Extract data from sheet called electrolyzer
desalination = wb['Desalination']           # Extract data from sheet called desalination
ammonia = wb['Ammonia']                     # Extract data from sheet called ammonia
liquidhydrogen = wb['Liquid hydrogen']      # Extract data from sheet called liquid hydrogen
landtransport = wb['Land transport']        # Extract data from sheet called landtransport
storage = wb['Storage']                     # Extract data from sheet called storage
fpso = wb['FPSO']                           # Extract data from sheet called fpso

# Time period data (Make sure this matches with excel file "Parameter Data")
Startyear = 2020                            # Starting year of time period
timeperiod = 30                             # Total time period span [years]
timestep = 3                                # Time step size within time period
Nsteps = int(timeperiod/timestep+1)         # Number of time steps

# Set coordinates of production and demand locations
# North-Sea case 
latitude = 54.35
longitude = 6.78
#  Eemshaven (Groningen)
coords_demand = (53.43, 6.84) 
coords_port = (53.43, 6.84) 

# Map visual output (0 if off, 1 if on)
MapOutput = 1 

# Era 5 weather data file location
era5_netcdf_filename = 'Era 5 test data\ERA5_weather_data_NorthSea_010120-300920.nc' #North-Sea Case data for 9 months in 2020
# Set to the time period ratio of the CDS dataset considered (1/12 means 1 month)
DataYearRatio = 9/12

# Solar and Wind capacity multiplier for later functions
multsolar = 524*1000/206.7989               # Installed capacity per unit [kWp] * 1000 (to make it [W]) / capacity from PVlib [Wp]
multwind = 12/3                             # Capacity per unit [MW] / capacity from Windlib [MW]

# Assumed rates [%/100]
DiscountRate = 0.08
InterestRate = 0.08

# Theoretical Demand Parameter (needed for now for transport calculation)
capacity = 700 # [MW] theoretical capacity of north sea case
demand = capacity*8760/1000/0.0505 # [tonH2/yr] theoretical demand at 100% CF

#%% 2.1 Sensitivity Analysis (Optional)

# # If you would like to perform a sensitivity analysis, uncomment the following code section and add indent from line 204 onwards
sensitivity_file_path = "Try 16/" # File path for results
# # Sensitivity scale parameter for sensitivity analysis
SSP_A1 = 1      # Conversion
SSP_A2 = 1      # Reconversion
SSP_B1 = 1      # Wind
SSP_B2 = 1      # Solar
SSP_B3 = 1      # Electrolysis
SSP_B4 = 1      # Desalination
SSP_C1 = 1      # Storage
SSP_C2 = 1      # FPSO vessel size
SSP_TC = 1      # Transport

# SSP_list = [SSP_A1,
# SSP_A2,
# SSP_B1,
# SSP_B2,
# SSP_B3,
# SSP_B4,
# SSP_C1,
# SSP_C2,
# SSP_TC]

# SSP_string_list = ["SSP_A1",
# "SSP_A2",
# "SSP_B1",
# "SSP_B2",
# "SSP_B3",
# "SSP_B4",
# "SSP_C1",
# "SSP_C2",
# "SSP_TC"]

# SSP_lower_case = 0.8
# SSP_reference_case = 1
# SSP_upper_case = 1.2



# for iteration_SSP in range(1):

    
#     # Sensitivity scale parameter for sensitivity analysis
#     SSP_A1 = 1      # Conversion
#     SSP_A2 = 1      # Reconversion
#     SSP_B1 = 1      # Wind
#     SSP_B2 = 1      # Solar
#     SSP_B3 = 1      # Electrolysis
#     SSP_B4 = 1      # Desalination
#     SSP_C1 = 1      # Storage
#     SSP_C2 = 1      # FPSO vessel size
#     SSP_TC = 1      # Transport
    
#     SSP_list = [SSP_A1,
#     SSP_A2,
#     SSP_B1,
#     SSP_B2,
#     SSP_B3,
#     SSP_B4,
#     SSP_C1,
#     SSP_C2,
#     SSP_TC]
    
#     SSP_list[iteration_SSP] = SSP_reference_case # lower bound (20% decrease), reference 0% difference, upper 20% increase
#     SSP_A1 = SSP_list[0]      # Conversion
#     SSP_A2 = SSP_list[1]      # Reconversion
#     SSP_B1 = SSP_list[2]      # Wind
#     SSP_B2 = SSP_list[3]      # Solar
#     SSP_B3 = SSP_list[4]      # Electrolysis
#     SSP_B4 = SSP_list[5]      # Desalination
#     SSP_C1 = SSP_list[6]      # Storage
#     SSP_C2 = SSP_list[7]      # FPSO vessel size
#     SSP_TC = SSP_list[8]      # Transport
    
#     SSP_list = [SSP_A1,
#     SSP_A2,
#     SSP_B1,
#     SSP_B2,
#     SSP_B3,
#     SSP_B4,
#     SSP_C1,
#     SSP_C2,
#     SSP_TC]
    
#     print(SSP_string_list[0], SSP_A1)
#     print(SSP_string_list[1], SSP_A2)
#     print(SSP_string_list[2], SSP_B1)
#     print(SSP_string_list[3], SSP_B2)
#     print(SSP_string_list[4], SSP_B3)
#     print(SSP_string_list[5], SSP_B4)
#     print(SSP_string_list[6], SSP_C1)
#     print(SSP_string_list[7], SSP_C2)
#     print(SSP_string_list[8], SSP_TC)
    
    # sensitivity_file_path = "Sensitivity Analysis/"
#     # if SSP_A1 == 1 and SSP_A2 == 1 and SSP_B1 == 1 and SSP_B2 == 1 and SSP_B3 == 1 and SSP_B4 == 1 and SSP_C1 == 1 and SSP_C2 == 1 and SSP_TC == 1:
#     #     print("Reference Case")
#     #     sensitivity_file_path = "Sensitivity Analysis/Reference Case/" + SSP_string_list[iteration_SSP] 
#     # elif SSP_A1 < 1 or SSP_A2 < 1 or SSP_B1 < 1 or SSP_B2 < 1 or SSP_B3 < 1 or SSP_B4 < 1 or SSP_C1 < 1 or SSP_C2 < 1 or SSP_TC < 1:
#     #     print("Lower Case")
#     #     sensitivity_file_path = "Sensitivity Analysis/Lower Case/" + SSP_string_list[iteration_SSP] 
#     # elif SSP_A1 > 1 or SSP_A2 > 1 or SSP_B1 > 1 or SSP_B2 > 1 or SSP_B3 > 1 or SSP_B4 > 1 or SSP_C1 > 1 or SSP_C2 > 1 or SSP_TC > 1:
#     #     print("Upper Case")
#     #     sensitivity_file_path = "Sensitivity Analysis/Upper Case/" + SSP_string_list[iteration_SSP]       


# If performing a sensitivity analysis, indent all code following this point (by highlighting and pressing TAB)
for iteration_j in range(0,4):        # range(2) to study optimal location of j1 and j2, range(2,4) for j3 and j4, range(0,4) for all j in J
    
    # Energy Medium (0: NH3 ship, 1: LH2 ship, 2: GH2 pipe, 3: NH3 pipe)
    E = iteration_j
    
    # Factor for transport cost, if demand and production locations cannot be reached using geodesic >1, otherwise 1
    distanceseafactor = 1
    
    # Technical parameters required for precomputed 
    beta = 0.4       # Hourly output capacity of PEM electrolyzer [tonH2/device/h]
    gamma = 50500000 # Energy requirement of an electrolyzer to produce ton of hydrogen [Wh/tonH2]
 
    
    
    #%% 3. Wind Parameters - Excel replacement
    
    # Input parameters for wind cost parameter calculation
    learning_rate_wind = 0.088                          # Learning rate [%/100]
    reduction_factor = np.log2(1 - learning_rate_wind)  # Computation for reduction factor
    Capacity_Wind = 12                                  # Capacity of a single wind turbine [MW]
    CAPEX_wind_initial = 1944                           # Total CAPEX of a wind turbine [Eur/kW] in 2020
    CAPEX_wind_initial_MW = CAPEX_wind_initial * 1000   # CAPEX conversion to Megawatts [Eur/MW]
    OPEX_wind_initial = 64                              # OPEX [Eur/kW] in 2020
    OPEX_wind_initial_MW = OPEX_wind_initial*1000       # OPEX conversion to Megawatts [Eur/MW]
    lftm = 25                                           # Expected lifetime of wind turbine [Years]
    a_wind = (InterestRate*(1+InterestRate)**lftm) / ((1+InterestRate)**lftm - 1) # Amortization factor
    
    # Calculate 1 dimsensional arrays 
    years = np.arange(2020, 2051)                                   # Array of years from 2021 to 2050
    annual_increase = (250000 - 50000) / (2031 - 2021)              # Global cumulative capacity increase per year [GW/y] 
    cumulative_capacity = 50000 + (years - 2021) * annual_increase  # Array of global cumulative capacity [GW]
    CAPEX_wind_year = CAPEX_wind_initial_MW * (cumulative_capacity / 50000) ** reduction_factor     # Array of CAPEX for year based on cumulative capacity [Eur/MW]
    CAPEX_wind_year_list = list(zip(years, CAPEX_wind_year*Capacity_Wind))                          # Conversion into list format
    OPEX_wind_year = OPEX_wind_initial_MW * (cumulative_capacity / 50000) ** reduction_factor       # Array of OPEX for year based on cumulative capacity [Eur/MW]
    OPEX_wind_year_list = list(zip(years, OPEX_wind_year*Capacity_Wind))                            # Conversion into list format
    
    Wind_Costs = np.zeros((5,31))   # Prepare empty 5x31 array
    
    # Input values into a complete wind cost array from 2020 to 2050
    yearstep = 0
    for i in range(31):
        Wind_Costs[0][i] = 2020 + yearstep
        yearstep += 1
        
    for i in range(31):
        Wind_Costs[1][i] = CAPEX_wind_year_list[i][1]
    
    for i in range(31):
        Wind_Costs[2][i] = lftm
        
    for i in range(31):
        Wind_Costs[3][i] = OPEX_wind_year_list[i][1]
    
    Wind_Costs[4] = a_wind*Wind_Costs[1] + Wind_Costs[3]
    
    # Create function to retrieve relevant wind array of the correct size
    def rel_wind_array(SY, NS, TS, WiCo):
        '''
        Inputs - SY: StartingYear
                  NS: NSteps
                  TS: Time Step
                  WiCo: Wind Costs array
        Outputs - Full Wind cost array of relevant format
        '''
        index = np.where(Wind_Costs[0] == SY)[0][0]
        relevant_array = np.zeros(NS)
        for i in range(NS):    
            relevant_array[i] = WiCo[4][index+i*TS]
        
        return relevant_array
    
    
    #%% 4. Solar Power Function 
    
    
    # area = [longitude, latitude]
        
    
    def func_PV(location):
        # Input:  location = [longitude, latitude] 
        #         type: list
        # Output: func_PV(location)[0] = feedinarray
        #         func_PV(location)[1] = feedin_index
        
        lat_location = location[1]
        lon_location = location[0]
        # get modules
        module_df = get_power_plant_data(dataset='SandiaMod') #retrieving dataset for PV modules
        
        # get inverter data
        inverter_df = get_power_plant_data(dataset='cecinverter') #retrieving dataset for inverters
        
        temp_params = pvlib.temperature.TEMPERATURE_MODEL_PARAMETERS['sapm']['open_rack_glass_polymer'] #defining temperature model parameters (leaving at default causes errors)
        
        #PV system definition
        system_data = {
            'module_name': 'Advent_Solar_Ventura_210___2008_',  # module name as in database
            'inverter_name': 'ABB__MICRO_0_25_I_OUTD_US_208__208V_',  # inverter name as in database
            'azimuth': 180, #angle of sun position with north in horizontal plane
            'tilt': 30, #angle of solar panels with horizontal
            'albedo': 0.075, #albedo (fraction of sun light reflected) of ocean water (https://geoengineering.global/ocean-albedo-modification/)
            'temperature_model_parameters': temp_params,
        }
        
        pv_system = Photovoltaic(**system_data)
        
        #getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
        # area = [lon, lat]
        pvlib_df = era5.weather_df_from_era5(     
            era5_netcdf_filename=era5_netcdf_filename,
            lib='pvlib', area=[lon_location, lat_location])
        
        #determining the zenith angle (angle of sun position with vertical in vertical plane) in the specified locations for the time instances downloaded from ERA5
        zenithcalc = pvlib.solarposition.get_solarposition(time=pvlib_df.index,latitude=lat_location, longitude=lon_location, altitude=None, pressure=None, method='nrel_numpy', temperature=pvlib_df['temp_air'])
        
        #determining DNI from GHI, DHI and zenith angle for the time instances downloaded from ERA5
        dni = pvlib.irradiance.dni(pvlib_df['ghi'],pvlib_df['dhi'], zenith=zenithcalc['zenith'], clearsky_dni=None, clearsky_tolerance=1.1, zenith_threshold_for_zero_dni=88.0, zenith_threshold_for_clearsky_limit=80.0)
        
        #adding DNI to dataframe with PV weather data
        pvlib_df['dni'] = dni
        
        #replacing 'NAN' in DNI column with 0 to prevent 'holes' in graphs (NANs are caused by the zenith angle being larger than the 'zenith threshold for zero dni' angle set with GHI and DHI not yet being zero. The DNI should be 0 in that case)
        pvlib_df['dni'] = pvlib_df['dni'].fillna(0)
        
        #determining PV power generation
        # location = (lat,lon)
        feedin = pv_system.feedin(
            weather=pvlib_df,
            location=(lat_location, lon_location))
        
        feedinarray = multsolar*feedin.values #hourly energy production over the year of 1 solar platform of the specified kind in the specified location
        feedinarray[feedinarray<0]=0
        
        # #plotting PV power generation
        # plt.figure()
        # feedin.plot(title='PV feed-in '+str(lat_location)+', '+str(lon_location))
        # plt.xlabel('Time')
        # plt.ylabel('Power in W');
        
        feedin_index = feedin.index
        
        return feedinarray, feedin_index    
    
    #%% 5. Wind Power Function 
    
    def func_Wind(location):
        
        lat_location = location[1]
        lon_location = location[0]
        
        df = get_turbine_types(print_out=False)
        
        #defining the wind turbine system
        turbine_data= {
            "turbine_type": "E-101/3050",  # turbine type as in register
            "hub_height": 130,  # in m
        }
        my_turbine = WindTurbine(**turbine_data)
        
        #getting the needed weather data for wind power calculations from the file as downloaded with a seperate script from ERA5
        # [lon,lat]
        windpowerlib_df = era5.weather_df_from_era5(
            era5_netcdf_filename=era5_netcdf_filename,
            lib='windpowerlib', area=[lon_location, lat_location])
        
        #increasing the time indices by half an hour because then they match the pvlib time indices so the produced energy can be added later
        #assumed wind speeds half an hour later are similar and this will not affect the results significantly
        windpowerlib_df.index = windpowerlib_df.index + timedelta(minutes=30)
        
        # power output calculation for e126
        
        # own specifications for ModelChain setup
        modelchain_data = {
            'wind_speed_model': 'logarithmic',      # 'logarithmic' (default),
                                                    # 'hellman' or
                                                    # 'interpolation_extrapolation'
            'density_model': 'ideal_gas',           # 'barometric' (default), 'ideal_gas'
                                                    #  or 'interpolation_extrapolation'
            'temperature_model': 'linear_gradient', # 'linear_gradient' (def.) or
                                                    # 'interpolation_extrapolation'
            'power_output_model': 'power_curve',    # 'power_curve' (default) or
                                                    # 'power_coefficient_curve'
            'density_correction': True,             # False (default) or True
            'obstacle_height': 0,                   # default: 0
            'hellman_exp': None}                    # None (default) or None
        
        # initialize ModelChain with own specifications and use run_model method to
        # calculate power output
        mc_my_turbine = ModelChain(my_turbine, **modelchain_data).run_model(windpowerlib_df)
        # write power output time series to WindTurbine object
        my_turbine.power_output = mc_my_turbine.power_output
        
        # title = 'Wind turbine power production at ' + str(location)
        # # plot turbine power output
        # plt.figure()
        # my_turbine.power_output.plot(title=title)
        # plt.xlabel('Time')
        # plt.ylabel('Power in W')
        # plt.show()
        
        my_turbinearray = multwind*my_turbine.power_output.values #hourly energy production over the year of 1 wind turbine of the specified kind in the specified location
        
        return my_turbinearray
    
    
    #%% 6. Transport Cost Functions & Parameters 
    
    # Transport cost (TC)
    coords_production = (latitude, longitude)
    distanceland = geopy.distance.geodesic(coords_port, coords_demand).km #distance to be travelled over land from production to demand location in km
    
    # Excel parameters
    Xtransport = DataYearRatio*demand #amount to be transported by ship in tons of hydrogen, assumed to be equal to demand as hydrogen losses are almost zero 
    Xkmland = DataYearRatio*demand*distanceland #yealy amount of tonkm over land
    Ckmammonia = np.array([float(cell.value) for cell in ammonia[158][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as ammonia in 10^3 euros over several years
    Cbasetransportammonia = np.array([float(cell.value) for cell in ammonia[159][2:2+Nsteps]])
    Ckmliquid = np.array([float(cell.value) for cell in liquidhydrogen[168][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as liquid hydrogen in 10^3 euros over several years
    Cbasetransportliquid =  np.array([float(cell.value) for cell in liquidhydrogen[169][2:2+Nsteps]])
    Ckmland = np.array([float(cell.value) for cell in landtransport[33][2:2+Nsteps]]) #costs per year per km of overland pipeline for 1 ton of hydrogen
    
    # Changes by transport mode: pipeline, vs shipping
    # Shipping cost
    Ckmsea = [Ckmammonia, Ckmliquid] #costs per ton per km of overseas transport, depending on whether ammonia or liquid hydrogen is chosen
    Cbasetransport = [Cbasetransportammonia, Cbasetransportliquid] #baserate of the transport per ton hydrogen
    
    # Transport Cost Pipeline computation
    # Parameters NH3 (j=4)
    Lifetime_Pipe_NH3 = 30 # Project Lifetime NH3 pipeline [years]
    a_pipe_NH3 =(InterestRate*(1+InterestRate)**Lifetime_Pipe_NH3)/((1+InterestRate)**Lifetime_Pipe_NH3) # Amortization factor
    C_pipe_NH3 = 771000 # NH3 Pipeline cost per km [Eur]
    C_pump_NH3 = 1800000 # NH3 pump cost per station [Eur/pumpstation]
    Pump_distance = 128.8 # Distance above which an additional pump is required [km]
    Offshore_factor = 2 # Multiplier for estimating pipe cost (1: onshore, 2: offshore)
    
    # Function calculating the transport cost parameter for NH3 pipeline
    def Transport_NH3_Pipe(distance):
        # Input: distance [km], Output: 3xNsteps Cost array for each year l, row 1: total yearly cost, row 2: CAPEX, row 3: OPEX
    
        n_pump = math.ceil(distance/128.8) # Number of pump stations required at selected distance [pumpstations]
        CAPEX_pipe_NH3 = a_pipe_NH3*Offshore_factor*(C_pipe_NH3*distance+C_pump_NH3*n_pump) # Yearly CAPEX
        OPEX_pipe_NH3 = 0.02*Offshore_factor*(C_pipe_NH3*distance+C_pump_NH3*n_pump) # Yearly OPEX (2% of CAPEX)
        
        # Prepare output cost array
        C_pipeline_NH3 = np.zeros(Nsteps)
        
        # Create data array for each studied year
        for l in range(Nsteps):
            C_pipeline_NH3[l] = CAPEX_pipe_NH3 + OPEX_pipe_NH3
        return C_pipeline_NH3
    
    
    # Parameters GH2 (j=3)
    Lifetime_Pipe_GH2 = 30 # Project Lifetime NH3 pipeline [years]
    peak_x3 = math.ceil(demand/(beta*8760)) # Peak number of electrolyzers required in project [# of electrolyzers]
    a_pipe_GH2 =(InterestRate*(1+InterestRate)**Lifetime_Pipe_GH2)/((1+InterestRate)**Lifetime_Pipe_GH2) # Amortization factor [-]
    v_pipe_GH2 = 15 # flow rate [m/s]
    rho_pipe_GH2 = 8 # density GH2 [kg/m3]
    Q_pipe_GH2 = 1000*beta*peak_x3/3600 # mass flow rate [kg/s]
    F_pipe_GH2 = Q_pipe_GH2/rho_pipe_GH2
    D_pipe_GH2 = math.sqrt((4*F_pipe_GH2)/(math.pi*v_pipe_GH2)) # Inner diameter GH2 pipe [m]
    C_pipe_GH2 = 4000000*D_pipe_GH2*D_pipe_GH2 + 598600*D_pipe_GH2 + 329000 # GH2 Pipeline cost per km [Eur]
    
    
    def Transport_GH2_Pipe(distance):
        
        CAPEX_pipe_GH2 = a_pipe_GH2*Offshore_factor*C_pipe_GH2*distance
        OPEX_pipe_GH2 = 0.02*Offshore_factor*C_pipe_GH2*distance
        
        # Prepare output cost array
        C_pipeline_GH2 = np.zeros(Nsteps)
        
        # Create data array for each studied year
        for l in range(Nsteps):
            C_pipeline_GH2[l] = CAPEX_pipe_GH2 + OPEX_pipe_GH2
        
        return C_pipeline_GH2
    
    
    # Energy medium set J (1 NH3 ship, 2 LH2 ship, 3 GH2 pipe, 4 NH3 pipe)
    J = [1, 2, 3, 4]
    
    
    if E==0:
        print('Energy Medium set to: NH3 ship')
    elif E==1:
        print('Energy Medium set to: LH2 ship')
    elif E==2:
        print('Energy Medium set to: GH2 pipe')
    elif E==3:
        print('Energy Medium set to: NH3 pipe')
    
    
    # Prepare TC_jl cost parameter array
    def func_TC(location, E):
        location = [location[1],location[0]]
        distancesea = distanceseafactor*geopy.distance.geodesic(location, coords_port).km
        Xkmsea = Xtransport*distancesea
        Cpipeline = np.array([Transport_NH3_Pipe(distancesea), Transport_GH2_Pipe(distancesea)])
        
        TC = []
        for j in J[0:2]:    # Shipping
            TCyear = Xtransport*Cbasetransport[j-1] + Xkmsea*Ckmsea[j-1] + Xkmland*Ckmland #Euro
            TC.append(TCyear)
    
        for j in J[2:4]:  # Pipeline
            TCyear = Cpipeline[j-3] # [Euro]
            TC.append(TCyear)
                
        return TC
    
    # TC = func_TC(area,E)
    
    
    
    #%% 7. Location Selection
    
    # # Define the dimensions of the matrix (play with these 4 values to determine start location)
    size = 6  # This will create a size*size matrix
    # Calculate the range for rows and columns
    start = 0 # Starting position relative to longitude
    end = 0 # Ending position relative to latitude
    resolution_map = 1 # Distance between locations
    
    
    
    # Create a matrix of the specified size, with each element being a tuple of length 2
    loc_matrix = np.empty((size, size), dtype=object)
    for i in range(size):
        for j in range(size):
            loc_matrix[i,j] = (start + j*resolution_map + longitude, end - i*resolution_map + latitude)
            
    if MapOutput == 1:
        df_map_test = pd.DataFrame(columns=['longitude','latitude','z'],index=[list(range(size*size))])
        
        fig_map = px.density_mapbox(df_map_test, lat = 'latitude', lon = 'longitude', z = 'z',
                                radius = 15,
                                center = dict(lat = latitude, lon = longitude),
                                zoom = 3,
                                mapbox_style = 'open-street-map',
                                color_continuous_scale = 'Rainbow')
        
        
        
        for i in range(size):
            for j in range(size):
                fig_map.add_trace(go.Scattermapbox(
                    lat=[loc_matrix[i,j][1]],
                    lon=[loc_matrix[i,j][0]],
                    mode='markers',
                    marker=dict(size=10, color="Orange"),
                    name="Location "+str(loc_matrix[i,j]),
            
                ))  
        
        
        pio.renderers.default='browser'
        fig_map.show()
        
        fig_map.data = []
    
    #%% 8. Prepare location loop
    
    # Denote lists to collect data
    Runtime_List = []           # Empty list for runtime informaiton
    list_vert_locs = []         # Empty list for location information
    CostDistribution_List = []  # Empty list for cost distribution information
    
    
    # For loop in vertical locations indices
    for vert in range(size):
        
        
        dict_of_df = {}
        for loci in loc_matrix[vert]:   # For loop in vertical locations coordinates
            dict_of_df[loci] = pd.DataFrame(index=['Demand (tons of hydrogen)', 'Usage location', 'Year','Production location','Transfer port','Total costs per year (euros)','Costs per kg hydrogen (euros)','Wind turbines', 'Solar platforms','Electrolyzers','Desalination equipment', 'Storage volume (m3)','Conversion devices','Reconversion devices','Transport medium', 'FPSO volume (m3)', 'Distance sea (km)','Distance land (km)'])
        
        feedinlist = []
        counter = 0     # Used for tracking progress of runs
        
        for loc in loc_matrix[vert]:    # For loop in horizontal locations coordinates
            # Select location 'loc' (varies by longitude)
            #loc = (lon,lat)
            
            # Calculate PV
            feedinarray = func_PV(list(loc))[0]
            
            # Calculate Wind
            my_turbinearray = func_Wind(loc)
            
            lon_prod = loc[0]
            lat_prod = loc[1]
            
            # Calculate TC_jl
            TC = func_TC(loc,E)                                                    # Transport cost at current location 'loc' for scenario 'E'
            TC = [TC[0]*SSP_TC, TC[1]*SSP_TC, TC[2]*SSP_TC, TC[3]*SSP_TC]          # Scaling for sensitivity analysis for transport
            
            print('Currently in location: (lat=', str(lat_prod),', lon=',str(lon_prod),')')
            distancesea = distanceseafactor*geopy.distance.geodesic((lat_prod,lon_prod), coords_port).km        # Distance calculated as geodesic
    
        
            #%% --- 9. Model parameters and sets ---
            
            # Set model name
            model = Model('GFPSO Cost Optimization')
            
            # ---- Sets ----
            
            I = [1, 2]                          # Conversion devices (1: Conversion, 2: Reconversion)
            J = [1, 2, 3, 4]                    # Energy Medium with Transport Mode combination (1: Ammonia Ship, 2: LiquidH2 Ship, 3: GaseousH2 Pipe, 4: Ammonia Pipe)
            K = [1, 2, 3, 4]                    # Device types (1: Wind, 2: Solar, 3: Elec, 4: Desal)
            L = range(Nsteps)                   # Years in time period
            N = [1, 2]                          # Volume based equipment (1: Storage, 2: FPSO)
            T = list(range(0,len(feedinarray))) # Operational hours in year
            
            
            # ---- Parameters ----
            
            # Cost parameters
            
            # Conversion and reconversion costs
            # Cconvammonia = DataYearRatio*rel_con_array(Startyear, Nsteps, timestep, Con_Costs)
            Cconvammonia = DataYearRatio*np.array([float(cell.value) for cell in ammonia[48][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
            Cconvliquid = DataYearRatio*np.array([float(cell.value) for cell in liquidhydrogen[61][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years
            Creconvammonia = DataYearRatio*np.array([float(cell.value) for cell in ammonia[111][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
            # Creconvammonia = np.zeros(Nsteps) # Uncomment this line to investigate the no NH3 reconversion situation
            Creconvliquid = DataYearRatio*np.array([float(cell.value) for cell in liquidhydrogen[125][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years
            Cconvgas = np.zeros(Nsteps) # j=3
            Creconvgas = np.zeros(Nsteps) #j=3
            
            
            # A_lij , l = year in time period, i = conversion device type, j = medium
            A = []
            for l in range(Nsteps):
                Aarray = np.array([[Cconvammonia[l], Cconvliquid[l], Cconvgas[l], Cconvammonia[l]], [Creconvammonia[l], Creconvliquid[l], Creconvgas[l], Creconvammonia[l]]])
                A.append(Aarray)
            
            for l in L:
                A[l][0][E] = A[l][0][E]*SSP_A1 # Conv
                A[l][1][E] = A[l][1][E]*SSP_A2 # Reconv
            
            
            Cs1 = DataYearRatio*np.array([float(cell.value) for cell in solar[51][2:2+Nsteps]])
            # Cw1 = DataYearRatio*np.array([float(cell.value) for cell in wind[48][2:2+Nsteps]])
            Cw1 = DataYearRatio*rel_wind_array(Startyear, Nsteps, timestep, Wind_Costs)
            Ce = DataYearRatio*np.array([float(cell.value) for cell in electrolyzer[50][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an electrolyzer in 10^3 euros over several years
            Cd = DataYearRatio*np.array([float(cell.value) for cell in desalination[49][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a desalination installation in 10^3 euros over several years
            
            # B_lk , l = year in time period, k = device type (solar, wind, elec, desal)
            B = []
            for l in L:
                Barray = np.array([Cw1[l], 
                                   Cs1[l], 
                                   Ce[l], 
                                   Cd[l]])
                B.append(Barray)
            
            for l in L:
                B[l][0] = B[l][0]*SSP_B1    # Wind
                B[l][1] = B[l][1]*SSP_B2    # Solar
                B[l][2] = B[l][2]*SSP_B3    # Electrolysis
                B[l][3] = B[l][3]*SSP_B4    # Desalination
            
            # C_lnj
            Cstliquid =  DataYearRatio*np.array([float(cell.value) for cell in storage[25][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of storage per m3 in euros over several years
            Cstammonia =  DataYearRatio*np.array([float(cell.value) for cell in storage[54][2:2+Nsteps]])
            Cstgas = np.zeros(Nsteps)  # Assumption: Pipeline counts as storage
            Cstammoniapipe = np.zeros(Nsteps) # Assumption: Pipeline counts as storage
            
            Cfpso = DataYearRatio*np.array([float(cell.value) for cell in fpso[53][2:2+Nsteps]]) #np.array([1,1,0.8,0.6,0.4]) #cost per year (depreciation+OPEX) of FPSO per m3 in 10^3 euros over several years
            Cst = [Cstammonia, Cstliquid, Cstgas, Cstammoniapipe]
            
            
            
            C = []
            for l in L:
                Carray = np.array([Cst[E][l], 
                                   Cfpso[l]])
                C.append(Carray)
            for l in L:
                C[l][0] = C[l][0]*SSP_C1 # Storage
                C[l][1] = C[l][1]*SSP_C2 # FPSO vessel size
            
            # D , Yearly demand
            D = (DataYearRatio)*demand # [tonH2/yr]
            
                
            
            
            # Non-cost parameters
            
            electrolyzer_energy = 1000*electrolyzer.cell(row=7, column=2).value #energy requirement of electrolyzer in Wh per ton of hydrogen
            gamma = electrolyzer_energy #50500000 [Wh/tonH2]
            
            Ed = 35
            Ec1 = 750/0.18
            Ec2 = 10180
            Ec3 = 0
            Ec4 = 750/0.18
            Es1 = 210
            Es2 = 600
            Es3 = 0
            Es4 = 0
            
            fracpowerelectrolyzerliquid = electrolyzer.cell(row=11, column=2).value #fraction of energy used by electrolyzers when using liquid hydrogen
            fracpowerelectrolyzerammonia = electrolyzer.cell(row=12, column=2).value #fraction of energy used by elecyrolyzers when using ammonia
            
            alpha_1 = (gamma/1000) / (Ed + Ec1 + Es1 + (gamma/1000))
            alpha_2 = (gamma/1000) / (Ed + Ec2 + Es2 + (gamma/1000))
            alpha_3 = (gamma/1000) / (Ed + Ec3 + Es3 + (gamma/1000))
            alpha_4 = (gamma/1000) / (Ed + Ec4 + Es4 + (gamma/1000))
            
            #alpha_j
            
            alpha = [alpha_1, alpha_2, alpha_3, alpha_4]
            
            # capelectrolyzerhour = electrolyzer.cell(row=9, column=2).value #hourly output capacity of electrolyzers in tons of hydrogen per hour
            # beta = capelectrolyzerhour
            
            
            
            electrolyzer_water = electrolyzer.cell(row=6, column=2).value #water requirement of electrolyzer in m3 of water per ton of hydrogen
            capdesalinationhour = desalination.cell(row=14, column=2).value #hourly output capacity desalination in m3 of water per hour
            epsilon = electrolyzer_water/capdesalinationhour
            
            eta_conversionammonia = ammonia.cell(row=13, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
            eta_conversionliquid =  liquidhydrogen.cell(row=16, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
            eta_reconversionammonia = eta_conversionammonia #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
            eta_reconversionliquid = eta_conversionliquid #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
            eta_conversiongas = 1
            eta_reconversiongas = 1
            
            # eta_ij , i = conv device type, j = energy medium and mode
            eta = 1/np.array([[eta_conversionammonia, eta_conversionliquid, eta_conversiongas, eta_conversionammonia],
                             [eta_reconversionammonia, eta_reconversionliquid, eta_reconversiongas, eta_reconversionammonia]])
            
            #nu_j
            volumefpsoliquid = 20*324.5*60.69*33.8/600 # (j=2) electrolyzer.cell(row=15, column=2).value #FPSO volume per electrolyzer liquid hydrogen
            volumefpsoammonia = volumefpsoliquid-(volumefpsoliquid*0.4469-(volumefpsoliquid*0.4469/(0.681*0.18/0.071)))  #(j=1) electrolyzer.cell(row=16, column=2).value #FPSO volume per electrolyzer ammonia
            volumefpsogaseouspipe = volumefpsoliquid # (j=3)
            volumefpsoammoniapipe = volumefpsoammonia # (j=4)
            
            nu = [volumefpsoammonia, volumefpsoliquid, volumefpsogaseouspipe, volumefpsoammoniapipe]
            
            #phi_j
            ratiostoragefpsoliquid = fpso.cell(row=18, column=2).value #ratio storage tanks in m3/fpso volume in m3 (j=2)
            ratiostoragefpsoammonia = fpso.cell(row=19, column=2).value #ratio storage tanks in m3/fpso volume in m3 (j=1)
            ratiostoragefpsogaseouspipe = 0 # Assumption: Pipeline counts as storage j=3
            ratiostoragefpsoammoniapipe = 0 # Assumption: Pipeline counts as storage j=4
            
            phi = [ratiostoragefpsoammonia, ratiostoragefpsoliquid, ratiostoragefpsogaseouspipe, ratiostoragefpsoammoniapipe]
            
            
            # Conversion
            capconvammonia = DataYearRatio*ammonia.cell(row=10, column=2).value #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for ammonia
            capconvliquid = DataYearRatio*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for liquid hydrogen
            
            w11 = math.ceil(1.6*D/(eta[1][0]*capconvammonia))   # NH3 j=1
            w12 = math.ceil(1.6*D/(eta[1][1]*capconvliquid))    # LH2 j=2
            w13 = 0                                             # GH2 j=3 (no need for conversion)
            w14 = math.ceil(1.6*D/(eta[1][0]*capconvammonia))   # NH3 j=4 (same as w11, both are ammonia)
            
            
            # Reconversion
            capreconvammonia = DataYearRatio*ammonia.cell(row=75, column=2).value #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for ammonia
            capreconvliquid = DataYearRatio*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for liquid hydrogen
            
            w21 = math.ceil(D/capreconvammonia)    # NH3 j=1
            w22 = math.ceil(D/capreconvliquid)     # LH2 j=2
            w23 = 0                                 # GH2 j=3
            w24 = math.ceil(D/capreconvammonia)    # NH3 j=4
            
            
            # w_ij Number of conversion and reconversion devices 
            W = [[w11, w12, w13, w14], [w21, w22, w23, w24]]
            
            
            #  For now the converters and reconverter amount is constant WC[i][j]
            # If not constant, A_lij could vary by l, and variable w_i should be added
            WC = quicksum(W[i-1][E]*A[l][i-1][E] for i in I) 
            
            #%% --- 10. Variables ---
            
            # x[k], unit based device variable
            x = {}
            for k in K:
                x[k] = model.addVar (lb = 0, vtype = GRB.INTEGER, name = 'x[' + str(k) + ']' )
                
            # y[n], volume based device variable
            y = {}
            for n in N:
                y[n] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'y[' + str(n) + ']' )
            
            # PU[t], Power Used by electrolyzer
            PU = {}
            for t in T:
                PU[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PU[' + str(t) + ']' )
                
            # PG[t], Power Generated (used to lower RHS values)
            PG = {}
            for t in T:
                PG[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'PG[' + str(t) + ']' )
            
            # h[t], Hydrogen produced
            h = {}
            for t in T:
                h[t] = model.addVar (lb = 0, vtype = GRB.CONTINUOUS, name = 'h[' + str(t) + ']' )
            
            model.update() # Intergrate new variables
          
            
            #%% --- 11. Constraints ---
            
            for t in T:
                model.addConstr(PG[t] == my_turbinearray[t]*x[1] + feedinarray[t]*x[2]) # Constraint 3.4
                model.addConstr(PU[t] <= alpha[E]*PG[t])                                # Constraint 3.5
                model.addConstr(1*PU[t] <= beta*gamma*x[3])                             # Constraint 3.6
                model.addConstr(h[t] == PU[t]/gamma)                                    # Constraint 3.7
            
            model.addConstr(quicksum(h[t] for t in T) >= D/(eta[0][E]*eta[1][E]))       # Constraint 3.8
            
            model.addConstr(x[4] >= x[3]*beta*epsilon)                                  # Constraint 3.9
            
            model.addConstr(y[2] == x[3]*nu[E])                                         # Constraint 3.10
            
            model.addConstr(y[1] == y[2]*phi[E])                                        # Constraint 3.11
            
            ## Additional constraints for specific scenarios, uncomment if needed
            # model.addConstr(x[1] == 0)# Force no wind, only solar
            # print('Forcing wind turbines to 0 units (Check Constraints)')
            # model.addConstr(x[2] == 0)# Force only wind, no solar
            # print('Forcing solar platforms to 0 units (Check Constraints)')
            
            
            #%% --- 12. Run Optimization ---
            
            # Gurobi parameters
            model.update()
            model.setParam( 'OutputFlag', False) # Gurobi output or not (If you want ouput, keep the line. If you dont want output, comment line out)
            model.Params.NumericFocus = 1
            model.Params.timeLimit = 400
            
            # Data collection preparation
            Result = []
            feedin_index = func_PV(list(loc))[1]
            hdata = pd.DataFrame(index=feedin_index)    # for plotting the hydrogen production
            exceldf = pd.DataFrame(index=['Demand (tons of hydrogen)', 'Usage location', 'Year','Production location','Transfer port','Total costs per year (euros)','Costs per kg hydrogen (euros)','Wind turbines', 'Solar platforms','Electrolyzers','Desalination equipment', 'Storage volume (m3)','Conversion devices','Reconversion devices','Transport medium', 'FPSO volume (m3)', 'Distance sea (km)','Distance land (km)'])
            demandlocation = 'Groningen'                # Location where hydrogen is needed
            productionlocation = 'North-Sea'            # Location where hydrogen is produced
            transferport = 'Groningen'                  # Port where hydrogen is transferred from sea to land transport
            
            # Information on selected configuration
            transportmedium = str('')
            if E == 0:
                transportmedium = 'NH3 ship'
            elif E == 1:
                transportmedium = 'LH2 ship'
            elif E == 2:
                transportmedium = 'GH2 pipe'
            elif E == 3:
                transportmedium = 'NH3 pipe'
            
            # --- Objective function (3.3)---
            
            for l in L: 
                model.reset()
                model.setObjective (WC + quicksum(x[k]*B[l][k-1] for k in K) + quicksum(y[n]*C[l][n-1] for n in N) + TC[E][l])
                model.modelSense = GRB.MINIMIZE
                model.update()
                model.optimize()
                
                # Due to assumption that shipping TC scales with demand, and pipeline TC does not, the total cost will be different in the scenarios
                if E == 0 or 1:
                    TotalCostYear = model.ObjVal*(1/DataYearRatio)
                    CostDistribution = [x[1].x*B[l][0]*(1/DataYearRatio),     # Wind cost total 
                                        x[2].x*B[l][1]*(1/DataYearRatio),     # Solar cost
                                        x[3].x*B[l][2]*(1/DataYearRatio),     # Elec
                                        x[4].x*B[l][3]*(1/DataYearRatio),     # Desal
                                        sum(W[i-1][E]*A[l][i-1][E] for i in I)*(1/DataYearRatio),                 # Conv
                                        y[1].x*C[l][0]*(1/DataYearRatio),     # Storage
                                        y[2].x*C[l][1]*(1/DataYearRatio),     # FPSO
                                        TC[E][l]*(1/DataYearRatio)                              # Transport
                                        ]     
                elif E == 2 or 3:
                    CostDistribution = [x[1].x*B[l][0]*(1/DataYearRatio),     # Wind cost total 
                                        x[2].x*B[l][1]*(1/DataYearRatio),     # Solar cost
                                        x[3].x*B[l][2]*(1/DataYearRatio),     # Elec
                                        x[4].x*B[l][3]*(1/DataYearRatio),     # Desal
                                        sum(W[i-1][E]*A[l][i-1][E] for i in I)*(1/DataYearRatio),                 # Conv
                                        y[1].x*C[l][0]*(1/DataYearRatio),     # Storage
                                        y[2].x*C[l][1]*(1/DataYearRatio),     # FPSO
                                        TC[E][l]                              # Transport
                                        ] 
                    TotalCostYear = sum(CostDistribution)
                    
                Result.append(model.ObjVal)
                
                
                
                
                # Extract output into lists and dataframes
                CostDistribution_List.append(CostDistribution)
                Runtime_List.append(model.Runtime)
                exceldf[timestep*l+Startyear] = [demand,demandlocation,timestep*l+Startyear, loc, transferport, TotalCostYear, TotalCostYear/demand/1000,x[1].X, x[2].X, x[3].X, x[4].x, y[1].X, W[0][E], W[1][E], transportmedium, y[2].X,distancesea,distanceland]
                dict_of_df[loc][timestep*l+Startyear] = [demand,demandlocation,timestep*l+Startyear, loc, transferport, TotalCostYear, TotalCostYear/demand/1000,x[1].X, x[2].X, x[3].X, x[4].x, y[1].X, W[0][E], W[1][E], transportmedium, y[2].X,distancesea,distanceland]
                print('------- Completed run: ', l+1, ' out of ', max(L)+1 , '     (Year: ',Startyear+timestep*l ,')')
            print('--------- Completed horizontal loc runs: ', counter+1, 'out of ', size)    
            print('--------- Completed vertical loc runs: ', vert+1, 'out of ', size)
            print('CostDistribution_List length: ', len(CostDistribution_List))
            
            
            
            #%% --- 13. Post-Processing ---
            
            # Data with latitude/longitude and values
            df = pd.DataFrame(columns=['longitude','latitude','Cost_per_kg'],index=[list(range(size))])
            # Add resulting cost per kg of h2 in 2050
            Cost_per_kg = dict_of_df[loc].loc['Costs per kg hydrogen (euros)',2050]
            df.loc[counter,'Cost_per_kg'] = Cost_per_kg
            df.loc[counter,'longitude'] = loc_matrix[vert][counter][0]
            df.loc[counter,'latitude'] = loc_matrix[vert][counter][1]
            list_vert_locs.append(dict_of_df)
            counter = counter + 1
    
    print("--- Runtime Statistics j=",E+1,"---")
    print("Total Runtime: ",sum(Runtime_List))
    print("Average Runtime: ", statistics.mean(Runtime_List))
    print("Max Runtime: ", max(Runtime_List))
    print("Min Runtime: ", min(Runtime_List))
        
    if E==0:
        title_str = ' NH3 ship'
    elif E==1:
        title_str = ' LH2 ship'
    elif E==2:
        title_str = ' GH2 pipe'
    elif E==3:
        title_str = ' NH3 pipe'
    
    
    #%% 14. Extract Complete Output
    
    counter = 0
    list_dfs2 = []
    for vert in range(size):
        for j in range(size):
            # print(counter)
            # print('Iteration: ', counter, ' Location: ', loc_matrix[vert][j]) 
            list_dfs2.append(list_vert_locs[counter][loc_matrix[vert][j]])
            counter = counter + 1
    
    COMPLETE_OUTPUT = list_dfs2
    
    # All values in 2050
    All_Headers = exceldf.index.to_list()
    
    counter = 0
    df_full_all = pd.DataFrame(columns=['longitude','latitude','Cost_per_kg', 'Wind turbines', 'Solar platforms', 'Electrolyzers', 'Desalination equipment', 'Storage volume', 'Conversion devices', 'Reconversion devices', 'Transport medium', 'FPSO volume', 'Distance sea'],index=[list(range(size*size))])
    for vert in range(size):
        counter2=0
        for j in range(size):
            Cost_per_kg = COMPLETE_OUTPUT[counter].loc['Costs per kg hydrogen (euros)',2050]
            df_full_all.loc[counter,'Cost_per_kg'] = Cost_per_kg
            df_full_all.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
            df_full_all.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
            df_full_all.loc[counter,'Wind turbines'] = COMPLETE_OUTPUT[counter].loc['Wind turbines',2050]
            df_full_all.loc[counter,'Solar platforms']= COMPLETE_OUTPUT[counter].loc['Solar platforms',2050]
            df_full_all.loc[counter,'Electrolyzers'] = COMPLETE_OUTPUT[counter].loc['Electrolyzers',2050]
            df_full_all.loc[counter,'Desalination equipment'] = COMPLETE_OUTPUT[counter].loc['Desalination equipment',2050]
            df_full_all.loc[counter,'Storage volume'] = COMPLETE_OUTPUT[counter].loc['Storage volume (m3)',2050]
            df_full_all.loc[counter,'Conversion devices'] = COMPLETE_OUTPUT[counter].loc['Conversion devices',2050]
            df_full_all.loc[counter,'Reconversion devices'] = COMPLETE_OUTPUT[counter].loc['Reconversion devices',2050]
            df_full_all.loc[counter,'Transport medium'] = COMPLETE_OUTPUT[counter].loc['Transport medium',2050]
            df_full_all.loc[counter,'FPSO volume'] = COMPLETE_OUTPUT[counter].loc['FPSO volume (m3)',2050]
            df_full_all.loc[counter,'Distance sea'] = COMPLETE_OUTPUT[counter].loc['Distance sea (km)',2050]
            counter2 = counter2 + 1     
            counter = counter + 1  
    
    
    
    # Cost_per_kg in 2050
    counter = 0
    df_full = pd.DataFrame(columns=['longitude','latitude','Cost_per_kg'],index=[list(range(size*size))])
    for vert in range(size):
        counter2=0
        for j in range(size):
            Cost_per_kg = list_dfs2[counter].loc['Costs per kg hydrogen (euros)',2050]
            df_full.loc[counter,'Cost_per_kg'] = Cost_per_kg
            df_full.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
            df_full.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
            counter2 = counter2 + 1     
            counter = counter + 1      
    
    
    
    
    df_full_backup = df_full
    df_full = df_full_backup
    avg = df_full['Cost_per_kg'].mean()
    
    df_relevant_lat = df_full
    
    
    
    #%% 15. Plot Cost per kg Map 2050
    
    if MapOutput == 1:
        fig = px.density_mapbox(df_full, lat = 'latitude', lon = 'longitude', z = 'Cost_per_kg',
                                radius = 15,
                                center = dict(lat = latitude, lon = longitude),
                                zoom = 3,
                                title = title_str,
                                mapbox_style = 'open-street-map',
                                color_continuous_scale = 'Rainbow')
        
        
        
        # Adjust color of heatmap by adding more points for density
        fig.add_trace(
            go.Scattermapbox(
                lat=df_full["latitude"],
                lon=df_full["longitude"],
                mode="markers",
                showlegend=False,
                hoverinfo="skip",
                marker={
                    "color": df_full["Cost_per_kg"],
                    "size": df_full["Cost_per_kg"].fillna(0).infer_objects(copy=False),
                    "coloraxis": "coloraxis",
                    # desired max size is 15. see https://plotly.com/python/bubble-maps/#united-states-bubble-map
                    "sizeref": (df_full["Cost_per_kg"].max()) / 15 ** 2,
                    "sizemode": "area",
                },
            )
        )
        
        # Usage location
        fig.add_trace(go.Scattermapbox(
                lat=[coords_demand[0]],
                lon=[coords_demand[1]],
                mode='markers',
                marker=dict(size=10, color="Orange"),
                name="Usage Location",
            
            ))
        
        
        pio.renderers.default='browser'
        fig.show()
    
    #%% 16. Plot Electrolyzer Map 2050
    df_map = df_full_all

    if MapOutput == 1:
        
    # Color palettes: 'RdBu', 
        fig = px.density_mapbox(df_map, lat = 'latitude', lon = 'longitude', z = 'Electrolyzers',
                                radius = 15,
                                center = dict(lat = latitude, lon = longitude),
                                zoom = 3,
                                mapbox_style = 'open-street-map',
                                title = title_str,
                                color_continuous_scale = 'Rainbow')
        
        
        
        # Adjust color of heatmap by adding more points for density
        fig.add_trace(
            go.Scattermapbox(
                lat=df_map["latitude"],
                lon=df_map["longitude"],
                mode="markers",
                showlegend=False,
                hoverinfo="skip",
                marker={
                    "color": df_map['Electrolyzers'],
                    "size": df_map['Electrolyzers'].fillna(0).infer_objects(copy=False),
                    "coloraxis": "coloraxis",
                    # desired max size is 15. see https://plotly.com/python/bubble-maps/#united-states-bubble-map
                    "sizeref": (df_map['Electrolyzers'].max()) / 15 ** 2,
                    "sizemode": "area",
                },
            )
        )
        
        # Usage location
        fig.add_trace(go.Scattermapbox(
                lat=[coords_demand[0]],
                lon=[coords_demand[1]],
                mode='markers',
                marker=dict(size=10, color="Orange"),
                name="Usage Location",
            
            ))
        
        
        pio.renderers.default='browser'
        fig.show()
    
    
    #%% 17. Excel output
    df_full.to_csv("df_full_csv.csv")
    
    
    list_df_all = []
    for year in list(COMPLETE_OUTPUT[0].columns):
        counter = 0
        df_full_all = pd.DataFrame(columns=['longitude','latitude','Cost_per_kg','Total Cost in Year', 'Wind turbines', 'Solar platforms', 'Electrolyzers', 'Desalination equipment', 'Storage volume', 'Conversion devices', 'Reconversion devices', 'Transport medium', 'FPSO volume', 'Distance sea'],index=[list(range(size*size))])
        for vert in range(size):
            counter2=0
            for j in range(size):
                Cost_per_kg = COMPLETE_OUTPUT[counter].loc['Costs per kg hydrogen (euros)',year]
                df_full_all.loc[counter,'Total Cost in Year'] = COMPLETE_OUTPUT[counter].loc['Total costs per year (euros)',year]
                df_full_all.loc[counter,'Cost_per_kg'] = Cost_per_kg
                df_full_all.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
                df_full_all.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
                df_full_all.loc[counter,'Wind turbines'] = COMPLETE_OUTPUT[counter].loc['Wind turbines',year]
                df_full_all.loc[counter,'Solar platforms']= COMPLETE_OUTPUT[counter].loc['Solar platforms',year]
                df_full_all.loc[counter,'Electrolyzers'] = COMPLETE_OUTPUT[counter].loc['Electrolyzers',year]
                df_full_all.loc[counter,'Desalination equipment'] = COMPLETE_OUTPUT[counter].loc['Desalination equipment',year]
                df_full_all.loc[counter,'Storage volume'] = COMPLETE_OUTPUT[counter].loc['Storage volume (m3)',year]
                df_full_all.loc[counter,'Conversion devices'] = COMPLETE_OUTPUT[counter].loc['Conversion devices',year]
                df_full_all.loc[counter,'Reconversion devices'] = COMPLETE_OUTPUT[counter].loc['Reconversion devices',year]
                df_full_all.loc[counter,'Transport medium'] = COMPLETE_OUTPUT[counter].loc['Transport medium',year]
                df_full_all.loc[counter,'FPSO volume'] = COMPLETE_OUTPUT[counter].loc['FPSO volume (m3)',year]
                df_full_all.loc[counter,'Distance sea'] = COMPLETE_OUTPUT[counter].loc['Distance sea (km)',year]
                df_full_all.loc[counter,'Demand [tonH2/yr]'] = COMPLETE_OUTPUT[counter].loc['Demand (tons of hydrogen)',year]
                counter2 = counter2 + 1     
                counter = counter + 1  
                
        # df_full_all.to_csv("Model Output/Output_pipeline_"+str(year)+".csv")
        list_df_all.append(df_full_all)
    
    # create a excel writer object
    with pd.ExcelWriter("Model Output/"+sensitivity_file_path+"Output_j"+str(E+1)+".xlsx") as writer:
       
        # use to_excel function and specify the sheet_name and index 
        # to store the dataframe in specified sheet
        counter = 0
        for year in list(COMPLETE_OUTPUT[0].columns):
            list_df_all[counter].to_excel(writer, sheet_name=str(year), index=False)
            counter += 1
    
   
    
    
    #%% 18. LCOH
    LCOH_df = pd.DataFrame(columns=['longitude','latitude','LCOH'],index=[list(range(size))])
    
    npvprod_list = []
    npvcost_list = []
    prodh2 = demand*1000
    # counter = 0
    counter3 = 0
    for vert in range(size):
        counter2=0
        for j in range(size):
            for l in L:
                year_prod = l*timestep
                dr = 1/((1+DiscountRate)**year_prod)
                Costperkg_ = COMPLETE_OUTPUT[counter3].loc['Costs per kg hydrogen (euros)',l*timestep+2020]
                # print(COMPLETE_OUTPUT[counter3].loc['Costs per kg hydrogen (euros)',l*timestep+2020])
                TotalYearlyCost = Costperkg_*prodh2
                npvcost_list.append(dr*TotalYearlyCost)
                npvprod_list.append(dr*prodh2)
                # print(counter3)
                # print(l*timestep+2020)
                # print(Costperkg_)
                
            npvcost = sum(npvcost_list)
            npvcost_list = []
            npvprod = sum(npvprod_list)        
            npvprod_list = []
    
            LCOH = npvcost / npvprod
            LCOH_df.loc[counter3,'LCOH'] = LCOH
            LCOH_df.loc[counter3,'longitude'] = loc_matrix[vert][counter2][0]
            LCOH_df.loc[counter3,'latitude'] = loc_matrix[vert][counter2][1]
            counter+=1
            counter2+=1
            counter3+=1
    
    min_index = LCOH_df['LCOH'].idxmin()
    optimal_location_df = LCOH_df.iloc[min_index[0]]
    optimal_location = "("+str(optimal_location_df['latitude'])+", "+str(optimal_location_df['longitude'])+")"
    
    LCOH_df['Optimal'] = (LCOH_df['latitude'] == optimal_location_df['latitude']) & (LCOH_df['longitude'] == optimal_location_df['longitude'])
    
    
    # create a excel writer object
    with pd.ExcelWriter("Model Output/"+sensitivity_file_path+"LCOH_j"+str(E+1)+".xlsx") as writer:
       
        # use to_excel function and specify the sheet_name and index 
        # to store the dataframe in specified sheet
        
        LCOH_df.to_excel(writer, sheet_name='j='+str(E+1), index=False)
    
    #%% 19. Cost distribution
    
    CostDistribution_df = pd.DataFrame(columns=['year','longitude','latitude','Wind','Solar','Elec','Desal','Conv','Storage','FPSO','Transport','Total Cost in Year'],index=[list(range(size*size))])
    counter = 0 # counter of location and year, size*size*Nsteps
    counter3 = 0 # location number (if size=6, 0-35)
    for vert in range(size):
        counter2=0  #0-5 if size=6, horizontal location index
        for j in range(size):
            
            for l in L:
               
                CostDistribution_df.loc[counter,'year'] = str(l*timestep+2020)
                CostDistribution_df.loc[counter,'Wind'] = CostDistribution_List[counter][0]
                CostDistribution_df.loc[counter,'Solar'] = CostDistribution_List[counter][1]
                CostDistribution_df.loc[counter,'Elec'] = CostDistribution_List[counter][2]
                CostDistribution_df.loc[counter,'Desal'] = CostDistribution_List[counter][3]
                CostDistribution_df.loc[counter,'Conv'] = CostDistribution_List[counter][4]
                CostDistribution_df.loc[counter,'Storage'] = CostDistribution_List[counter][5]
                CostDistribution_df.loc[counter,'FPSO'] = CostDistribution_List[counter][6]
                CostDistribution_df.loc[counter,'Transport'] = CostDistribution_List[counter][7]
                
                CostDistribution_df.loc[counter,'Total Cost in Year'] = COMPLETE_OUTPUT[counter3].loc['Costs per kg hydrogen (euros)',l*timestep+2020]*prodh2
                
                CostDistribution_df.loc[counter,'longitude'] = loc_matrix[vert][counter2][0]
                CostDistribution_df.loc[counter,'latitude'] = loc_matrix[vert][counter2][1]
                counter+=1
                # print(counter)
            counter2+=1
            counter3+=1
            
            # print(counter2)
    
    
    columns_to_normalize = ['Wind', 'Solar', 'Elec', 'Desal', 'Conv', 'Storage', 'FPSO', 'Transport']
    
    # Create a new DataFrame with normalized values (divide each by 'Total Cost in Year')
    CostDistributionPercent_df = CostDistribution_df.copy()
    
    # Perform the normalization
    CostDistributionPercent_df[columns_to_normalize] = CostDistribution_df[columns_to_normalize].div(CostDistribution_df['Total Cost in Year'], axis=0)
    
    
    # create a excel writer object
    with pd.ExcelWriter("Model Output/"+sensitivity_file_path+"CostDistribution_j"+str(E+1)+".xlsx") as writer:
       
        # use to_excel function and specify the sheet_name and index 
        # to store the dataframe in specified sheet
        
        CostDistribution_df.to_excel(writer, sheet_name='j='+str(E+1), index=False)
        CostDistributionPercent_df.to_excel(writer, sheet_name="Percent", index=False)