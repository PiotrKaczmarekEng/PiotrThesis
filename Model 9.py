# -*- coding: utf-8 -*-
"""
Created on Tue Nov 28 12:13:58 2023

@author: tmell
"""

# -*- coding: utf-8 -*-
"""
Created on Fri Oct 13 11:11:14 2023

@author: tmell
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import geopy.distance
import math
#import openpyxl
from openpyxl import load_workbook
#from mip import *
from feedinlib import era5
#import cdsapi
#import matplotlib.pyplot as plt
import pvlib
from feedinlib import Photovoltaic
from feedinlib import get_power_plant_data
#from shapely.geometry import Point
#from pvlib import pvsystem
#from pvlib import spectrum
#import pandas as pd
#from windpowerlib.wind_turbine import load_turbine_data_from_oedb
#from windpowerlib import WindTurbine
#import os
from windpowerlib.modelchain import ModelChain
from windpowerlib.wind_turbine import WindTurbine
from windpowerlib import get_turbine_types
from datetime import timedelta
import gurobipy as gp
#from gurobipy import GRB
#import sys




#loading excel to later retrieve input data using openpyxl
data_file = os.getcwd() 
data_file = os.path.dirname(os.path.realpath('__file__')) + '\Inputdata econ.xlsx'
# data_file = 'C:/Users/tmell/Documents/TU Delft/Master/Afstuderen/Model/Inputdata econ.xlsx' #importing data from input data excel
wb = load_workbook(data_file,data_only=True) # creating workbook

general = wb['General'] #selecting data from sheet called general
wind = wb['Wind'] #selecting data from sheet called wind
solar = wb['Solar'] #selecting data from sheet called solar
electrolyzer = wb['Electrolyzer'] #selecting data from sheet called electrolyzer
desalination = wb['Desalination'] #selecting data from sheet called desalination
ammonia = wb['Ammonia'] #selecting data from sheet called ammonia
liquidhydrogen = wb['Liquid hydrogen'] #selecting data from sheet called liquid hydrogen
landtransport = wb['Land transport'] #selecting data from sheet called landtransport
storage = wb['Storage'] #selecting data from sheet called storage
fpso = wb['FPSO'] #selecting data from sheet called fpso

#creating the model
m = gp.Model() #Model(sense=MINIMIZE,solver_name=general.cell(row=4, column=2).value) #solver_name=GRB for gurobi, solver_name=CBC for 'standard' CBC

#time period data
Startyear = general.cell(row=9, column=2).value #first year to be reviewed
timeperiod = general.cell(row=10, column=2).value #time period of simulation in years
timestep = general.cell(row=11, column=2).value #time step of simulation in years
Nsteps = int(timeperiod/timestep+1) #number of time steps (important for data selection from excel and loop at the end)

#scenario data
demand = 10/12*general.cell(row=21, column=2).value #demand in tons of hydrogen per year
demandlocation = general.cell(row=17, column=2).value #location where hydrogen is asked
transferport = general.cell(row=16, column=2).value #port where hydrogen is transferred from sea to land transport
productionlocation = general.cell(row=15, column=2).value #location where hydrogen is produced
reconversion = general.cell(row=23, column=2).value #equal to 1 if reconversion has to be done for delivery and to 0 when it does not 

#capacity data
capconvammonia = 10/12*ammonia.cell(row=10, column=2).value #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for ammonia
capconvliquid = 10/12*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per hour after conversion of one conversion installation for liquid hydrogen
capreconvammonia = 10/12*ammonia.cell(row=75, column=2).value #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for ammonia
capreconvliquid = 10/12*liquidhydrogen.cell(row=25, column=2).value  #yearly output capacity in tons of hydrogen per year after reconversion of one reconversion installation for liquid hydrogen
capdesalinationhour = desalination.cell(row=14, column=2).value #hourly output capacity desalination in m3 of water per hour
capelectrolyzerhour = electrolyzer.cell(row=9, column=2).value #hourly output capacity of electrolyzers in tons of hydrogen per hour


#efficiency data 
fracpowerelectrolyzerliquid = electrolyzer.cell(row=11, column=2).value #fraction of energy used by electrolyzers when using liquid hydrogen
fracpowerelectrolyzerammonia = electrolyzer.cell(row=12, column=2).value #fraction of energy used by elecyrolyzers when using ammonia
fracpowerelectrolyzer = m.addVar(vtype=gp.GRB.CONTINUOUS, name='fracpowerelectrolyzer')  #fraction of energy used by elecyrolyzers
eta_conversionammonia = ammonia.cell(row=13, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
eta_conversionliquid =  liquidhydrogen.cell(row=16, column=2).value #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the conversion process and the amount of hydrogen coming out
eta_reconversionammonia = eta_conversionammonia #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
eta_reconversionliquid = eta_conversionliquid #PLEAS NOTE THAT THESE ARE ACTUALLY EQUAL TO 1/ETA FOR PROGRAMMING PURPOSES conversion efficiency ratio between amount of hydrogen going into the reconversion process and the amount of hydrogen coming out
eta_transportammonia = 1 #transport efficiency (to take into account hydrogen losses during transport)
eta_transportliquid = 1 #transport efficiency (to take into account hydrogen losses during transport)
electrolyzer_water = electrolyzer.cell(row=6, column=2).value #water requirement of electrolyzer in m3 of water per ton of hydrogen
electrolyzer_energy = 1000*electrolyzer.cell(row=7, column=2).value #energy requirement of electrolyzer in Wh per ton of hydrogen

eta_conversion = m.addVar(vtype=gp.GRB.CONTINUOUS, name='eta_conversion')
eta_reconversion = m.addVar(vtype=gp.GRB.CONTINUOUS, name='eta_reconversion')

#distance calculation ('as the crow flies') and distance variables
coords_production = (general.cell(row=15, column=4).value, general.cell(row=15, column=5).value) #coordinates (latitude and longitude) of production location
coords_port = (general.cell(row=16, column=4).value, general.cell(row=16, column=5).value) #coordinates (latitude and longitude) of port
coords_demand = (general.cell(row=17, column=4).value, general.cell(row=17, column=5).value)  #coordinates (latitude and longitude) of demand location

distanceseafactor = general.cell(row=25, column=2).value
distancesea = distanceseafactor*geopy.distance.geodesic(coords_production, coords_port).km #distance to be travelled over sea from production to demand location in km
distanceland = geopy.distance.geodesic(coords_port, coords_demand).km #distance to be travelled over land from production to demand location in km

Xtransport = demand #amount to be transported by ship in tons of hydrogen, assumed to be equal to demand as hydrogen losses are almost zero 
Xkmsea = Xtransport*distancesea #yearly amount of tonkm to be made over sea
Xkmland = demand*distanceland #yealy amount of tonkm over land

#two binary variables to decide the transport mode over sea
Tammonia = m.addVar(vtype=gp.GRB.BINARY, name='Tammonia') #if equal to 1, overseas transport is done with ammonia
Tliquid = m.addVar(vtype=gp.GRB.BINARY, name='Tliquid') #if equal to 1, overseas transport is done with liquid hydrogen


#wind variables
Ntypewind = 1#wind.cell(row=4, column=2).value #number of wind turbine types included in the analysis


Xw1 = m.addVar(vtype=gp.GRB.INTEGER, name='Xwind1') #number of wind turbines of type 1
Xw2 = m.addVar(vtype=gp.GRB.INTEGER, name='Xwind2') #number of wind turbines of type 2
Xw3 = m.addVar(vtype=gp.GRB.INTEGER, name='Xwind3') #number of wind turbines of type 3
Xw4 = m.addVar(vtype=gp.GRB.INTEGER, name='Xwind4') #number of wind turbines of type 4
Xw5 = m.addVar(vtype=gp.GRB.INTEGER, name='Xwind5') #number of wind turbines of type 5

Cw1 = 10/12*np.array([float(cell.value) for cell in wind[48][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 1 floating wind turbine in 10^3 euros over several years
#Cw2 = 10/12*np.array([float(cell.value) for cell in wind[52][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 2 floating wind turbine in 10^3 euros over several years
#Cw3 = 10/12*np.array([float(cell.value) for cell in wind[56][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 3 floating wind turbine in 10^3 euros over several years
#Cw4 = 10/12*np.array([float(cell.value) for cell in wind[60][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 4 floating wind turbine in 10^3 euros over several years
#Cw5 = 10/12*np.array([float(cell.value) for cell in wind[64][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 5 floating wind turbine in 10^3 euros over several years


if Ntypewind == 1: #including variables and cost vectors for the amount of wind turbine types
    x = np.array([Xw1])
    C = np.array([Cw1])
elif Ntypewind == 2:
    x = np.array([Xw1,Xw2])
    C = np.array([Cw1,Cw2])
elif Ntypewind == 3:
    x = np.array([Xw1,Xw2,Xw3])
    C = np.array([Cw1,Cw2,Cw3])
elif Ntypewind == 4:
    x = np.array([Xw1,Xw2,Xw3,Xw4])
    C = np.array([Cw1,Cw2,Cw3,Cw4])
elif Ntypewind == 5:
    x = np.array([Xw1,Xw2,Xw3,Xw4,Xw5])
    C = np.array([Cw1,Cw2,Cw3,Cw4,Cw5])
    
#solar variables
Ntypesolar = 1 #solar.cell(row=4, column=2).value #number of solar platform types included in the analysis

Xs1 = m.addVar(vtype=gp.GRB.INTEGER, name='Xsolar1') #number of solar panel platforms of type 1
Xs2 = m.addVar(vtype=gp.GRB.INTEGER, name='Xsolar2') #number of solar panel platforms of type 2
Xs3 = m.addVar(vtype=gp.GRB.INTEGER, name='Xsolar3') #number of solar panel platforms of type 3
Xs4 = m.addVar(vtype=gp.GRB.INTEGER, name='Xsolar4') #number of solar panel platforms of type 4
Xs5 = m.addVar(vtype=gp.GRB.INTEGER, name='Xsolar5') #number of solar panel platforms of type 5
   
Cs1 = 10/12*np.array([float(cell.value) for cell in solar[51][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 1 floating solar platform in 10^3 euros over several years
# Cs2 = 10/12*np.array([float(cell.value) for cell in solar[60][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 2 floating solar platform in 10^3 euros over several years
# Cs3 = 10/12*np.array([float(cell.value) for cell in solar[64][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 3 floating solar platform in 10^3 euros over several years
# Cs4 = 10/12*np.array([float(cell.value) for cell in solar[68][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 4 floating solar platform in 10^3 euros over several years
# Cs5 = 10/12*np.array([float(cell.value) for cell in solar[72][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a type 5 floating solar platform in 10^3 euros over several years

if Ntypesolar == 1: #including variables and cost vectors for the amount of solar platform types
    x = np.append(x,[Xs1])
    C = np.vstack([C,Cs1])
elif Ntypesolar == 2:
    x = np.append(x,[Xs1,Xs2])
    C = np.vstack([C,Cs1,Cs2])
elif Ntypesolar == 3:
    x = np.append(x,[Xs1,Xs2,Xs3])
    C = np.vstack([C,Cs1,Cs2,Cs3])
elif Ntypesolar == 4:
    x = np.append(x,[Xs1,Xs2,Xs3,Xs4])
    C = np.vstack([C,Cs1,Cs2,Cs3,Cs4])
elif Ntypesolar == 5:
    x = np.append(x,[Xs1,Xs2,Xs3,Xs4,Xs5])
    C = np.vstack([C,Cs1,Cs2,Cs3,Cs4,Cs5]) 



#production equipment
convcapfactor = general.cell(row=27, column=2).value #conversion capacity factor; used to increase the conversion capacity with respect to the capacity needed for constant production over the year. Determined based on hourly production of electrolyzers and conversion equipment, and the results of some initial simulations. Couldn't determine the amount of conversion devices with a constraint like we did it with the desalination devices, because the model indicated it caused numerical problems and it started solving with a heuristic method, which did not lead to the optimal solutions (just a feasible solution)

Xe =  m.addVar(vtype=gp.GRB.INTEGER, name='Xelectrolyzers') #number of electrolyzers
Xd =  m.addVar(vtype=gp.GRB.INTEGER, name='Xdesalination') #number of desalination equipment
Xconvammonia = math.ceil(convcapfactor*(demand*(eta_reconversionammonia*eta_transportammonia))/capconvammonia) #number of conversion installations to ammonia. we nemen steady productie aan over het jaar net als in A040
Xconvliquid = math.ceil(convcapfactor*(demand*(eta_reconversionliquid*eta_transportliquid))/capconvliquid) #number of conversion installations to liquid hydrogen
Xreconvammonia = math.ceil(reconversion*demand/capreconvammonia) #number of conversion installations to ammonia
Xreconvliquid = math.ceil(reconversion*demand/capreconvliquid) #number of conversion installations to liquid hydrogen

#fpso dimension ratios and electrolyzer area (to determine fpso size), and fpso and storage variables
ratiostoragefpsoliquid = fpso.cell(row=18, column=2).value #ratio storage tanks in m3/fpso volume in m3
ratiostoragefpsoammonia = fpso.cell(row=19, column=2).value #ratio storage tanks in m3/fpso volume in m3
volumefpsoliquid = electrolyzer.cell(row=15, column=2).value #FPSO volume per electrolyzer liquid hydrogen
volumefpsoammonia = electrolyzer.cell(row=16, column=2).value #FPSO volume per electrolyzer ammonia


Xst =  m.addVar(vtype=gp.GRB.CONTINUOUS, name='Xstorage') #storage capacity in m3
Xfpso = m.addVar(name='Xfpso', vtype=gp.GRB.CONTINUOUS, lb=0) #volume of FPSO in m3


ratiostoragefpso =  m.addVar(vtype=gp.GRB.CONTINUOUS, name='ratiostoragefpso')
volumefpso =  m.addVar(vtype=gp.GRB.CONTINUOUS, name='volumefpso')

#creating the variable vector
x = np.append(x,[Xe,Xd,Xst,Xfpso,Xkmsea,Xtransport,Xkmland,Tammonia*Xconvammonia,Tliquid*Xconvliquid,Tammonia*Xreconvammonia,Tliquid*Xreconvliquid]) #creating quantity vector

#
#costs input data
#

#equipment costs
Ce = 10/12*np.array([float(cell.value) for cell in electrolyzer[50][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an electrolyzer in 10^3 euros over several years
Cd = 10/12*np.array([float(cell.value) for cell in desalination[49][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a desalination installation in 10^3 euros over several years
Cconvammonia = 10/12*np.array([float(cell.value) for cell in ammonia[48][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
Cconvliquid =10/12*np.array([float(cell.value) for cell in liquidhydrogen[61][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years
Creconvammonia = 10/12*np.array([float(cell.value) for cell in ammonia[111][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of an ammonia conversion installation in 10^3 euros over several years
Creconvliquid = 10/12*np.array([float(cell.value) for cell in liquidhydrogen[125][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of a liquid hydrogen conversion installation in 10^3 euros over several years

#transport and storage costs (including fpso)
Cstliquid =  10/12*np.array([float(cell.value) for cell in storage[25][2:2+Nsteps]]) #cost per year (depreciation+OPEX) of storage per m3 in euros over several years
Cstammonia =  10/12*np.array([float(cell.value) for cell in storage[54][2:2+Nsteps]])
Cfpso = 10/12* np.array([float(cell.value) for cell in fpso[53][2:2+Nsteps]]) #np.array([1,1,0.8,0.6,0.4]) #cost per year (depreciation+OPEX) of FPSO per m3 in 10^3 euros over several years
Ckmammonia = np.array([float(cell.value) for cell in ammonia[158][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as ammonia in 10^3 euros over several years
Cbasetransportammonia = np.array([float(cell.value) for cell in ammonia[159][2:2+Nsteps]])
Ckmliquid = np.array([float(cell.value) for cell in liquidhydrogen[168][2:2+Nsteps]]) #costs per km of overseas transport of 1 ton of hydrogen as liquid hydrogen in 10^3 euros over several years
Cbasetransportliquid =  np.array([float(cell.value) for cell in liquidhydrogen[169][2:2+Nsteps]])
Ckmland = np.array([float(cell.value) for cell in landtransport[33][2:2+Nsteps]]) #costs per year per km of overland pipeline for 1 ton of hydrogen

Cst = Tammonia * Cstammonia + Tliquid * Cstliquid

Ckmsea = Ckmammonia * Tammonia + Ckmliquid * Tliquid #costs per km of overseas transport, depending on whether ammonia or liquid hydrogen is chosen
Cbasetransport = Cbasetransportammonia * Tammonia + Cbasetransportliquid * Tliquid #baserate of the transport per ton hydrogen

#creating costs vector
C = np.vstack([C,Ce,Cd,Cst,Cfpso,Ckmsea,Cbasetransport,Ckmland,Cconvammonia,Cconvliquid,Creconvammonia,Creconvliquid]) #creating costs matrix with the costs of every aspect over the years

#
#wind and solar energy generation calculations based on ERA5 weather data, pvlib, windpowerlib and feedinlib
#

#
#general
#

#this script works with version 0.1.0rc4 of feedinlib, version 0.9.5 of pvlib and version 0.2.1 of windpowerlib

#coordinates production location
latitude =  general.cell(row=15, column=4).value
longitude = general.cell(row=15, column=5).value


# set start and end date (end date will be included
# in the time period for which data is downloaded)
start_date, end_date = '2020-01-01', '2020-10-31'
# set variable set to download
variable = 'feedinlib'

era5_netcdf_filename = 'ERA5_weather_data_location4.nc' #referring to file with weather data downloaded earlier using the ERA5 API

area = [longitude, latitude] #location of production

#
#PV
#

# get modules
module_df = get_power_plant_data(dataset='SandiaMod') #retrieving dataset for PV modules
# print the first four modules
#module_df.iloc[:, 1:5]

# get inverter data
inverter_df = get_power_plant_data(dataset='cecinverter') #retrieving dataset for inverters
# print the first four inverters
#inverter_df.iloc[:, 1:5]

temp_params = pvlib.temperature.TEMPERATURE_MODEL_PARAMETERS['sapm']['open_rack_glass_polymer'] #defining temperature model parameters (leaving at default causes errors)

#PV system definition
system_data = {
    'module_name': 'Advent_Solar_Ventura_210___2008_',  # module name as in database
    'inverter_name': 'ABB__MICRO_0_25_I_OUTD_US_208__208V_',  # inverter name as in database
    'azimuth': 180, #angle of sun position with north in horizontal plane
    'tilt': 30, #angle of solar panels with horizontal
    'albedo': 0.075, #albedo (fraction of sun light reflected) of ocean water (https://geoengineering.global/ocean-albedo-modification/)
    'temperature_model_parameters': temp_params,
}


pv_system = Photovoltaic(**system_data)


#getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
pvlib_df = era5.weather_df_from_era5(     
    era5_netcdf_filename='ERA5_weather_data_location4.nc',
    lib='pvlib', area=area)


#determining the zenith angle (angle of sun position with vertical in vertical plane) in the specified locations for the time instances downloaded from ERA5
zenithcalc = pvlib.solarposition.get_solarposition(time=pvlib_df.index,latitude=latitude, longitude=longitude, altitude=None, pressure=None, method='nrel_numpy', temperature=pvlib_df['temp_air'])

#determining DNI from GHI, DHI and zenith angle for the time instances downloaded from ERA5
dni = pvlib.irradiance.dni(pvlib_df['ghi'],pvlib_df['dhi'], zenith=zenithcalc['zenith'], clearsky_dni=None, clearsky_tolerance=1.1, zenith_threshold_for_zero_dni=88.0, zenith_threshold_for_clearsky_limit=80.0)

#adding DNI to dataframe with PV weather data
pvlib_df['dni'] = dni

#replacing 'NAN' in DNI column with 0 to prevent 'holes' in graphs (NANs are caused by the zenith angle being larger than the 'zenith threshold for zero dni' angle set with GHI and DHI not yet being zero. The DNI should be 0 in that case)
pvlib_df['dni'] = pvlib_df['dni'].fillna(0)

#plotting dhi and ghi
plt.figure
pvlib_df.loc[:, ['dhi', 'ghi']].plot(title='Irradiance')
plt.xlabel('Time')
plt.ylabel('Irradiance in $W/m^2$');

#determining PV power generation
feedin = pv_system.feedin(
    weather=pvlib_df,
    location=(latitude, longitude))


#plotting PV power generation
plt.figure()
feedin.plot(title='PV feed-in')
plt.xlabel('Time')
plt.ylabel('Power in W');

#calculating total PV power generated in selected period
PVpower = pd.Series.sum(feedin) #power produced over time period in Wh

print('One panel with the chosen input parameters in location', latitude,',',longitude, 'will produce',PVpower,'Wh of electricity between', start_date, 'and', end_date )

#
#Wind
#


# get power curves
# get names of wind turbines for which power curves and/or are provided
# set print_out=True to see the list of all available wind turbines
df = get_turbine_types(print_out=False)

#defining the wind turbine system
turbine_data= {
    "turbine_type": "E-101/3050",  # turbine type as in register
    "hub_height": 130,  # in m
}
my_turbine = WindTurbine(**turbine_data)

#getting the needed weather data for PV calculations from the file as downloaded with a seperate script from ERA5
windpowerlib_df = era5.weather_df_from_era5(
    era5_netcdf_filename='ERA5_weather_data_location4.nc',
    lib='windpowerlib', area=area)

#increasing the time indices by half an hour because then they match the pvlib time indices so the produced energy can be added later
#assumed wind speeds half an hour later are similar and this will not affect the results significantly
windpowerlib_df.index = windpowerlib_df.index + timedelta(minutes=30)

# power output calculation for e126

# own specifications for ModelChain setup
modelchain_data = {
    'wind_speed_model': 'logarithmic',      # 'logarithmic' (default),
                                            # 'hellman' or
                                            # 'interpolation_extrapolation'
    'density_model': 'ideal_gas',           # 'barometric' (default), 'ideal_gas'
                                            #  or 'interpolation_extrapolation'
    'temperature_model': 'linear_gradient', # 'linear_gradient' (def.) or
                                            # 'interpolation_extrapolation'
    'power_output_model': 'power_curve',    # 'power_curve' (default) or
                                            # 'power_coefficient_curve'
    'density_correction': True,             # False (default) or True
    'obstacle_height': 0,                   # default: 0
    'hellman_exp': None}                    # None (default) or None

# initialize ModelChain with own specifications and use run_model method to
# calculate power output
mc_my_turbine = ModelChain(my_turbine, **modelchain_data).run_model(windpowerlib_df)
# write power output time series to WindTurbine object
my_turbine.power_output = mc_my_turbine.power_output


# plot turbine power output
plt.figure()
my_turbine.power_output.plot(title='Wind turbine power production')
plt.xlabel('Time')
plt.ylabel('Power in W')
plt.show()

#
#Total power and hydrogen production
#

multsolar = solar.cell(row=16, column=2).value*1000/206.7989
multwind = wind.cell(row=7, column=2).value/3

feedinarray = multsolar*feedin.values #hourly energy production over the year of 1 solar platform of the specified kind in the specified location
feedinarray[feedinarray<0]=0
my_turbinearray = multwind*my_turbine.power_output.values #hourly energy production over the year of 1 wind turbine of the specified kind in the specified location

powerovertime = m.addVars(len(feedinarray),vtype=gp.GRB.CONTINUOUS,name='powerovertime') #hourly energy production from all wind turbines and solar platforms combined
powerovertimeelectrolyzer = m.addVars(len(feedinarray),vtype=gp.GRB.CONTINUOUS,name='powerovertimeelectrolyzer') #hourly energy available for the elctrolyzers
electricityusedelectrolyzer = m.addVars(len(feedinarray),vtype=gp.GRB.CONTINUOUS,name='electricityudeselectrolyzer') #hourly energy actually used by the electrolyzers (limited by their capacity and the energy availability)
hydrogenproduced = m.addVars(len(feedinarray),vtype=gp.GRB.CONTINUOUS,name='hydrogenproduced') #hourly hydrogen production

#
#Constraints
#

m.params.NonConvex = 2 #allowing quadratic constraints in Gurobi
m.Params.NumericFocus=1

print('Optimalisatie start!!!!')

#the transport over sea is done completely with either ammonia or liquid hydrogen
m.addConstr(Tammonia + Tliquid == 1) 


for i in range(len(electricityusedelectrolyzer)):
    #produced power depends on the amounts of wind turbines and solar platforms and the production per unit
    m.addConstr(powerovertime[i] == feedinarray[i]*Xs1 + my_turbinearray[i]*Xw1)
    #fraction of the power used by the electrolyzers depends on whether ammonia or liquid hydrogen is used for storage and transport
    m.addConstr(fracpowerelectrolyzer == Tammonia*fracpowerelectrolyzerammonia + Tliquid*fracpowerelectrolyzerliquid)
    #then the power available for the electrolyzers can be determined
    m.addConstr(powerovertimeelectrolyzer[i] == powerovertime[i]*fracpowerelectrolyzer) 
    #the power actually used by the electrolyzers must be lower than or equal to the available power for the electrolyzers
    m.addConstr(electricityusedelectrolyzer[i] <= powerovertimeelectrolyzer[i])
    #the power actually used by the electrolyzers must also be lower than or equal to the capacity of the electrolyzers
    m.addConstr(electricityusedelectrolyzer[i] <= capelectrolyzerhour*Xe*electrolyzer_energy)
    #from the power actually used by the electrolyzers, the produced hydrogen can be determined
    m.addConstr(hydrogenproduced[i] == electricityusedelectrolyzer[i]/electrolyzer_energy)

m.addConstr(eta_conversion==Tammonia*eta_conversionammonia+Tliquid*eta_conversionliquid)


m.addConstr(eta_reconversion==Tammonia*eta_reconversionammonia+Tliquid*eta_reconversionliquid)

    
#the hydrogen produced in the whole period should be bigger than or equal to the demand (with correction for conversion, transport and reconversion efficiencies)
m.addConstr(hydrogenproduced.sum() >= demand*(eta_conversion*eta_conversion))

#the capacity of the desalination equipment should match the capacity of the electrolyzers to be able to produce sufficient water
m.addConstr(Xd >=  Xe*capelectrolyzerhour*electrolyzer_water/capdesalinationhour)

#the capacity of the conversion equipment should match the electrolyzer capacity
#m.addConstr(Xconvammonia >= Xe*capelectrolyzerhour/capconvammoniahour)
#m.addConstr(Xconvliquid >= Xe*capelectrolyzerhour/capconvliquidhour)

#determining FPSO volume per installed electrolyzer based on liquid hydrogen or ammonia fpso
m.addConstr(volumefpso == Tammonia*volumefpsoammonia + Tliquid*volumefpsoliquid)


#with the amount of electrolyzers and FPSO volume per electrolyzer, the total FPSO volume is determined
m.addConstr(Xfpso == Xe*volumefpso)

#determining storage as ratio of fpso volime based on liquid hydrogen or ammonia fpso (ammonia requires smaller tank to store same amount of hydrogen)
m.addConstr(ratiostoragefpso == Tammonia*ratiostoragefpsoammonia + Tliquid*ratiostoragefpsoliquid)

#based on the volume of the FPSO and the example FPSO design we can determine the storage capacity
m.addConstr(Xst == Xfpso*ratiostoragefpso)

powerovertimeelectrolyzerdata = pd.DataFrame(index=feedin.index) #for plotting the available electrolyzer power
electricityusedelectrolyzerdata = pd.DataFrame(index=feedin.index) #for plotting the used electrolyzer power
hydrogenproduceddata = pd.DataFrame(index=feedin.index) #for plotting the hydrogen production
exceldf = pd.DataFrame(index=['Demand (tons of hydrogen)', 'Usage location', 'Year','Production location','Transfer port','Total costs per year (euros)','Costs per kg hydrogen (euros)','Wind turbines', 'Solar platforms','Electrolyzers','Desalination equipment', 'Storage volume (m3)','Conversion devices','Reconversion devices','Transport medium', 'FPSO volume (m3)', 'Distance sea (km)','Distance land (km)'])
 

# #objective
# for k in range(Nsteps): #for loop allows to evaluate several years in one run   
#     m.reset()
#     m.setObjective(sum(x[i]*C[i][k] for i in range(len(x))),gp.GRB.MINIMIZE) #minimization of the total costs with the given demand (aka minimization of the costs/ton hydrogen, aka optimization of the hydrogen supply chain in the given scenario)
#     m.optimize()
#     #determine fuel
#     if Tammonia.x>0 :
#         Transportmedium ="AMMONIA"
#     else:
#         Transportmedium = "LIQUID HYDROGEN"
#     print("Optimal solution found!")
#     print("For a demand of", demand*12/10 ,"tons of hydrogen per year in", demandlocation ,"in year", timestep*k+Startyear, "with production in", productionlocation ,"and transfer in the port of", transferport, "the yearly costs are equal to", m.ObjVal*12/10,"euros in total.\n This is equal to a cost of", m.ObjVal/demand, "euros per ton of hydrogen.\n In this configuration, we will use:\n",Xw1.X, "wind turbines of type 1\n",Xw2.X, "wind turbines of type 2\n",Xw3.X, "wind turbines of type 3\n",Xw4.X, "wind turbines of type 4\n",Xw5.X, "wind turbines of type 5\n", Xs1.X, "solar platforms of type 1\n", Xs2.X, "solar platforms of type 2\n", Xs3.X, "solar platforms of type 3\n", Xs4.X, "solar platforms of type 4\n", Xs5.X, "solar platforms of type 5\n", Xe.X, "electrolyzers\n", Xd.x , "desalination installations\n", Xst.X, "m3 of storage capacity \n ",Tammonia.X*Xconvammonia + Tliquid.X*Xconvliquid , "conversion and", Tammonia.X*Xreconvammonia + Tliquid.X*Xreconvliquid , "reconversion installations for", Transportmedium ,"\n in an FPSO of", Xfpso.X, "m3") #prints the values of all variables in the optimized solution
#     #export to excel
#     exceldf[timestep*k+Startyear] = [demand*12/10,demandlocation,timestep*k+Startyear, productionlocation, transferport, m.ObjVal*12/10, m.ObjVal/demand/1000,Xw1.X, Xs1.X, Xe.X, Xd.x , Xst.X, Tammonia.X*Xconvammonia + Tliquid.X*Xconvliquid, Tammonia.X*Xreconvammonia + Tliquid.X*Xreconvliquid , Transportmedium , Xfpso.X,distancesea,distanceland]

#     #plot power available electrolyzers
#     powerovertimeelectrolyzervalues = np.empty(len(powerovertimeelectrolyzer), dtype=object)
#     for i in range(len(powerovertimeelectrolyzer)):
#         powerovertimeelectrolyzervalues[i] = powerovertimeelectrolyzer[i].x
#     powerovertimeelectrolyzerdata['production'] = powerovertimeelectrolyzervalues.tolist()
#     powerovertimeelectrolyzerdata.plot(title=' ') #power available for electrolyzers
#     plt.xlabel('Time')
#     plt.ylabel('W')
    
#     #plot power used electrolyzers
#     electricityusedelectrolyzervalues = np.empty(len(electricityusedelectrolyzer), dtype=object)
#     for i in range(len(electricityusedelectrolyzer)):
#         electricityusedelectrolyzervalues[i] = electricityusedelectrolyzer[i].x
#     electricityusedelectrolyzerdata['production'] = electricityusedelectrolyzervalues.tolist()
#     electricityusedelectrolyzerdata.plot(title=' ')
#     plt.xlabel('Time')
#     plt.ylabel('W')
#     plt.ylim([0,700000000])
    
#     #plot produced hydrogen
#     hydrogenproducedvalues = np.empty(len(hydrogenproduced), dtype=object)
#     for i in range(len(hydrogenproduced)):
#         hydrogenproducedvalues[i] = hydrogenproduced[i].x
#     hydrogenproduceddata['production'] = hydrogenproducedvalues.tolist()
#     hydrogenproduceddata.plot(title='')
#     plt.xlabel('Time')
#     plt.ylabel('Tons per hour')
   
# exceldf.to_excel('Simulation_results.xlsx', sheet_name='Results')